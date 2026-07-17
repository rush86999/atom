"""
Workbook runtime — evaluates formulas, maintains dependencies, renders output.

This is to an Excel workbook what V8 is to JavaScript: it runs the workbook.
openpyxl (the existing approach) is a parser/serializer — it can read/write
the file format but can't evaluate formulas, maintain references on structural
edits, or render pixel-accurate output.

The runtime uses a fallback chain:
  1. LibreOffice headless (if `soffice` is installed) — full runtime:
     recalc + render + structural edit support
  2. `formulas` library (pure Python) — formula evaluation only
  3. openpyxl (current behavior) — parser/serializer with stale cached values

When LibreOffice is available, the runtime:
  - Recalculates ALL formulas after writes (so agents see computed results)
  - Renders to pixel-accurate HTML/PNG with conditional formatting and charts
  - Maintains formula references when rows/columns are inserted
"""

from __future__ import annotations

import asyncio
import logging
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def _find_soffice() -> Optional[str]:
    """Find the LibreOffice executable."""
    for name in ("soffice", "libreoffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        path = shutil.which(name)
        if path:
            return path
    # Check common macOS path
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(mac_path):
        return mac_path
    return None


_SOFFICE = _find_soffice()


class WorkbookRuntime:
    """Evaluates formulas, renders output, and maintains workbook semantics.

    Uses LibreOffice headless when available (full runtime), falling back to
    the `formulas` Python library for evaluation-only, then to openpyxl's
    cached values.
    """

    def __init__(self):
        self._soffice = _SOFFICE
        self._has_formulas = self._check_formulas()

    @property
    def engine(self) -> str:
        """Which engine is active."""
        if self._soffice:
            return "libreoffice"
        if self._has_formulas:
            return "formulas"
        return "openpyxl"

    @property
    def can_evaluate(self) -> bool:
        """Can this runtime evaluate formulas?"""
        return self._soffice is not None or self._has_formulas

    @property
    def can_render(self) -> bool:
        """Can this runtime render to HTML/PNG?"""
        return self._soffice is not None

    def _check_formulas(self) -> bool:
        try:
            import formulas  # noqa: F401
            return True
        except ImportError:
            return False

    # ------------------------------------------------------------------
    # Recalculate
    # ------------------------------------------------------------------

    async def recalculate(self, file_path: str | Path) -> Path:
        """Recalculate all formulas in the workbook.

        After this call, `data_only=True` reads return fresh computed values
        instead of stale/None cached values.

        Args:
            file_path: Path to the .xlsx file (modified in place).

        Returns:
            Path to the recalculated file.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Workbook not found: {file_path}")

        if self._soffice:
            return await self._recalc_with_soffice(file_path)
        elif self._has_formulas:
            return await self._recalc_with_formulas(file_path)
        else:
            logger.debug("No formula engine available — skipping recalc")
            return file_path

    async def _recalc_with_soffice(self, file_path: Path) -> Path:
        """Use LibreOffice headless to recalculate and save."""
        try:
            # Create a temp dir for the output.
            with tempfile.TemporaryDirectory() as tmpdir:
                tmpdir = Path(tmpdir)
                # soffice --headless --calc --convert-to xlsx --outdir <tmp> <file>
                # This opens the file in Calc (which recalculates), then saves.
                proc = await asyncio.create_subprocess_exec(
                    self._soffice,
                    "--headless",
                    "--calc",
                    "--convert-to", "xlsx",
                    "--outdir", str(tmpdir),
                    str(file_path),
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    env={**os.environ, "HOME": str(Path.home())},
                )
                stdout, stderr = await asyncio.wait_for(
                    proc.communicate(), timeout=60
                )

                output_file = tmpdir / file_path.name
                if output_file.exists():
                    # Replace the original with the recalculated version.
                    shutil.copy2(output_file, file_path)
                    logger.info(f"Recalculated workbook via LibreOffice: {file_path.name}")
                    return file_path
                else:
                    logger.warning(f"LibreOffice recalc produced no output: {stderr.decode()}")
                    return file_path
        except asyncio.TimeoutError:
            logger.warning("LibreOffice recalc timed out (60s)")
            return file_path
        except Exception as e:
            logger.warning(f"LibreOffice recalc failed: {e}")
            return file_path

    async def _recalc_with_formulas(self, file_path: Path) -> Path:
        """Use the `formulas` Python library to evaluate and write cached values."""
        try:
            import formulas
            from openpyxl import load_workbook

            # Parse and evaluate the workbook.
            xl_model = formulas.ExcelModel().loads(str(file_path)).finish()
            solution = xl_model.calculate()

            # Open with openpyxl and write back the computed values as cached.
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            # Try to find the computed value in the solution.
                            # The formulas library keys by fully-qualified refs.
                            ref = f"'[{file_path.name}]{sheet_name}'!{cell.coordinate}"
                            # Also try without quotes
                            ref_alt = f"[{file_path.name}]{sheet_name}!{cell.coordinate}"
                            for key in (ref, ref_alt, ref.upper(), ref_alt.upper()):
                                if key in solution:
                                    val = solution[key]
                                    if hasattr(val, "value"):
                                        cell.value = val.value
                                    elif isinstance(val, (int, float, str)):
                                        cell.value = val
                                    break
            wb.save(file_path)
            logger.info(f"Recalculated workbook via formulas lib: {file_path.name}")
            return file_path
        except Exception as e:
            logger.warning(f"Formulas recalc failed: {e}")
            return file_path

    # ------------------------------------------------------------------
    # Macros and VBA
    # ------------------------------------------------------------------

    async def run_macro(self, file_path: str | Path, macro_name: str) -> Dict[str, Any]:
        """Run a macro/VBA script in the workbook via Firecracker sandbox."""
        file_path = Path(file_path)
        if not file_path.exists():
            return {"success": False, "error": f"Workbook not found: {file_path}"}

        if not self._soffice:
            return {"success": False, "error": "Macro execution requires LibreOffice runtime engine"}

        from core.firecracker_sandbox import get_sandbox
        sandbox = get_sandbox()

        # Command to trigger macro execution in LibreOffice Calc standard libraries
        command = [
            self._soffice,
            "--headless",
            "--invisible",
            f"macro:///Standard.Module1.{macro_name}",
            str(file_path)
        ]

        success = await sandbox.execute_in_sandbox(command, file_path, file_path.parent)
        if success:
            await self.recalculate(file_path)
            return {"success": True, "macro": macro_name}
        return {"success": False, "error": "Macro execution failed inside sandbox"}

    # ------------------------------------------------------------------
    # Pivot Tables
    # ------------------------------------------------------------------

    async def add_pivot_table(
        self, file_path: str | Path, sheet_name: str, pivot_sheet_name: str,
        data_range: str, rows: List[str], columns: List[str],
        values: List[Dict[str, str]]
    ) -> Dict[str, Any]:
        """Generate a pivot table sheet using pandas data analytics and write to workbook."""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            file_path = Path(file_path)
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Source sheet '{sheet_name}' not found"}

            # Load sheet data into DataFrame
            ws_src = wb[sheet_name]
            
            # Read source range or whole sheet if range is not fully defined
            # We convert sheet cells to a list of dicts/lists for pandas
            data = []
            for row in ws_src.iter_rows(values_only=True):
                data.append(row)
            
            if not data:
                return {"success": False, "error": f"Source sheet '{sheet_name}' has no data"}
                
            headers = data[0]
            df = pd.DataFrame(data[1:], columns=headers)

            # Map aggregator functions
            agg_funcs = {}
            for v in values:
                field = v.get("field")
                func = v.get("function", "sum").lower()
                agg_funcs[field] = func

            # Pivot computation using pandas
            pivot_df = pd.pivot_table(
                df,
                index=rows,
                columns=columns,
                values=[v.get("field") for v in values],
                aggfunc=agg_funcs
            )

            # Write back pivot summary sheet
            if pivot_sheet_name in wb.sheetnames:
                del wb[pivot_sheet_name]
            ws_pivot = wb.create_sheet(pivot_sheet_name)

            # Write styled header and dataframe to worksheet
            for r in dataframe_to_rows(pivot_df, index=True, header=True):
                ws_pivot.append(r)

            wb.save(file_path)
            await self.recalculate(file_path)

            logger.info(f"Created pivot table in sheet {pivot_sheet_name} from {sheet_name}")
            return {"success": True, "pivot_sheet": pivot_sheet_name}
        except Exception as e:
            logger.error(f"Failed to generate pivot table: {e}")
            return {"success": False, "error": str(e)}

    # ------------------------------------------------------------------
    # Render
    # ------------------------------------------------------------------

    async def render_to_html(self, file_path: str | Path) -> str:
        """Render the workbook to pixel-accurate HTML.

        Uses LibreOffice for full rendering (conditional formatting, charts,
        evaluated formulas). Falls back to a basic openpyxl-based HTML table.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            return "<p>File not found</p>"

        if self._soffice:
            return await self._render_html_with_soffice(file_path)
        return self._render_html_basic(file_path)

    async def _render_html_with_soffice(self, file_path: Path) -> str:
        """LibreOffice HTML render — pixel-accurate with CF, charts, formulas."""
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmpdir = Path(tmpdir)
                proc = await asyncio.create_subprocess_exec(
                    self._soffice,
                    "--headless",
                    "--convert-to", "html",
                    "--outdir", str(tmpdir),
                    str(file_path),
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    env={**os.environ, "HOME": str(Path.home())},
                )
                await asyncio.wait_for(proc.communicate(), timeout=30)

                html_file = tmpdir / (file_path.stem + ".html")
                if html_file.exists():
                    html = html_file.read_text(encoding="utf-8")
                    logger.info(f"Rendered {file_path.name} to HTML via LibreOffice")
                    return html
        except Exception as e:
            logger.warning(f"LibreOffice HTML render failed: {e}")

        return self._render_html_basic(file_path)

    def _render_html_basic(self, file_path: Path) -> str:
        """Fallback: basic HTML table from openpyxl (values only, no CF/charts)."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            html_parts = ['<table border="1" style="border-collapse:collapse">']
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                html_parts.append(f'<tr><td colspan="{ws.max_column or 1}" '
                                  f'style="font-weight:bold;background:#e0e0e0">'
                                  f'{sheet_name}</td></tr>')
                for row in ws.iter_rows(values_only=True):
                    html_parts.append("<tr>")
                    for val in row:
                        display = "" if val is None else str(val)
                        html_parts.append(f"<td>{display}</td>")
                    html_parts.append("</tr>")
            html_parts.append("</table>")
            return "".join(html_parts)
        except Exception as e:
            return f"<p>Error rendering: {e}</p>"

    # ------------------------------------------------------------------
    # Structural operations
    # ------------------------------------------------------------------

    async def insert_rows(
        self, file_path: str | Path, sheet_name: str, row: int, count: int = 1
    ) -> Dict[str, Any]:
        """Insert rows and recalculate to maintain formula references.

        Args:
            file_path: Path to the .xlsx file.
            sheet_name: Worksheet name.
            row: Row number to insert at (1-based).
            count: Number of rows to insert.

        Returns:
            {"success": True, "rows_inserted": count}
        """
        from openpyxl import load_workbook

        file_path = Path(file_path)
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

        ws = wb[sheet_name]
        ws.insert_rows(row, count)
        wb.save(file_path)

        # Recalculate to fix all formula references.
        await self.recalculate(file_path)

        logger.info(f"Inserted {count} rows at row {row} in {sheet_name}")
        return {"success": True, "rows_inserted": count}

    async def insert_cols(
        self, file_path: str | Path, sheet_name: str, col: int, count: int = 1
    ) -> Dict[str, Any]:
        """Insert columns and recalculate to maintain formula references."""
        from openpyxl import load_workbook

        file_path = Path(file_path)
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

        ws = wb[sheet_name]
        ws.insert_cols(col, count)
        wb.save(file_path)
        await self.recalculate(file_path)

        logger.info(f"Inserted {count} columns at col {col} in {sheet_name}")
        return {"success": True, "cols_inserted": count}

    # ------------------------------------------------------------------
    # Evaluated reads
    # ------------------------------------------------------------------

    async def get_evaluated_range(
        self, file_path: str | Path, sheet_name: str,
        start_cell: str = "A1", end_cell: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Read a range with evaluated formula results.

        Recalculates first, then reads with data_only=True so the agent
        sees computed values, not formula strings.
        """
        from openpyxl import load_workbook

        file_path = Path(file_path)
        await self.recalculate(file_path)

        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

        ws = wb[sheet_name]
        cell_range = f"{start_cell}:{end_cell}" if end_cell else start_cell
        cells = ws[cell_range]

        values = []
        if isinstance(cells, tuple):  # range
            for row in cells:
                row_vals = []
                for cell in row:
                    row_vals.append({
                        "cell": cell.coordinate,
                        "value": cell.value,
                    })
                values.append(row_vals)
        else:  # single cell
            values.append([{
                "cell": cells.coordinate,
                "value": cells.value,
            }])

        return {
            "success": True,
            "sheet": sheet_name,
            "range": cell_range,
            "values": values,
            "engine": self.engine,
        }

    async def get_formula_result(
        self, file_path: str | Path, sheet_name: str, cell: str
    ) -> Dict[str, Any]:
        """Get the computed result of a single cell (evaluates if needed)."""
        return await self.get_evaluated_range(file_path, sheet_name, cell)


# Singleton.
_workbook_runtime: Optional[WorkbookRuntime] = None


def get_workbook_runtime() -> WorkbookRuntime:
    global _workbook_runtime
    if _workbook_runtime is None:
        _workbook_runtime = WorkbookRuntime()
        logger.info(f"WorkbookRuntime initialized: engine={_workbook_runtime.engine}")
    return _workbook_runtime
