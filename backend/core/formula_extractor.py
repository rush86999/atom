"""
Formula Extractor for ATOM Platform
Extracts formulas from Excel spreadsheets and stores them in Atom's memory.
"""

import logging
from pathlib import Path
import re
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


class FormulaExtractor:
    """
    Extracts formulas from Excel files and maps them to semantic descriptions.
    Uses column headers as parameter names for better searchability.
    """

    # Common Excel formula patterns
    FORMULA_PATTERNS = {
        "SUM": r"=SUM\(([^)]+)\)",
        "AVERAGE": r"=AVERAGE\(([^)]+)\)",
        "IF": r"=IF\(([^)]+)\)",
        "VLOOKUP": r"=VLOOKUP\(([^)]+)\)",
        "COUNT": r"=COUNT\(([^)]+)\)",
        "MAX": r"=MAX\(([^)]+)\)",
        "MIN": r"=MIN\(([^)]+)\)",
        "SUMIF": r"=SUMIF\(([^)]+)\)",
        "CONCATENATE": r"=CONCATENATE\(([^)]+)\)",
    }

    # Domain keywords for classification
    DOMAIN_KEYWORDS = {
        "finance": ["revenue", "cost", "profit", "expense", "budget", "margin", "roi", "income", "loss", "tax"],
        "sales": ["sales", "quota", "target", "commission", "deal", "pipeline", "conversion", "lead"],
        "operations": ["inventory", "order", "shipment", "delivery", "capacity", "utilization", "efficiency"],
        "hr": ["salary", "headcount", "turnover", "retention", "attendance", "fte", "employee"],
        "marketing": ["cac", "ltv", "engagement", "reach", "impression", "click", "conversion", "campaign"]
    }

    def __init__(self, workspace_id: str = "default"):
        self.workspace_id = workspace_id
        self._formula_manager = None

    # Supported file extensions for formula extraction
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".csv", ".ods", ".gsheet", ".numbers"]

    def _get_formula_manager(self):
        """Lazy load formula manager."""
        if self._formula_manager is None:
            from core.formula_memory import get_formula_manager
            self._formula_manager = get_formula_manager(self.workspace_id)
        return self._formula_manager

    def extract_from_file(
        self,
        file_path: str,
        user_id: str = "default_user",
        auto_store: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Extract formulas from any Excel-like file.
        Supported formats: .xlsx, .xls, .csv, .ods, .gsheet, .numbers

        Args:
            file_path: Path to the file
            user_id: Owner of the extracted formulas
            auto_store: If True, automatically store in Atom's memory

        Returns:
            List of extracted formula definitions
        """
        ext = Path(file_path).suffix.lower()
        logger.info(f"Extracting formulas from {file_path} (type: {ext})")

        if ext in [".xlsx"]:
            return self.extract_from_excel(file_path, user_id, auto_store)
        elif ext in [".xls"]:
            return self.extract_from_xls(file_path, user_id, auto_store)
        elif ext in [".csv"]:
            return self.extract_from_csv(file_path, user_id, auto_store)
        elif ext in [".ods"]:
            return self.extract_from_ods(file_path, user_id, auto_store)
        elif ext in [".gsheet", ".numbers"]:
            # Google Sheets and Numbers export as xlsx/csv, handle if possible
            logger.info(f"{ext} file detected, attempting xlsx extraction")
            return self.extract_from_excel(file_path, user_id, auto_store)
        else:
            logger.warning(f"Unsupported file type for formula extraction: {ext}")
            return []

    def extract_from_excel(
        self,
        file_path: str,
        user_id: str = "default_user",
        auto_store: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Extract all formulas from an Excel file.

        Args:
            file_path: Path to the Excel file
            user_id: Owner of the extracted formulas
            auto_store: If True, automatically store in Atom's memory

        Returns:
            List of extracted formula definitions
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return []

        # Raw-XML scan FIRST: it reads every <f> element straight from the
        # package and works for openpyxl-hostile workbooks (Zoho Sheet
        # exports trip openpyxl's strict reader — either raising outright or,
        # worse, "succeeding" degenerately with shared strings half-parsed).
        try:
            raw_extracted = self._extract_formulas_raw_xml(file_path)
            if raw_extracted:
                if auto_store:
                    self._store_formulas(raw_extracted, user_id, file_path)
                logger.info(
                    f"Extracted {len(raw_extracted)} formulas from {file_path} "
                    f"(raw-XML)"
                )
                return raw_extracted
        except Exception as raw_err:
            logger.debug(f"raw-XML formula scan failed for {file_path}: {raw_err}")

        extracted = []

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_formulas = self._extract_from_sheet(sheet, sheet_name)
                extracted.extend(sheet_formulas)

            workbook.close()

            if auto_store and extracted:
                self._store_formulas(extracted, user_id, file_path)

            logger.info(f"Extracted {len(extracted)} formulas from {file_path}")
            return extracted

        except Exception as e:
            # openpyxl is strict about workbook XML and hard-fails on
            # spreadsheets written by non-Excel tools (Zoho Sheet exports trip
            # read_strings) — the same files that broke content ingestion
            # before the raw-XML parser existed. Their formulas are plain
            # well-formed XML: extract them directly instead of returning [].
            logger.warning(
                f"openpyxl cannot read {file_path} ({e}) — trying raw-XML formula fallback"
            )
            try:
                extracted = self._extract_formulas_raw_xml(file_path)
                if auto_store and extracted:
                    self._store_formulas(extracted, user_id, file_path)
                if extracted:
                    logger.info(
                        f"Extracted {len(extracted)} formulas from {file_path} "
                        f"(raw-XML fallback)"
                    )
                return extracted
            except Exception as raw_err:
                logger.error(f"Raw-XML formula fallback failed for {file_path}: {raw_err}")
                return []

    def _extract_formulas_raw_xml(
        self,
        file_path: str,
        max_sheets: int = 10,
        max_formulas: int = 200,
    ) -> List[Dict[str, Any]]:
        """Extract formulas straight from the xlsx zip (stdlib XML only).

        Fallback for workbooks openpyxl rejects. Formula text lives in <f>
        elements; headers come from spreadsheet row 1; the same semantic
        assembly as the XLS path (_parse_xls_formula) builds definitions.
        Dedupes identical (name, expression) pairs — price books repeat the
        same markup formula across hundreds of row cells — capped at
        max_formulas to bound formula-memory noise.
        """
        import zipfile
        import xml.etree.ElementTree as ET

        def _local(tag: str) -> str:
            return tag.rsplit("}", 1)[-1]

        def _col_num(ref: str) -> int:
            n = 0
            for ch in ref:
                if ch.isalpha():
                    n = n * 26 + (ord(ch.upper()) - 64)
                else:
                    break
            return n

        out: List[Dict[str, Any]] = []
        seen: set = set()
        with zipfile.ZipFile(file_path) as zf:
            # Header cells reference shared strings by index — resolve them.
            # <rPh> phonetic runs are pronunciation hints, not content.
            shared: List[str] = []
            if "xl/sharedStrings.xml" in zf.namelist():
                sst_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in sst_root:
                    if _local(si.tag) != "si":
                        continue
                    parts: List[str] = []
                    for child in si:
                        tag = _local(child.tag)
                        if tag == "t":
                            parts.append(child.text or "")
                        elif tag == "r":
                            parts.extend(
                                sub.text or ""
                                for sub in child
                                if _local(sub.tag) == "t"
                            )
                    shared.append("".join(parts))

            def _cell_text(c) -> str:
                ctype = c.get("t")
                v = next((ch for ch in c if _local(ch.tag) == "v"), None)
                if v is None:
                    return ""
                if ctype == "s":
                    try:
                        return shared[int(v.text)]
                    except (ValueError, IndexError):
                        return v.text or ""
                return v.text or ""

            sheet_display: Dict[str, str] = {}
            try:
                wb = ET.fromstring(zf.read("xl/workbook.xml"))
                rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                rid_target = {
                    rel.get("Id"): rel.get("Target", "")
                    for rel in rels
                    if _local(rel.tag) == "Relationship"
                }
                for sh in wb.iter():
                    if _local(sh.tag) != "sheet":
                        continue
                    rid = sh.get(
                        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                    )
                    target = (rid_target.get(rid) or "").lstrip("/")
                    if target and not target.startswith("xl/"):
                        target = f"xl/{target}"
                    if target:
                        sheet_display[target] = sh.get("name") or target
            except Exception:
                pass  # names best-effort; part filenames still work

            sheets = sorted(
                n for n in zf.namelist()
                if n.startswith("xl/worksheets/") and n.endswith(".xml")
            )
            for sheet_path in sheets[:max_sheets]:
                sheet_name = sheet_display.get(
                    sheet_path, sheet_path.rsplit("/", 1)[-1]
                )
                root = ET.fromstring(zf.read(sheet_path))
                headers: Dict[int, str] = {}
                for row in root.iter():
                    if _local(row.tag) != "row":
                        continue
                    try:
                        row_num = int(row.get("r") or "0")
                    except ValueError:
                        row_num = 0
                    for c in row:
                        if _local(c.tag) != "c":
                            continue
                        ref = c.get("r") or ""
                        col = _col_num(ref)
                        if row_num == 1:
                            header_text = _cell_text(c).strip()
                            if header_text:
                                headers[col] = header_text
                            continue
                        f_el = next(
                            (ch for ch in c if _local(ch.tag) == "f"), None
                        )
                        ftext = (f_el.text or "").strip() if f_el is not None else ""
                        if not ftext:
                            continue  # shared-formula dependents carry no body
                        formula = ftext if ftext.startswith("=") else f"={ftext}"
                        formula_def = self._parse_xls_formula(
                            formula_str=formula,
                            row=max(row_num - 1, 0),
                            col=max(col - 1, 0),
                            headers=headers,
                            sheet_name=sheet_name,
                        )
                        if not formula_def:
                            continue
                        key = (formula_def.get("name"), formula_def.get("expression"))
                        if key in seen:
                            continue
                        seen.add(key)
                        out.append(formula_def)
                        if len(out) >= max_formulas:
                            return out
            return out

    def extract_from_xls(
        self,
        file_path: str,
        user_id: str = "default_user",
        auto_store: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Extract formulas from legacy .xls Excel files.
        Uses xlrd for older Excel format support.
        """
        try:
            import xlrd
        except ImportError:
            logger.warning("xlrd not installed. Run: pip install xlrd")
            # Fallback: try openpyxl with conversion
            return self.extract_from_excel(file_path, user_id, auto_store)

        extracted = []

        try:
            workbook = xlrd.open_workbook(file_path, formatting_info=False)

            for sheet in workbook.sheets():
                headers = self._get_xls_headers(sheet)
                for row_idx in range(1, sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        # XLS stores formula text differently
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            cell_value = str(cell.value)
                            if cell_value.startswith("="):
                                formula_def = self._parse_xls_formula(
                                    formula_str=cell_value,
                                    row=row_idx,
                                    col=col_idx,
                                    headers=headers,
                                    sheet_name=sheet.name
                                )
                                if formula_def:
                                    extracted.append(formula_def)

            if auto_store and extracted:
                self._store_formulas(extracted, user_id, file_path)

            logger.info(f"Extracted {len(extracted)} formulas from XLS {file_path}")
            return extracted

        except Exception as e:
            logger.error(f"Failed to extract from XLS {file_path}: {e}")
            return []

    def _get_xls_headers(self, sheet) -> Dict[int, str]:
        """Get column headers from XLS sheet."""
        headers = {}
        if sheet.nrows > 0:
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(0, col_idx)
                if cell.value:
                    headers[col_idx + 1] = str(cell.value).strip()
        return headers

    def _parse_xls_formula(
        self,
        formula_str: str,
        row: int,
        col: int,
        headers: Dict[int, str],
        sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """Parse a formula from XLS file."""
        result_name = headers.get(col + 1, f"Column_{col + 1}")
        formula_type = self._detect_formula_type(formula_str)
        referenced_cols = self._extract_cell_references(formula_str)
        parameters = []
        semantic_parts = []

        for col_letter, col_num in referenced_cols:
            param_name = headers.get(col_num, col_letter)
            parameters.append({"name": param_name, "type": "number"})
            semantic_parts.append(param_name)

        semantic_expression = self._create_semantic_expression(
            formula_str=formula_str,
            formula_type=formula_type,
            parameters=parameters
        )

        domain = self._detect_domain(result_name, semantic_parts)
        use_case = self._generate_use_case(result_name, formula_type, parameters)

        return {
            "expression": semantic_expression,
            "original_formula": formula_str,
            "name": result_name,
            "domain": domain,
            "use_case": use_case,
            "parameters": parameters,
            "source_sheet": sheet_name,
            "source_cell": f"R{row+1}C{col+1}"
        }

    def extract_from_csv(
        self,
        file_path: str,
        user_id: str = "default_user",
        auto_store: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Extract formulas from CSV files.
        CSV typically doesn't store formulas, but some tools export formula text.
        Also detects column relationships for implicit formulas.
        """
        import csv

        extracted = []

        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                return []

            headers = {i + 1: h.strip() for i, h in enumerate(rows[0]) if h.strip()}

            # Look for formula-like content in cells
            for row_idx, row in enumerate(rows[1:], start=2):
                for col_idx, cell in enumerate(row):
                    if cell.strip().startswith("="):
                        formula_def = self._parse_csv_formula(
                            formula_str=cell.strip(),
                            row=row_idx,
                            col=col_idx,
                            headers=headers
                        )
                        if formula_def:
                            extracted.append(formula_def)

            # Also detect implicit formulas from column patterns
            implicit = self._detect_implicit_formulas(rows, headers)
            extracted.extend(implicit)

            if auto_store and extracted:
                self._store_formulas(extracted, user_id, file_path)

            logger.info(f"Extracted {len(extracted)} formulas from CSV {file_path}")
            return extracted

        except Exception as e:
            logger.error(f"Failed to extract from CSV {file_path}: {e}")
            return []

    def _parse_csv_formula(
        self,
        formula_str: str,
        row: int,
        col: int,
        headers: Dict[int, str]
    ) -> Optional[Dict[str, Any]]:
        """Parse a formula from CSV cell."""
        result_name = headers.get(col + 1, f"Column_{col + 1}")
        formula_type = self._detect_formula_type(formula_str)
        referenced_cols = self._extract_cell_references(formula_str)
        parameters = []

        for col_letter, col_num in referenced_cols:
            param_name = headers.get(col_num, col_letter)
            parameters.append({"name": param_name, "type": "number"})

        semantic_expression = self._create_semantic_expression(
            formula_str=formula_str,
            formula_type=formula_type,
            parameters=parameters
        )

        domain = self._detect_domain(result_name, [p["name"] for p in parameters])
        use_case = self._generate_use_case(result_name, formula_type, parameters)

        return {
            "expression": semantic_expression,
            "original_formula": formula_str,
            "name": result_name,
            "domain": domain,
            "use_case": use_case,
            "parameters": parameters,
            "source_sheet": "CSV",
            "source_cell": f"R{row}C{col+1}"
        }

    def _detect_implicit_formulas(
        self,
        rows: List[List[str]],
        headers: Dict[int, str]
    ) -> List[Dict[str, Any]]:
        """
        Detect implicit formulas in CSV by analyzing column relationships.
        E.g., if 'Total' = 'Price' * 'Quantity' in most rows.
        """
        implicit = []

        if len(rows) < 3 or not headers:
            return implicit

        # Look for columns that might be calculated from others
        # This is a heuristic approach
        numeric_cols = []
        for col_idx in range(len(rows[0])):
            try:
                # Check if column is numeric
                values = [float(rows[i][col_idx]) for i in range(1, min(10, len(rows)))
                          if col_idx < len(rows[i]) and rows[i][col_idx].strip()]
                if len(values) >= 3:
                    numeric_cols.append(col_idx)
            except (ValueError, IndexError):
                continue

        # Try to detect multiplication/sum patterns
        # (Simplified - in production, use more sophisticated analysis)
        for result_col in numeric_cols:
            result_name = headers.get(result_col + 1, f"Column_{result_col + 1}")
            result_lower = result_name.lower()

            # Check for common calculated column names
            if any(kw in result_lower for kw in ["total", "sum", "amount", "subtotal"]):
                # Find potential input columns
                input_cols = [c for c in numeric_cols if c != result_col]
                if len(input_cols) >= 2:
                    param_names = [headers.get(c + 1, f"Column_{c + 1}") for c in input_cols[:2]]
                    implicit.append({
                        "expression": f"{param_names[0]} * {param_names[1]}",
                        "original_formula": "(implicit)",
                        "name": result_name,
                        "domain": self._detect_domain(result_name, param_names),
                        "use_case": f"Calculate {result_name} from {' and '.join(param_names)}",
                        "parameters": [{"name": p, "type": "number"} for p in param_names],
                        "source_sheet": "CSV",
                        "source_cell": "implicit"
                    })
                    break  # Only add one implicit formula per file

        return implicit

    def extract_from_ods(
        self,
        file_path: str,
        user_id: str = "default_user",
        auto_store: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Extract formulas from LibreOffice Calc (.ods) files.
        Uses odfpy or pyexcel-ods for ODS support.
        """
        try:
            from odf import table, text
            from odf.opendocument import load
        except ImportError:
            logger.warning("odfpy not installed. Run: pip install odfpy")
            return []

        extracted = []

        try:
            doc = load(file_path)
            spreadsheet = doc.spreadsheet

            for sheet in spreadsheet.getElementsByType(table.Table):
                sheet_name = sheet.getAttribute("name")
                rows = sheet.getElementsByType(table.TableRow)

                # Get headers from first row
                headers = {}
                if rows:
                    first_row = rows[0]
                    cells = first_row.getElementsByType(table.TableCell)
                    for idx, cell in enumerate(cells):
                        p = cell.getElementsByType(text.P)
                        if p:
                            headers[idx + 1] = str(p[0]).strip()

                # Look for formulas in subsequent rows
                for row_idx, row in enumerate(rows[1:], start=2):
                    cells = row.getElementsByType(table.TableCell)
                    for col_idx, cell in enumerate(cells):
                        formula = cell.getAttribute("formula")
                        if formula:
                            formula_def = self._parse_ods_formula(
                                formula_str=formula,
                                row=row_idx,
                                col=col_idx,
                                headers=headers,
                                sheet_name=sheet_name
                            )
                            if formula_def:
                                extracted.append(formula_def)

            if auto_store and extracted:
                self._store_formulas(extracted, user_id, file_path)

            logger.info(f"Extracted {len(extracted)} formulas from ODS {file_path}")
            return extracted

        except Exception as e:
            logger.error(f"Failed to extract from ODS {file_path}: {e}")
            return []

    def _parse_ods_formula(
        self,
        formula_str: str,
        row: int,
        col: int,
        headers: Dict[int, str],
        sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """Parse a formula from ODS file."""
        # ODS formulas often have 'of:' prefix, normalize to Excel-like
        if formula_str.startswith("of:"):
            formula_str = formula_str[3:]
        if not formula_str.startswith("="):
            formula_str = "=" + formula_str

        result_name = headers.get(col + 1, f"Column_{col + 1}")
        formula_type = self._detect_formula_type(formula_str)
        referenced_cols = self._extract_cell_references(formula_str)
        parameters = []

        for col_letter, col_num in referenced_cols:
            param_name = headers.get(col_num, col_letter)
            parameters.append({"name": param_name, "type": "number"})

        semantic_expression = self._create_semantic_expression(
            formula_str=formula_str,
            formula_type=formula_type,
            parameters=parameters
        )

        domain = self._detect_domain(result_name, [p["name"] for p in parameters])
        use_case = self._generate_use_case(result_name, formula_type, parameters)

        return {
            "expression": semantic_expression,
            "original_formula": formula_str,
            "name": result_name,
            "domain": domain,
            "use_case": use_case,
            "parameters": parameters,
            "source_sheet": sheet_name,
            "source_cell": f"R{row}C{col+1}"
        }


    def _extract_from_sheet(
        self,
        sheet,
        sheet_name: str
    ) -> List[Dict[str, Any]]:
        """Extract formulas from a single worksheet."""
        formulas = []
        headers = self._get_column_headers(sheet)

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_def = self._parse_formula(
                        cell=cell,
                        formula_str=cell.value,
                        headers=headers,
                        sheet_name=sheet_name
                    )
                    if formula_def:
                        formulas.append(formula_def)

        return formulas

    def _get_column_headers(self, sheet) -> Dict[int, str]:
        """Get column headers from the first row."""
        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[idx] = str(cell.value).strip()
        return headers

    def _parse_formula(
        self,
        cell,
        formula_str: str,
        headers: Dict[int, str],
        sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """Parse a formula cell and create a definition."""
        # Get result column name
        result_name = headers.get(cell.column, f"Column_{cell.column}")

        # Detect formula type
        formula_type = self._detect_formula_type(formula_str)

        # Extract cell references and map to headers
        referenced_cols = self._extract_cell_references(formula_str)
        parameters = []
        semantic_parts = []

        for col_letter, col_num in referenced_cols:
            param_name = headers.get(col_num, col_letter)
            parameters.append({"name": param_name, "type": "number"})
            semantic_parts.append(param_name)

        # Create semantic expression
        semantic_expression = self._create_semantic_expression(
            formula_str=formula_str,
            formula_type=formula_type,
            parameters=parameters
        )

        # Detect domain
        domain = self._detect_domain(result_name, semantic_parts)

        # Generate use case description
        use_case = self._generate_use_case(
            result_name=result_name,
            formula_type=formula_type,
            parameters=parameters
        )

        return {
            "expression": semantic_expression,
            "original_formula": formula_str,
            "name": result_name,
            "domain": domain,
            "use_case": use_case,
            "parameters": parameters,
            "source_sheet": sheet_name,
            "source_cell": cell.coordinate
        }

    def _detect_formula_type(self, formula_str: str) -> str:
        """Detect the type of Excel formula."""
        formula_upper = formula_str.upper()

        for func_name in self.FORMULA_PATTERNS.keys():
            if func_name in formula_upper:
                return func_name

        # Check for simple arithmetic
        if any(op in formula_str for op in ["+", "-", "*", "/"]):
            return "ARITHMETIC"

        return "CUSTOM"

    def _extract_cell_references(self, formula_str: str) -> List[Tuple[str, int]]:
        """Extract cell references and their column numbers."""
        # Match patterns like A1, B2, $C$3
        pattern = r"\$?([A-Z]+)\$?\d+"
        matches = re.findall(pattern, formula_str.upper())

        # Dedupe while preserving first-occurrence order. A `set` here made
        # the parameter order of the extracted semantic expression depend on
        # PYTHONHASHSEED (set iteration order), so the same formula could be
        # stored as "sum(Revenue, Cost)" in one run and "sum(Cost, Revenue)"
        # in another — nondeterministic memory content across workers.
        results = []
        seen = set()
        for col_letter in matches:
            if col_letter not in seen:
                seen.add(col_letter)
                results.append((col_letter, self._column_letter_to_number(col_letter)))

        return results

    def _column_letter_to_number(self, col_letter: str) -> int:
        """Convert column letter to number (A=1, B=2, etc.)."""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    def _create_semantic_expression(
        self,
        formula_str: str,
        formula_type: str,
        parameters: List[Dict[str, str]]
    ) -> str:
        """Create a semantic expression from the formula."""
        param_names = [p["name"] for p in parameters]

        if formula_type == "SUM":
            return f"sum({', '.join(param_names)})"
        elif formula_type == "AVERAGE":
            return f"average({', '.join(param_names)})"
        elif formula_type == "ARITHMETIC" and len(param_names) == 2:
            # Try to infer the operation
            if "-" in formula_str:
                return f"{param_names[0]} - {param_names[1]}"
            elif "+" in formula_str:
                return f"{param_names[0]} + {param_names[1]}"
            elif "*" in formula_str:
                return f"{param_names[0]} * {param_names[1]}"
            elif "/" in formula_str:
                return f"{param_names[0]} / {param_names[1]}"

        # Default: use parameter names with the formula type
        return f"{formula_type.lower()}({', '.join(param_names)})"

    def _detect_domain(self, result_name: str, param_names: List[str]) -> str:
        """Detect the domain based on column names."""
        all_names = [result_name.lower()] + [n.lower() for n in param_names]
        text = " ".join(all_names)

        for domain, keywords in self.DOMAIN_KEYWORDS.items():
            for keyword in keywords:
                if keyword in text:
                    return domain

        return "general"

    def _generate_use_case(
        self,
        result_name: str,
        formula_type: str,
        parameters: List[Dict[str, str]]
    ) -> str:
        """Generate a natural language use case description."""
        param_names = [p["name"] for p in parameters]

        if formula_type == "SUM":
            return f"Calculate the total {result_name} by summing {' and '.join(param_names)}"
        elif formula_type == "AVERAGE":
            return f"Calculate the average {result_name} from {' and '.join(param_names)}"
        elif formula_type == "ARITHMETIC":
            return f"Compute {result_name} using {' and '.join(param_names)}"
        else:
            return f"Calculate {result_name} using {formula_type} formula"

    def _store_formulas(
        self,
        formulas: List[Dict[str, Any]],
        user_id: str,
        source_file: str
    ):
        """Store extracted formulas in Atom's memory."""
        manager = self._get_formula_manager()
        source_name = Path(source_file).name

        stored_count = 0
        for formula in formulas:
            try:
                formula_id = manager.add_formula(
                    expression=formula["expression"],
                    name=formula["name"],
                    domain=formula["domain"],
                    use_case=formula["use_case"],
                    parameters=formula["parameters"],
                    source=f"excel:{source_name}",
                    user_id=user_id
                )
                if formula_id:
                    stored_count += 1
            except Exception as e:
                logger.warning(f"Failed to store formula {formula['name']}: {e}")

        logger.info(f"Stored {stored_count}/{len(formulas)} formulas from {source_name}")


def get_formula_extractor(workspace_id: str = "default") -> FormulaExtractor:
    """Get a formula extractor instance."""
    return FormulaExtractor(workspace_id)
