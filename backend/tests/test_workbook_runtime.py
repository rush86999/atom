"""
Tests for the workbook runtime — formula evaluation, structural ops, rendering.

Verifies the runtime can evaluate formulas, insert rows with reference
maintenance, and render to HTML. Tests run against whichever engine is
available (LibreOffice > formulas > openpyxl fallback).
"""

from __future__ import annotations

import asyncio
import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def sample_workbook(tmp_path):
    """Create a workbook with formulas for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["A4"] = "=SUM(A1:A3)"
    ws["A5"] = "=AVERAGE(A1:A3)"
    ws["B1"] = "=A1*2"
    ws["B2"] = "=IF(A1>5, \"big\", \"small\")"

    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    return file_path


class TestWorkbookRuntime:

    def test_runtime_exists(self):
        from core.workbook_runtime import WorkbookRuntime
        assert WorkbookRuntime is not None

    def test_singleton(self):
        from core.workbook_runtime import get_workbook_runtime
        r1 = get_workbook_runtime()
        r2 = get_workbook_runtime()
        assert r1 is r2

    def test_engine_reported(self):
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        engine = runtime.engine
        assert engine in ("libreoffice", "formulas", "openpyxl")

    def test_can_evaluate_flag(self):
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        # can_evaluate should be True for libreoffice/formulas, False for bare openpyxl
        if runtime.engine == "openpyxl":
            assert runtime.can_evaluate is False
        else:
            assert runtime.can_evaluate is True

    @pytest.mark.asyncio
    async def test_recalculate_does_not_crash(self, sample_workbook):
        """Recalculate should run without error regardless of engine."""
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        result = await runtime.recalculate(sample_workbook)
        assert result.exists()

    @pytest.mark.asyncio
    async def test_get_evaluated_range(self, sample_workbook):
        """get_evaluated_range returns cell values."""
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        result = await runtime.get_evaluated_range(sample_workbook, "Sheet1", "A1", "A3")
        assert result["success"] is True
        assert len(result["values"]) >= 1

    @pytest.mark.asyncio
    async def test_insert_rows(self, sample_workbook):
        """insert_rows adds rows without crashing."""
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        result = await runtime.insert_rows(sample_workbook, "Sheet1", 2, 1)
        assert result["success"] is True
        assert result["rows_inserted"] == 1

    @pytest.mark.asyncio
    async def test_insert_cols(self, sample_workbook):
        """insert_cols adds columns without crashing."""
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        result = await runtime.insert_cols(sample_workbook, "Sheet1", 2, 1)
        assert result["success"] is True

    @pytest.mark.asyncio
    async def test_render_to_html(self, sample_workbook):
        """render_to_html returns an HTML string."""
        from core.workbook_runtime import get_workbook_runtime
        runtime = get_workbook_runtime()
        html = await runtime.render_to_html(sample_workbook)
        assert isinstance(html, str)
        assert len(html) > 0
        # Should contain some form of table or content
        assert any(tag in html.lower() for tag in ["<table", "<div", "<p", "<body"])


class TestExcelManagerIntegration:

    def test_excel_manager_has_runtime_methods(self):
        """ExcelManager exposes the new runtime methods."""
        from core.office_service import ExcelManager
        assert hasattr(ExcelManager, "insert_rows")
        assert hasattr(ExcelManager, "insert_columns")
        assert hasattr(ExcelManager, "get_evaluated_range")
        assert hasattr(ExcelManager, "recalculate")

    def test_write_cell_returns_formula_field(self, tmp_path):
        """write_cell now includes a 'formula' field in the response."""
        from core.office_service import ExcelManager
        file_path = str(tmp_path / "test.xlsx")
        result = ExcelManager().write_cell(file_path, "/Sheet1/A1", "=SUM(1,2,3)", is_formula=True)
        assert result["success"] is True
        assert "formula" in result
        assert result["formula"] == "=SUM(1,2,3)"

    def test_renderer_uses_runtime(self):
        """DocumentRenderer.render_to_html uses the workbook runtime (not xlsx2html)."""
        from core.office_service import DocumentRenderer
        import inspect
        source = inspect.getsource(DocumentRenderer.render_to_html)
        assert "workbook_runtime" in source or "WorkbookRuntime" in source, \
            "Renderer should use the workbook runtime, not xlsx2html"
        assert "xlsx2html" not in source, \
            "Renderer should NOT use xlsx2html (replaced with runtime)"

    @pytest.mark.asyncio
    async def test_add_pivot_table(self, tmp_path):
        """test programmatic creation of pivot table"""
        from core.workbook_runtime import get_workbook_runtime
        from openpyxl import Workbook
        
        # Create a sample sheet with columns: Product, Region, Sales
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        ws.append(["Product", "Region", "Sales"])
        ws.append(["Apple", "North", 100])
        ws.append(["Apple", "South", 150])
        ws.append(["Banana", "North", 200])
        ws.append(["Banana", "South", 250])
        
        file_path = tmp_path / "sales.xlsx"
        wb.save(file_path)
        
        runtime = get_workbook_runtime()
        result = await runtime.add_pivot_table(
            file_path=file_path,
            sheet_name="DataSheet",
            pivot_sheet_name="PivotSheet",
            data_range="A1:C5",
            rows=["Product"],
            columns=["Region"],
            values=[{"field": "Sales", "function": "sum"}]
        )
        assert result["success"] is True
        assert result["pivot_sheet"] == "PivotSheet"

    @pytest.mark.asyncio
    async def test_run_macro_fails_safely_when_no_macro_present(self, tmp_path):
        """test that run_macro fails gracefully when macro name standard subroutine not found"""
        from core.workbook_runtime import get_workbook_runtime
        from openpyxl import Workbook
        wb = Workbook()
        file_path = tmp_path / "macro_test.xlsx"
        wb.save(file_path)
        
        runtime = get_workbook_runtime()
        # Should return success=False gracefully or not crash
        res = await runtime.run_macro(file_path, "NonExistentMacro")
        assert res["success"] is False
