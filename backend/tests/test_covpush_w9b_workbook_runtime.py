"""
Coverage wave 9b — core/workbook_runtime.py (62% -> 90%+ target).

Bugs fixed (TDD):
1. str(e) leak: add_pivot_table returned the raw exception text
   ("Failed to generate pivot table: {e}") — internal paths/frames reach
   the agent; now generic ("Failed to generate pivot table").
2. str(e) leak: _render_html_basic leaked the exception into rendered
   HTML ("<p>Error rendering: {e}</p>"); now generic.
"""
import asyncio
import sys
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import pytest


@pytest.fixture
def workbook(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)
    return path


class TestEngineSelection:
    def test_engine_libreoffice_when_soffice(self, monkeypatch):
        from core import workbook_runtime as wb

        monkeypatch.setattr(wb, "_SOFFICE", "/usr/bin/soffice")
        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"
        assert rt.engine == "libreoffice"
        assert rt.can_evaluate is True
        assert rt.can_render is True

    def test_engine_formulas_when_importable(self, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = True
        assert rt.engine == "formulas"
        assert rt.can_evaluate is True
        assert rt.can_render is False

    def test_engine_openpyxl_fallback(self):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = False
        assert rt.engine == "openpyxl"
        assert rt.can_evaluate is False
        assert rt.can_render is False

    def test_find_soffice_no_match(self, monkeypatch):
        from core import workbook_runtime as wb

        monkeypatch.setattr(wb.shutil, "which", lambda name: None)
        monkeypatch.setattr(wb.os.path, "exists", lambda p: False)
        assert wb._find_soffice() is None

    def test_find_soffice_which_hit(self, monkeypatch):
        from core import workbook_runtime as wb

        monkeypatch.setattr(wb.shutil, "which", lambda name: "/usr/bin/soffice")
        assert wb._find_soffice() == "/usr/bin/soffice"

    def test_check_formulas_importable(self, monkeypatch):
        from core import workbook_runtime as wb

        monkeypatch.setitem(sys.modules, "formulas", MagicMock())
        assert wb.WorkbookRuntime()._check_formulas() is True

    def test_find_soffice_mac_path(self, monkeypatch):
        from core import workbook_runtime as wb

        monkeypatch.setattr(wb.shutil, "which", lambda name: None)
        monkeypatch.setattr(wb.os.path, "exists", lambda p: True)
        assert wb._find_soffice() == "/Applications/LibreOffice.app/Contents/MacOS/soffice"


class TestRecalculate:
    @pytest.mark.asyncio
    async def test_missing_file_raises(self, tmp_path):
        from core.workbook_runtime import WorkbookRuntime

        with pytest.raises(FileNotFoundError):
            await WorkbookRuntime().recalculate(tmp_path / "nope.xlsx")

    @pytest.mark.asyncio
    async def test_no_engine_returns_unchanged(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = False
        assert await rt.recalculate(workbook) == workbook

    @pytest.mark.asyncio
    async def test_recalculate_dispatches_soffice(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"
        rt._recalc_with_soffice = AsyncMock(return_value=workbook)
        assert await rt.recalculate(workbook) == workbook
        rt._recalc_with_soffice.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_recalculate_dispatches_formulas(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = True
        rt._recalc_with_formulas = AsyncMock(return_value=workbook)
        assert await rt.recalculate(workbook) == workbook
        rt._recalc_with_formulas.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_soffice_recalc_success(self, workbook, tmp_path, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        out_dir = tmp_path / "out"
        out_dir.mkdir()
        (out_dir / workbook.name).write_bytes(b"recalced")

        proc = AsyncMock()
        proc.communicate.return_value = (b"", b"")

        class FakeTmpDir:
            def __enter__(self):
                return out_dir

            def __exit__(self, *a):
                return False

        monkeypatch.setattr(wb.tempfile, "TemporaryDirectory", lambda: FakeTmpDir())
        monkeypatch.setattr(
            wb.asyncio, "create_subprocess_exec", AsyncMock(return_value=proc)
        )
        monkeypatch.setattr(wb.shutil, "copy2", Mock())

        result = await rt._recalc_with_soffice(workbook)
        assert result == workbook
        wb.shutil.copy2.assert_called_once()

    @pytest.mark.asyncio
    async def test_soffice_recalc_no_output(self, workbook, tmp_path, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        proc = AsyncMock()
        proc.communicate.return_value = (b"", b"no output")
        monkeypatch.setattr(
            wb.asyncio, "create_subprocess_exec", AsyncMock(return_value=proc)
        )

        result = await rt._recalc_with_soffice(workbook)
        assert result == workbook

    @pytest.mark.asyncio
    async def test_soffice_recalc_timeout(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        proc = AsyncMock()
        proc.communicate.side_effect = asyncio.TimeoutError
        monkeypatch.setattr(
            wb.asyncio, "create_subprocess_exec", AsyncMock(return_value=proc)
        )

        assert await rt._recalc_with_soffice(workbook) == workbook

    @pytest.mark.asyncio
    async def test_soffice_recalc_generic_exception(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        monkeypatch.setattr(
            wb.asyncio,
            "create_subprocess_exec",
            AsyncMock(side_effect=RuntimeError("soffice blew up")),
        )

        assert await rt._recalc_with_soffice(workbook) == workbook

    @pytest.mark.asyncio
    async def test_formulas_recalc(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = True

        fake_formulas = MagicMock()
        model = MagicMock()
        solution = {"'[book.xlsx]Sheet1'!A3": 3}
        fake_formulas.ExcelModel.return_value.loads.return_value.finish.return_value = model
        model.calculate.return_value = solution

        fake_openpyxl = MagicMock()
        fake_ws = MagicMock()
        cell = MagicMock()
        cell.value = "=SUM(A1:A2)"
        cell.coordinate = "A3"
        fake_ws.iter_rows.return_value = [[cell]]
        fake_wb = MagicMock()
        fake_wb.sheetnames = ["Sheet1"]
        fake_wb.__getitem__.return_value = fake_ws
        fake_openpyxl.load_workbook.return_value = fake_wb

        monkeypatch.setitem(sys.modules, "formulas", fake_formulas)
        monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)

        result = await rt._recalc_with_formulas(workbook)
        assert result == workbook
        assert cell.value == 3
        fake_wb.save.assert_called_once_with(workbook)

    @pytest.mark.asyncio
    async def test_formulas_recalc_value_object(self, workbook, monkeypatch):
        """Solution entries with a .value attribute take that value."""
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = True

        fake_formulas = MagicMock()
        model = MagicMock()
        solution = {"'[book.xlsx]Sheet1'!A3": Mock(value=9)}
        fake_formulas.ExcelModel.return_value.loads.return_value.finish.return_value = model
        model.calculate.return_value = solution

        fake_openpyxl = MagicMock()
        fake_ws = MagicMock()
        cell = MagicMock()
        cell.value = "=SUM(A1:A2)"
        cell.coordinate = "A3"
        fake_ws.iter_rows.return_value = [[cell]]
        fake_wb = MagicMock()
        fake_wb.sheetnames = ["Sheet1"]
        fake_wb.__getitem__.return_value = fake_ws
        fake_openpyxl.load_workbook.return_value = fake_wb

        monkeypatch.setitem(sys.modules, "formulas", fake_formulas)
        monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)

        await rt._recalc_with_formulas(workbook)
        assert cell.value == 9

    @pytest.mark.asyncio
    async def test_formulas_recalc_exception(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = None
        rt._has_formulas = True

        fake_formulas = MagicMock()
        fake_formulas.ExcelModel.side_effect = RuntimeError("formulas broke")
        monkeypatch.setitem(sys.modules, "formulas", fake_formulas)

        assert await rt._recalc_with_formulas(workbook) == workbook


class TestRunMacro:
    @pytest.mark.asyncio
    async def test_missing_file(self, tmp_path):
        from core.workbook_runtime import WorkbookRuntime

        res = await WorkbookRuntime().run_macro(tmp_path / "nope.xlsx", "M")
        assert res["success"] is False

    @pytest.mark.asyncio
    async def test_no_soffice(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = None
        res = await rt.run_macro(workbook, "M")
        assert res["success"] is False
        assert "LibreOffice" in res["error"]

    @pytest.mark.asyncio
    async def test_sandbox_success(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"
        rt.recalculate = AsyncMock(return_value=workbook)

        fake_sandbox = MagicMock()
        fake_sandbox.execute_in_sandbox = AsyncMock(return_value=True)
        with patch("core.firecracker_sandbox.get_sandbox", return_value=fake_sandbox):
            res = await rt.run_macro(workbook, "MyMacro")
        assert res["success"] is True
        assert res["macro"] == "MyMacro"
        rt.recalculate.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_sandbox_failure(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        fake_sandbox = MagicMock()
        fake_sandbox.execute_in_sandbox = AsyncMock(return_value=False)
        with patch("core.firecracker_sandbox.get_sandbox", return_value=fake_sandbox):
            res = await rt.run_macro(workbook, "MyMacro")
        assert res["success"] is False


class TestAddPivotTable:
    @pytest.mark.asyncio
    async def test_source_sheet_missing(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        res = await WorkbookRuntime().add_pivot_table(
            workbook, "Nope", "Pivot", "A1:A3", ["A"], ["B"], [{"field": "C"}]
        )
        assert res["success"] is False
        assert "not found" in res["error"]

    @pytest.mark.asyncio
    async def test_empty_source_sheet(self, tmp_path):
        from openpyxl import Workbook

        from core.workbook_runtime import WorkbookRuntime

        path = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.create_sheet("Empty")
        wb.save(path)

        res = await WorkbookRuntime().add_pivot_table(
            path, "Empty", "Pivot", "A1:A3", ["A"], ["B"], [{"field": "C"}]
        )
        assert res["success"] is False
        assert "no data" in res["error"]

    @pytest.mark.asyncio
    async def test_pivot_success_replaces_existing(self, tmp_path):
        from openpyxl import Workbook

        from core.workbook_runtime import WorkbookRuntime

        path = tmp_path / "pivot.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Region", "Amount"])
        ws.append(["East", 100])
        ws.append(["West", 50])
        ws.append(["East", 25])
        wb.save(path)

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=path)

        # Pre-existing pivot sheet must be replaced.
        wb = __import__("openpyxl").load_workbook(path)
        wb.create_sheet("Pivot")
        wb.save(path)

        res = await rt.add_pivot_table(
            path,
            "Data",
            "Pivot",
            "A1:B4",
            ["Region"],
            [],
            [{"field": "Amount", "function": "sum"}],
        )
        assert res["success"] is True
        assert res["pivot_sheet"] == "Pivot"
        rt.recalculate.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_exception_generic_no_str_leak(self, workbook, monkeypatch):
        """RED (bug 1): exceptions must not leak their message to the caller."""
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()

        fake_openpyxl = MagicMock()
        fake_openpyxl.load_workbook.side_effect = RuntimeError("pivot-internal-secret")
        monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)

        res = await rt.add_pivot_table(
            workbook, "Sheet1", "Pivot", "A1:A3", ["A"], ["B"], [{"field": "A"}]
        )
        assert res["success"] is False
        assert res["error"] == "Failed to generate pivot table"
        assert "pivot-internal-secret" not in res["error"]


class TestRender:
    @pytest.mark.asyncio
    async def test_render_missing_file(self, tmp_path):
        from core.workbook_runtime import WorkbookRuntime

        html = await WorkbookRuntime().render_to_html(tmp_path / "nope.xlsx")
        assert "File not found" in html

    @pytest.mark.asyncio
    async def test_render_soffice_branch(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"
        rt._render_html_with_soffice = AsyncMock(return_value="<html>lo</html>")
        assert await rt.render_to_html(workbook) == "<html>lo</html>"

    @pytest.mark.asyncio
    async def test_render_basic_fallback(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt._soffice = None
        html = await rt.render_to_html(workbook)
        assert "<table" in html
        assert "Sheet1" in html
        assert "<td>1</td>" in html

    @pytest.mark.asyncio
    async def test_render_soffice_success(self, workbook, tmp_path, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        out_dir = tmp_path / "out"
        out_dir.mkdir()
        (out_dir / (workbook.stem + ".html")).write_text("<html>pixel-perfect</html>")

        proc = AsyncMock()
        proc.communicate.return_value = (b"", b"")

        class FakeTmpDir:
            def __enter__(self):
                return out_dir

            def __exit__(self, *a):
                return False

        monkeypatch.setattr(wb.tempfile, "TemporaryDirectory", lambda: FakeTmpDir())
        monkeypatch.setattr(
            wb.asyncio, "create_subprocess_exec", AsyncMock(return_value=proc)
        )

        html = await rt._render_html_with_soffice(workbook)
        assert html == "<html>pixel-perfect</html>"

    @pytest.mark.asyncio
    async def test_render_soffice_failure_falls_back(self, workbook, monkeypatch):
        from core import workbook_runtime as wb

        rt = wb.WorkbookRuntime()
        rt._soffice = "/usr/bin/soffice"

        monkeypatch.setattr(
            wb.asyncio,
            "create_subprocess_exec",
            AsyncMock(side_effect=RuntimeError("soffice render broke")),
        )

        html = await rt._render_html_with_soffice(workbook)
        assert "<table" in html

    def test_render_basic_exception_generic(self, tmp_path, monkeypatch):
        """RED (bug 2): internal render errors must not leak into HTML."""
        from core import workbook_runtime as wb

        fake_openpyxl = MagicMock()
        fake_openpyxl.load_workbook.side_effect = RuntimeError("render-internal-secret")
        monkeypatch.setitem(sys.modules, "openpyxl", fake_openpyxl)

        html = wb.WorkbookRuntime()._render_html_basic(tmp_path / "bad.xlsx")
        assert "render-internal-secret" not in html


class TestStructuralOps:
    @pytest.mark.asyncio
    async def test_insert_rows_missing_sheet(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        res = await WorkbookRuntime().insert_rows(workbook, "Nope", 2)
        assert res["success"] is False

    @pytest.mark.asyncio
    async def test_insert_rows_success(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.insert_rows(workbook, "Sheet1", 2, count=2)
        assert res["success"] is True
        assert res["rows_inserted"] == 2
        rt.recalculate.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_insert_cols_missing_sheet(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        res = await WorkbookRuntime().insert_cols(workbook, "Nope", 2)
        assert res["success"] is False

    @pytest.mark.asyncio
    async def test_insert_cols_success(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.insert_cols(workbook, "Sheet1", 2, count=3)
        assert res["success"] is True
        assert res["cols_inserted"] == 3
        rt.recalculate.assert_awaited_once()


class TestEvaluatedReads:
    @pytest.mark.asyncio
    async def test_get_evaluated_range_range(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.get_evaluated_range(workbook, "Sheet1", "A1", "A2")
        assert res["success"] is True
        assert res["range"] == "A1:A2"
        assert res["values"][0][0]["cell"] == "A1"
        assert res["values"][0][0]["value"] == 1

    @pytest.mark.asyncio
    async def test_get_evaluated_range_single_cell(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.get_evaluated_range(workbook, "Sheet1", "A1")
        assert res["success"] is True
        assert res["values"][0][0]["cell"] == "A1"

    @pytest.mark.asyncio
    async def test_get_evaluated_range_missing_sheet(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.get_evaluated_range(workbook, "Nope", "A1")
        assert res["success"] is False

    @pytest.mark.asyncio
    async def test_get_formula_result_delegates(self, workbook):
        from core.workbook_runtime import WorkbookRuntime

        rt = WorkbookRuntime()
        rt.recalculate = AsyncMock(return_value=workbook)
        res = await rt.get_formula_result(workbook, "Sheet1", "A1")
        assert res["success"] is True


def test_get_workbook_runtime_singleton():
    from core import workbook_runtime as wb

    orig = wb._workbook_runtime
    wb._workbook_runtime = None
    try:
        rt1 = wb.get_workbook_runtime()
        assert wb.get_workbook_runtime() is rt1
    finally:
        wb._workbook_runtime = orig
