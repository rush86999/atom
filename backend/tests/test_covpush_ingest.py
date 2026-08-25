"""
Coverage-push tests for assigned modules (Round: covpush_ingest).

Covers:
- core.auto_document_ingestion
- core.ai_workflow_optimizer / core.ai_workflow_optimization_endpoints
- core.automation_insight_manager / core.automation_settings
- core.backfill_job_queue
- core.alert_service
- core.atom_saas_client / core.atom_saas_websocket
- core.accounting_validator

External HTTP/WS/DB dependencies are mocked.
"""

import asyncio
import json
import os
import sqlite3
import time
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import Any, Optional
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from core.accounting_validator import (
    DoubleEntryValidationError,
    EntryType,
    check_balance_sheet,
    validate_double_entry,
    validate_journal_entries,
)
from core.alert_service import (
    AlertEvaluationResult,
    AlertSeverity,
    AlertStatus,
    AlertThresholdService,
    AlertViolation,
)
from core.ai_workflow_optimization_endpoints import router as optimizer_router
from core.ai_workflow_optimizer import (
    AIWorkflowOptimizer,
    ImpactLevel,
    OptimizationRecommendation,
    OptimizationType,
    WorkflowAnalysis,
    get_ai_workflow_optimizer,
)
from core.automation_settings import AutomationSettingsManager
from core.backfill_job_queue import (
    BackfillJobQueue,
    BackfillJobStatus,
    get_backfill_job_queue,
)
from core.atom_saas_client import AtomSaaSConfig, AtomSaaSClient
from core.atom_saas_websocket import (
    AtomSaaSWebSocketClient,
    WebSocketConnectionError,
)


# ============================================================================
# BUG 1: alert_service imports a non-existent model -> AlertThresholdService
# construction raises ImportError. (TDD: red -> green)
# ============================================================================

class TestAlertServiceInitBug:
    def test_alert_service_constructs_without_models_mock(self):
        service = AlertThresholdService(Mock())
        assert service.db is not None
        assert service.redis is None


# ============================================================================
# BUG 2: backfill get_job_status crashes on non-numeric retry counter.
# ============================================================================

class TestBackfillJobStatusNonNumeric:
    @pytest.mark.asyncio
    async def test_get_job_status_non_numeric_retry(self):
        queue = BackfillJobQueue()
        redis = MagicMock()
        redis.get = AsyncMock(side_effect=[b"processing", b"oops"])
        redis.hgetall = AsyncMock(return_value={})
        queue._client = redis
        status = await queue.get_job_status("job-1")
        assert status["status"] == "processing"
        assert status["retry_count"] == 0


# ============================================================================
# BUG 3: optimization-plan endpoint KeyError when workflow_analysis lacks
# optimization_opportunities -> 500. (TDD: red -> green)
# ============================================================================

class TestOptimizationPlanMissingKey:
    @pytest.fixture
    def app_client(self):
        optimizer = Mock()
        optimizer.optimize_workflow_plan = AsyncMock(return_value={
            "optimization_plan": {"steps": ["s1"]},
            "workflow_analysis": {
                "workflow_id": "wf-1",
                "workflow_name": "Test",
                "complexity_score": 0.5,
                "failure_points": [],
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })
        app = FastAPI()
        app.include_router(optimizer_router)
        app.dependency_overrides[get_ai_workflow_optimizer] = lambda: optimizer
        return TestClient(app)

    def test_plan_without_opportunities_key(self, app_client):
        response = app_client.post(
            "/api/v1/workflows/optimization-plan",
            json={"workflow_data": {"nodes": []}, "optimization_goals": ["performance"]},
        )
        assert response.status_code == 200
        assert response.json()["success"] is True


# ============================================================================
# accounting_validator
# ============================================================================

class TestAccountingValidator:
    def test_validate_double_entry_balanced(self):
        result = validate_double_entry([
            {"account_id": "a1", "type": EntryType.DEBIT, "amount": "100.00"},
            {"account_id": "a2", "type": EntryType.CREDIT, "amount": "100.00"},
        ])
        assert result["balanced"] is True
        assert result["debits"] == Decimal("100.00")
        assert result["credits"] == Decimal("100.00")

    def test_validate_double_entry_unbalanced_raises(self):
        with pytest.raises(DoubleEntryValidationError) as exc:
            validate_double_entry([
                {"account_id": "a1", "type": "debit", "amount": "100.00"},
                {"account_id": "a2", "type": "credit", "amount": "99.99"},
            ])
        assert exc.value.debits == Decimal("100.00")
        assert exc.value.credits == Decimal("99.99")
        assert exc.value.difference == Decimal("0.01")

    def test_validate_double_entry_empty_raises(self):
        with pytest.raises(DoubleEntryValidationError):
            validate_double_entry([])

    def test_validate_double_entry_single_entry_raises(self):
        with pytest.raises(DoubleEntryValidationError):
            validate_double_entry([{"account_id": "a1", "type": "debit", "amount": "5"}])

    def test_validate_double_entry_missing_fields_raises(self):
        with pytest.raises(ValueError):
            validate_double_entry([{"account_id": "a1"}, {"account_id": "a2", "type": "credit"}])

    def test_validate_double_entry_invalid_amount_raises(self):
        with pytest.raises(ValueError):
            validate_double_entry([
                {"account_id": "a1", "type": "debit", "amount": "abc"},
                {"account_id": "a2", "type": "credit", "amount": "1"},
            ])

    def test_validate_double_entry_negative_amount_raises(self):
        with pytest.raises(DoubleEntryValidationError):
            validate_double_entry([
                {"account_id": "a1", "type": "debit", "amount": "-1"},
                {"account_id": "a2", "type": "credit", "amount": "1"},
            ])

    def test_validate_double_entry_invalid_type_raises(self):
        with pytest.raises(ValueError):
            validate_double_entry([
                {"account_id": "a1", "type": "weird", "amount": "1"},
                {"account_id": "a2", "type": "credit", "amount": "1"},
            ])

    def test_validate_double_entry_rounds_amounts(self):
        result = validate_double_entry([
            {"account_id": "a1", "type": "DEBIT", "amount": "10.005"},
            {"account_id": "a2", "type": "CREDIT", "amount": "10.005"},
        ])
        assert result["balanced"] is True

    def test_check_balance_sheet_balanced(self):
        result = check_balance_sheet({
            "assets": ["100.00"],
            "liabilities": ["60.00"],
            "equity": ["40.00"],
        })
        assert result["balanced"] is True
        assert result["discrepancy"] is None
        assert result["assets"] == Decimal("100.00")

    def test_check_balance_sheet_unbalanced(self):
        result = check_balance_sheet({
            "assets": ["100.00"],
            "liabilities": ["50.00"],
            "equity": ["45.00"],
        })
        assert result["balanced"] is False
        assert result["discrepancy"] == Decimal("5.00")

    def test_check_balance_sheet_missing_keys(self):
        result = check_balance_sheet({})
        assert result["balanced"] is True

    def test_validate_journal_entries_no_errors(self):
        errors = validate_journal_entries([
            {"account_id": "a1", "type": "debit", "amount": "10.00"},
            {"account_id": "a2", "type": "credit", "amount": "10.00"},
        ])
        assert errors == []

    def test_validate_journal_entries_collects_errors(self):
        errors = validate_journal_entries([
            {},
            {"account_id": "a1", "type": "debit", "amount": "-5"},
            {"account_id": "a2", "type": "credit", "amount": "nan"},
        ])
        assert any("missing account_id" in e for e in errors)
        assert any("missing type" in e for e in errors)
        assert any("missing amount" in e for e in errors)
        assert any("negative amount" in e for e in errors)
        assert any("invalid amount" in e for e in errors)


# ============================================================================
# automation_settings
# ============================================================================

class TestAutomationSettings:
    def test_defaults_when_no_file(self, tmp_path, monkeypatch):
        settings_file = tmp_path / "automation_settings.json"
        monkeypatch.setattr(AutomationSettingsManager, "SETTINGS_FILE", str(settings_file))
        manager = AutomationSettingsManager()
        assert manager.get_settings()["enable_automatic_knowledge_extraction"] is True
        assert settings_file.exists()
        assert manager.is_extraction_enabled() is True
        assert manager.is_automations_enabled() is True
        assert manager.is_accounting_enabled() is True
        assert manager.is_sales_enabled() is True
        assert manager.get_settings() is not manager._settings

    def test_loads_existing_file(self, tmp_path, monkeypatch):
        settings_file = tmp_path / "automation_settings.json"
        settings_file.write_text(json.dumps({"enable_automatic_knowledge_extraction": False}))
        monkeypatch.setattr(AutomationSettingsManager, "SETTINGS_FILE", str(settings_file))
        manager = AutomationSettingsManager()
        assert manager.is_extraction_enabled() is False

    def test_corrupt_file_falls_back_to_defaults(self, tmp_path, monkeypatch):
        settings_file = tmp_path / "automation_settings.json"
        settings_file.write_text("{not valid json")
        monkeypatch.setattr(AutomationSettingsManager, "SETTINGS_FILE", str(settings_file))
        manager = AutomationSettingsManager()
        assert manager.is_extraction_enabled() is True

    def test_update_settings_persists(self, tmp_path, monkeypatch):
        settings_file = tmp_path / "automation_settings.json"
        monkeypatch.setattr(AutomationSettingsManager, "SETTINGS_FILE", str(settings_file))
        manager = AutomationSettingsManager()
        manager.update_settings({"enable_accounting_automations": False})
        assert manager.is_accounting_enabled() is False
        assert json.loads(settings_file.read_text())["enable_accounting_automations"] is False

    def test_save_error_is_logged(self, tmp_path, monkeypatch):
        settings_file = tmp_path / "automation_settings.json"
        monkeypatch.setattr(AutomationSettingsManager, "SETTINGS_FILE", str(settings_file))
        manager = AutomationSettingsManager()
        real_open = open
        def failing_open(path, *args, **kwargs):
            if "w" in str(args) or kwargs.get("mode") == "w" or path == str(settings_file):
                raise OSError("disk full")
            return real_open(path, *args, **kwargs)
        with patch("builtins.open", side_effect=failing_open):
            manager.update_settings({"enable_sales_automations": False})
        assert manager.is_sales_enabled() is False

    def test_get_automation_settings_returns_global(self):
        from core.automation_settings import get_automation_settings
        assert get_automation_settings() is not None


# ============================================================================
# automation_insight_manager
# ============================================================================

class TestAutomationInsightManager:
    @pytest.fixture
    def manager(self, tmp_path):
        from core.automation_insight_manager import AutomationInsightManager
        db = tmp_path / "analytics.db"
        conn = sqlite3.connect(str(db))
        conn.execute(
            "CREATE TABLE workflow_events (workflow_id TEXT, event_type TEXT, "
            "timestamp TEXT, user_id TEXT)"
        )
        now = (datetime.now() - timedelta(days=1)).isoformat()
        rows = [
            ("wf-1", "step_completed", now, "user-1"),
            ("wf-1", "step_completed", now, "user-1"),
            ("wf-1", "manual_override", now, "user-1"),
            ("wf-1", "manual_override", now, "user-1"),
            ("wf-3", "manual_override", now, "user-1"),
        ]
        rows += [("wf-2", "step_completed", now, "user-1") for _ in range(11)]
        conn.executemany(
            "INSERT INTO workflow_events VALUES (?, ?, ?, ?)", rows
        )
        conn.commit()
        conn.close()
        return AutomationInsightManager(db_path=str(db))

    def test_drift_metrics_recommendations(self, manager):
        insights = manager.get_drift_metrics("user-1")
        by_id = {i["workflow_id"]: i for i in insights}
        assert by_id["wf-1"]["recommendation"] == "OPTIMIZE (High Overrides)"
        assert by_id["wf-1"]["drift_score"] == 1.0
        assert by_id["wf-2"]["recommendation"] == "HIGH_CONFIDENCE"
        assert by_id["wf-3"]["recommendation"] == "STABLE"

    def test_drift_metrics_filtered_by_workflow(self, manager):
        insights = manager.get_drift_metrics("user-1", workflow_id="wf-1")
        assert len(insights) == 1
        assert insights[0]["workflow_id"] == "wf-1"

    def test_drift_metrics_empty_user(self, manager):
        assert manager.get_drift_metrics("nobody") == []

    def test_drift_metrics_db_error_returns_empty(self, tmp_path):
        from core.automation_insight_manager import AutomationInsightManager
        manager = AutomationInsightManager(db_path=str(tmp_path / "missing.db"))
        with patch.object(manager, "db_path", tmp_path / "missing.db"):
            assert manager.get_drift_metrics("user-1") == []

    def test_underutilization_insights(self, manager):
        mock_session = MagicMock()
        row1 = MagicMock()
        row1.workflow_id = "wf-a"
        row1.execution_count = 1
        row2 = MagicMock()
        row2.workflow_id = "wf-b"
        row2.execution_count = 2
        row3 = MagicMock()
        row3.workflow_id = "wf-c"
        row3.execution_count = 5
        mock_session.query.return_value.filter.return_value.group_by.return_value.all.return_value = [row1, row2, row3]
        with patch("core.database.get_db_session") as mock_get_session:
            mock_get_session.return_value.__enter__.return_value = mock_session
            result = manager.get_underutilization_insights()
        assert len(result) == 2
        assert {r["workflow_id"] for r in result} == {"wf-a", "wf-b"}
        assert all(r["status"] == "UNDERUTILIZED" for r in result)

    def test_underutilization_insights_error(self, manager):
        with patch("core.database.get_db_session", side_effect=RuntimeError("boom")):
            assert manager.get_underutilization_insights() == []

    def test_generate_all_insights(self, manager):
        with patch.object(manager, "get_drift_metrics") as mock_drift:
            mock_drift.return_value = [
                {"workflow_id": "w1", "recommendation": "OPTIMIZE (High Overrides)"},
                {"workflow_id": "w2", "recommendation": "STABLE"},
            ]
            report = manager.generate_all_insights("user-1")
        assert report["summary"] == {
            "total_monitored": 2,
            "needs_optimization": 1,
            "stable": 1,
        }
        assert "timestamp" in report

    def test_get_insight_manager_singleton(self):
        import core.automation_insight_manager as aim
        original = aim._insight_manager
        try:
            aim._insight_manager = None
            instance = aim.get_insight_manager()
            assert aim.get_insight_manager() is instance
        finally:
            aim._insight_manager = original


# ============================================================================
# backfill_job_queue
# ============================================================================

class TestBackfillJobQueue:
    @pytest.fixture
    def redis_mock(self):
        redis = MagicMock()
        redis.hset = AsyncMock()
        redis.set = AsyncMock()
        redis.get = AsyncMock()
        redis.rpush = AsyncMock()
        redis.blpop = AsyncMock()
        redis.llen = AsyncMock(return_value=7)
        redis.delete = AsyncMock()
        redis.hgetall = AsyncMock(return_value={})
        redis.incr = AsyncMock()
        redis.expire = AsyncMock()
        redis.close = AsyncMock()
        return redis

    @pytest.fixture
    def queue(self, redis_mock):
        q = BackfillJobQueue(redis_url="redis://localhost:6379/0")
        q._client = redis_mock
        return q

    def test_default_retry_delays(self):
        q = BackfillJobQueue()
        assert q.retry_delays == [60, 300, 900, 3600]
        q2 = BackfillJobQueue(retry_delays=[1, 2])
        assert q2.retry_delays == [1, 2]

    @pytest.mark.asyncio
    async def test_get_client_creates_pool(self):
        with patch("core.backfill_job_queue.ConnectionPool.from_url") as mock_pool, \
                patch("core.backfill_job_queue.redis.Redis") as mock_redis_cls:
            q = BackfillJobQueue()
            client = await q.get_client()
        mock_pool.assert_called_once_with("redis://localhost:6379/0")
        mock_redis_cls.assert_called_once()
        assert client is mock_redis_cls.return_value

    @pytest.mark.asyncio
    async def test_get_client_reuses(self, queue, redis_mock):
        client = await queue.get_client()
        assert client is redis_mock

    @pytest.mark.asyncio
    async def test_close(self, queue, redis_mock):
        pool = MagicMock()
        pool.disconnect = AsyncMock()
        queue._pool = pool
        await queue.close()
        redis_mock.close.assert_awaited_once()
        pool.disconnect.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_schedule_entity_type_backfill(self, queue, redis_mock):
        job_id = await queue.schedule_entity_type_backfill(
            tenant_id="t1", slug="customer", display_name="Customer",
            json_schema={"type": "object"}, source="salesforce", ttl_hours=24,
        )
        assert "entity_type" in job_id
        redis_mock.hset.assert_awaited_once()
        redis_mock.rpush.assert_awaited_once()
        stored = redis_mock.hset.call_args.kwargs["mapping"]
        assert stored["job_type"] == "entity_type_backfill"
        assert stored["json_schema"] == '{"type": "object"}'

    @pytest.mark.asyncio
    async def test_schedule_node_migration(self, queue, redis_mock):
        job_id = await queue.schedule_node_migration("t1", "ws1", "entity", batch_size=500)
        assert "node_migration" in job_id
        stored = redis_mock.hset.call_args.kwargs["mapping"]
        assert stored["batch_size"] == 500

    @pytest.mark.asyncio
    async def test_schedule_ttl_cleanup(self, queue, redis_mock):
        job_id = await queue.schedule_ttl_cleanup("t1", interval_hours=2)
        assert "ttl_cleanup" in job_id
        stored = redis_mock.hset.call_args.kwargs["mapping"]
        assert stored["interval_hours"] == 2

    @pytest.mark.asyncio
    async def test_get_job_status_full(self, queue, redis_mock):
        redis_mock.get = AsyncMock(side_effect=[b"processing", b"3"])
        redis_mock.hgetall = AsyncMock(side_effect=[
            {b"job_type": b"entity_type_backfill", b"json_schema": b'{"type": "object"}',
             b"ttl_hours": b"48", b"available_skills": b'["a"]'},
            {b"progress": b"50", b"message": b"halfway"},
        ])
        status = await queue.get_job_status("job-1")
        assert status["status"] == "processing"
        assert status["retry_count"] == 3
        assert status["json_schema"] == {"type": "object"}
        assert status["available_skills"] == ["a"]
        assert status["ttl_hours"] == 48
        assert status["progress"] == {"progress": "50", "message": "halfway"}

    @pytest.mark.asyncio
    async     def test_get_job_status_unparsable_json(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={b"json_schema": b"{broken"})
        redis_mock.get = AsyncMock(return_value=b"0")
        status = await queue.get_job_status("job-1")
        assert status["json_schema"] == "{broken"

    @pytest.mark.asyncio
    async def test_update_job_progress(self, queue, redis_mock):
        await queue.update_job_progress("job-1", 42.5, "working")
        mapping = redis_mock.hset.call_args.kwargs["mapping"]
        assert mapping["progress"] == "42.5"
        assert mapping["message"] == "working"

    @pytest.mark.asyncio
    async def test_set_job_status(self, queue, redis_mock):
        await queue.set_job_status("job-1", BackfillJobStatus.PROCESSING)
        redis_mock.set.assert_awaited_with("job:status:job-1", "processing")

    @pytest.mark.asyncio
    async def test_process_job_success(self, queue, redis_mock):
        redis_mock.get = AsyncMock(return_value=b"0")
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"entity_type_backfill", b"json_schema": b'{"type": "object"}',
        })
        await queue.process_job_with_retry("job-1")
        assert redis_mock.set.await_args_list[-1].args[1] == "completed"

    @pytest.mark.asyncio
    async def test_process_job_retry_path(self, queue, redis_mock):
        redis_mock.get = AsyncMock(return_value=b"1")
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"node_migration", b"batch_size": b"0",
        })
        await queue.process_job_with_retry("job-1")
        assert redis_mock.incr.await_count == 1
        redis_mock.expire.assert_awaited_once()
        assert redis_mock.set.await_args_list[-1].args[1] == "retrying"

    @pytest.mark.asyncio
    async def test_process_job_dead_letter(self, queue, redis_mock):
        redis_mock.get = AsyncMock(return_value=b"4")
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"node_migration", b"batch_size": b"0",
        })
        queue.max_retries = 4
        await queue.process_job_with_retry("job-1")
        assert redis_mock.set.await_args_list[-1].args[1] == "dead_letter"

    @pytest.mark.asyncio
    async def test_execute_job_node_migration_valid(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"node_migration", b"batch_size": b"1000",
        })
        await queue._execute_job("job-1")

    @pytest.mark.asyncio
    async def test_execute_job_node_migration_invalid(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"node_migration", b"batch_size": b"10001",
        })
        with pytest.raises(ValueError):
            await queue._execute_job("job-1")

    @pytest.mark.asyncio
    async def test_execute_job_ttl_cleanup(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={b"job_type": b"ttl_cleanup"})
        await queue._execute_job("job-1")

    @pytest.mark.asyncio
    async def test_execute_job_unknown_type(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={b"job_type": b"mystery"})
        await queue._execute_job("job-1")

    @pytest.mark.asyncio
    async def test_execute_job_invalid_schema(self, queue, redis_mock):
        redis_mock.hgetall = AsyncMock(return_value={
            b"job_type": b"entity_type_backfill", b"json_schema": b"not json {",
        })
        with pytest.raises(ValueError):
            await queue._execute_job("job-1")

    @pytest.mark.asyncio
    async def test_get_next_job(self, queue, redis_mock):
        redis_mock.blpop = AsyncMock(return_value=[b"job:queue:t1", b"job-1"])
        assert await queue.get_next_job("t1") == "job-1"
        redis_mock.blpop = AsyncMock(return_value=None)
        assert await queue.get_next_job("t1") is None

    @pytest.mark.asyncio
    async def test_get_queue_size(self, queue, redis_mock):
        assert await queue.get_queue_size("t1") == 7

    @pytest.mark.asyncio
    async def test_clear_queue(self, queue, redis_mock):
        await queue.clear_queue("t1")
        redis_mock.delete.assert_awaited_with("job:queue:t1")

    def test_singleton(self):
        with patch.dict(os.environ, {"REDIS_URL": "redis://localhost:9999/0"}):
            with patch("core.backfill_job_queue._job_queue", None):
                q = get_backfill_job_queue()
        assert q is not None


# ============================================================================
# alert_service
# ============================================================================

class TestAlertThresholdServiceCoverage:
    @pytest.fixture
    def service(self):
        return AlertThresholdService(Mock())

    def _config(self, **overrides):
        config = Mock()
        config.tenant_id = "t1"
        config.connector_id = "slack"
        config.window_seconds = 300
        config.error_rate_threshold = 10.0
        config.latency_threshold_ms = 500
        config.notification_channels = []
        config.slack_channel_id = None
        config.email_recipients = None
        config.is_active = True
        for k, v in overrides.items():
            setattr(config, k, v)
        return config

    def _metrics(self, successes=0, failures=0, p95=0.0):
        metrics = Mock()
        metrics.success_counts = {"slack:t1:agent": successes}
        metrics.failure_counts = {"slack:t1:agent": failures}
        metrics._make_key = Mock(return_value="slack:t1:agent")
        metrics.get_duration_percentiles = Mock(return_value={"p95": p95})
        return metrics

    def test_constructs_with_stub_config(self):
        service = AlertThresholdService(Mock())
        assert service.AlertConfiguration.is_active is True

    def test_evaluate_error_rate_no_config(self, service):
        service.db.query.return_value.filter.return_value.first.return_value = None
        assert service.evaluate_error_rate_threshold("t1", "slack") is None

    def test_evaluate_error_rate_severity_critical(self, service):
        config = self._config()
        with patch("core.integration_metrics.get_integration_metrics", return_value=self._metrics(successes=70, failures=30)):
            violation = service.evaluate_error_rate_threshold("t1", "slack", config)
        assert violation is not None
        assert violation.severity == AlertSeverity.CRITICAL
        assert violation.actual_value == 30.0

    def test_evaluate_error_rate_hysteresis_clears(self, service):
        config = self._config()
        service._redis_state = {}
        with patch("core.integration_metrics.get_integration_metrics", return_value=self._metrics(successes=96, failures=4)):
            with patch.object(service, "_get_alert_state", return_value="violated"):
                with patch.object(service, "_set_alert_state") as mock_set:
                    violation = service.evaluate_error_rate_threshold("t1", "slack", config)
        assert violation is None
        mock_set.assert_called_once_with("t1", "slack", "error_rate", "cleared")

    def test_evaluate_latency_no_config(self, service):
        service.db.query.return_value.filter.return_value.first.return_value = None
        assert service.evaluate_latency_threshold("t1", "slack") is None

    def test_evaluate_latency_no_threshold(self, service):
        config = self._config(latency_threshold_ms=None)
        service.db.query.return_value.filter.return_value.first.return_value = config
        assert service.evaluate_latency_threshold("t1", "slack") is None

    def test_evaluate_latency_violation(self, service):
        config = self._config(latency_threshold_ms=500)
        with patch("core.integration_metrics.get_integration_metrics", return_value=self._metrics(p95=900.0)):
            violation = service.evaluate_latency_threshold("t1", "slack", config)
        assert violation is not None
        assert violation.metric_type == "latency_p95"
        assert violation.severity == AlertSeverity.WARNING

    def test_evaluate_all_thresholds_groups(self, service):
        c1 = self._config()
        c2 = self._config(connector_id="slack")
        c3 = self._config(connector_id="gmail")
        query = service.db.query.return_value.filter.return_value
        query.filter.return_value.all.return_value = [c1, c2, c3]
        with patch("core.integration_metrics.get_integration_metrics", return_value=self._metrics(successes=0, failures=0)):
            results = service.evaluate_all_thresholds(tenant_id="t1")
        assert len(results) == 2
        for r in results:
            assert isinstance(r, AlertEvaluationResult)

    def test_evaluate_all_no_tenant_filter(self, service):
        service.db.query.return_value.filter.return_value.all.return_value = [self._config()]
        with patch("core.integration_metrics.get_integration_metrics", return_value=self._metrics()):
            results = service.evaluate_all_thresholds()
        assert len(results) == 1

    def test_calculate_error_rate_in_window_zero_total(self, service):
        metrics = self._metrics(successes=0, failures=0)
        assert service._calculate_error_rate_in_window(metrics, "t1", "slack", datetime.now(timezone.utc)) == 0.0

    def test_alert_state_redis_get_set(self):
        redis = Mock()
        redis.get = Mock(return_value=b"violated")
        service = AlertThresholdService(Mock(), redis_client=redis)
        assert service._get_alert_state("t1", "slack", "error_rate") == "violated"
        service._set_alert_state("t1", "slack", "error_rate", "violated")
        redis.setex.assert_called_once()

    def test_alert_state_no_redis(self, service):
        assert service._get_alert_state("t1", "slack", "error_rate") == "ok"
        service._set_alert_state("t1", "slack", "error_rate", "ok")

    def test_get_violations_for_tenant(self, service):
        service.evaluate_all_thresholds = Mock(return_value=[
            AlertEvaluationResult(
                tenant_id="t1", connector_id="slack", status=AlertStatus.VIOLATED,
                violations=[Mock()], evaluated_at=datetime.now(timezone.utc),
            )
        ])
        violations = service.get_violations_for_tenant("t1")
        assert len(violations) == 1

    @pytest.mark.asyncio
    async def test_send_notifications_no_channels(self, service):
        config = self._config(notification_channels=[])
        assert await service.send_notifications(Mock(), config) == {}

    @pytest.mark.asyncio
    async def test_send_notifications_slack_only(self, service):
        config = self._config(notification_channels=["slack"], slack_channel_id="C1")
        violation = AlertViolation(
            tenant_id="t1", connector_id="slack", metric_type="error_rate",
            actual_value=15.0, threshold=10.0, severity=AlertSeverity.WARNING,
            timestamp=datetime.now(timezone.utc),
            window_start=datetime.now(timezone.utc), window_end=datetime.now(timezone.utc),
        )
        with patch.object(service, "send_slack_notification", new=AsyncMock(return_value=True)) as mock_slack:
            results = await service.send_notifications(violation, config)
        assert results == {"slack": True}

    @pytest.mark.asyncio
    async def test_send_slack_no_token(self, service):
        violation = Mock(tenant_id="t1")
        config = self._config(slack_channel_id="C1")
        with patch("core.token_storage.token_storage.get_token", return_value=None):
            assert await service.send_slack_notification(violation, config) is False

    @pytest.mark.asyncio
    async def test_send_slack_exception(self, service):
        violation = Mock(tenant_id="t1", metric_type="error_rate")
        config = self._config(slack_channel_id="C1")
        with patch("core.token_storage.token_storage.get_token", side_effect=RuntimeError("boom")):
            assert await service.send_slack_notification(violation, config) is False

    @pytest.mark.asyncio
    async def test_send_email_no_recipients(self, service):
        violation = Mock(tenant_id="t1")
        config = self._config(notification_channels=["email"], email_recipients=[])
        results = await service.send_notifications(violation, config)
        assert results == {}

    @pytest.mark.asyncio
    async def test_send_email_exception(self, service):
        violation = Mock(tenant_id="t1", metric_type="error_rate", severity=AlertSeverity.CRITICAL,
                         connector_id="slack", actual_value=1.0, threshold=0.5,
                         timestamp=datetime.now(timezone.utc))
        config = self._config(email_recipients=["a@b.c"])
        with patch("integrations.email_routes.EmailService") as email_cls:
            email_cls.side_effect = RuntimeError("smtp down")
            assert await service.send_email_notification(violation, config) is False

    @pytest.mark.asyncio
    async def test_send_email_success(self, service):
        violation = Mock(tenant_id="t1", metric_type="error_rate", severity=AlertSeverity.INFO,
                         connector_id="slack", actual_value=1.0, threshold=0.5,
                         timestamp=datetime.now(timezone.utc))
        email_service = Mock()
        email_service.send_email = AsyncMock(return_value=True)
        with patch.dict("sys.modules", {"core.email_service": MagicMock(EmailService=Mock(return_value=email_service))}):
            assert await service.send_email_notification(violation, self._config(email_recipients=["a@b.c"])) is True

    @pytest.mark.asyncio
    async def test_send_alert_cleared_no_channels(self, service):
        assert await service.send_alert_cleared_notification("t1", "slack", "error_rate", self._config()) is False

    @pytest.mark.asyncio
    async def test_send_alert_cleared_slack_failure(self, service):
        config = self._config(notification_channels=["slack"], slack_channel_id="C1")
        with patch.dict("sys.modules", {
            "integrations.slack_enhanced_service": MagicMock(),
            "core.token_storage": MagicMock(),
        }):
            from integrations.slack_enhanced_service import SlackEnhancedService
            SlackEnhancedService.return_value.send_message = AsyncMock(side_effect=RuntimeError("fail"))
            assert await service.send_alert_cleared_notification("t1", "slack", "error_rate", config) is False

    @pytest.mark.asyncio
    async def test_send_alert_cleared_email(self, service):
        config = self._config(notification_channels=["email"], email_recipients=["a@b.c"])
        email_service = Mock()
        email_service.send_email = AsyncMock(return_value=True)
        with patch.dict("sys.modules", {"core.email_service": MagicMock(EmailService=Mock(return_value=email_service))}):
            assert await service.send_alert_cleared_notification("t1", "slack", "error_rate", config) is True

    @pytest.mark.asyncio
    async def test_send_alert_cleared_exception(self, service):
        with patch.dict("sys.modules", {"core.email_service": MagicMock()}):
            from core.email_service import EmailService
            EmailService.side_effect = RuntimeError("boom")
            assert await service.send_alert_cleared_notification("t1", "slack", "error_rate", self._config()) is False

    def test_format_helpers(self, service):
        violation = AlertViolation(
            tenant_id="t1", connector_id="slack", metric_type="error_rate",
            actual_value=15.0, threshold=10.0, severity=AlertSeverity.WARNING,
            timestamp=datetime.now(timezone.utc),
            window_start=datetime.now(timezone.utc), window_end=datetime.now(timezone.utc),
        )
        msg = service._format_slack_message(violation, self._config())
        assert "Alert Violation Detected" in msg
        assert service._get_emoji_for_severity(AlertSeverity.CRITICAL) == ":rotating_light:"
        assert service._get_emoji_for_severity("unknown") == ":warning:"
        subject = service._format_email_subject(violation)
        assert "Alert" in subject
        html = service._format_email_html(violation, self._config())
        assert "<html>" in html and "#ffc107" in html

    @pytest.mark.asyncio
    async def test_check_and_send_cleared_alerts_no_redis(self, service):
        assert await service.check_and_send_cleared_alerts("t1", "slack") is None

    @pytest.mark.asyncio
    async def test_check_and_send_cleared_alerts_flow(self, service):
        redis = Mock()
        redis.get = Mock(return_value=b"cleared")
        service.redis = redis
        config = self._config()
        query = service.db.query.return_value.filter.return_value
        query.first.return_value = config
        with patch.object(service, "send_alert_cleared_notification", new=AsyncMock(return_value=True)):
            await service.check_and_send_cleared_alerts("t1", "slack")
        assert redis.setex.call_count == 2

    @pytest.mark.asyncio
    async def test_check_and_send_cleared_alerts_no_config(self, service):
        redis = Mock()
        redis.get = Mock(return_value=b"cleared")
        service.redis = redis
        query = service.db.query.return_value.filter.return_value
        query.first.return_value = None
        await service.check_and_send_cleared_alerts("t1", "slack")
        assert redis.setex.call_count == 2


# ============================================================================
# ai_workflow_optimizer
# ============================================================================

class TestAIWorkflowOptimizer:
    @pytest.fixture
    def optimizer(self):
        return AIWorkflowOptimizer()

    @pytest.fixture
    def workflow(self):
        return {
            "id": "wf-1",
            "name": "Test Workflow",
            "nodes": [
                {"id": "n1", "type": "trigger", "config": {}},
                {"id": "n2", "type": "action", "config": {"integration": "salesforce"}},
                {"id": "n3", "type": "action", "config": {"integration": "salesforce"}},
                {"id": "n4", "type": "action", "config": {"integration": "salesforce"}},
                {"id": "n5", "type": "action", "config": {"integration": "openai", "error_handling": True}},
                {"id": "n6", "type": "condition", "config": {"integration": "openai", "error_handling": True}},
                {"id": "n7", "type": "action", "config": {"integration": "openai", "error_handling": True}},
                {"id": "n8", "type": "action", "config": {"integration": "openai", "error_handling": True}},
                {"id": "n9", "type": "action", "config": {"integration": "openai", "error_handling": True}},
                {"id": "n10", "type": "action", "config": {"integration": "gmail", "batch_size": 2000}},
                {"id": "n11", "type": "action", "config": {"integration": "unknown_int"}, "label": "Approval needed"},
            ],
            "edges": [{"source": "n1", "target": "n2"}],
        }

    @pytest.mark.asyncio
    async def test_analyze_workflow_full(self, optimizer, workflow):
        analysis = await optimizer.analyze_workflow(workflow, {"success_rate": 0.9, "avg_execution_time": 5.0})
        assert isinstance(analysis, WorkflowAnalysis)
        assert analysis.total_nodes == 11
        assert "salesforce" in analysis.integrations_used
        assert 0 <= analysis.complexity_score <= 100
        assert analysis.estimated_execution_time == 5.0

    @pytest.mark.asyncio
    async def test_analyze_workflow_minimal(self, optimizer):
        analysis = await optimizer.analyze_workflow({})
        assert analysis.workflow_id == "unknown"
        assert analysis.workflow_name == "Unnamed Workflow"
        assert analysis.estimated_execution_time == 0.0

    @pytest.mark.asyncio
    async def test_optimize_workflow_plan_goals(self, optimizer, workflow):
        plan = await optimizer.optimize_workflow_plan(workflow, [OptimizationType.PERFORMANCE, OptimizationType.COST])
        assert plan["optimization_plan"]["goals"] == ["performance", "cost"]
        assert "phases" in plan["optimization_plan"]
        assert "estimated_total_improvement" in plan["optimization_plan"]

    @pytest.mark.asyncio
    async def test_monitor_performance_urgent_recommendations(self, optimizer):
        result = await optimizer.monitor_workflow_performance(
            "wf-1",
            {"success_rate": 0.5},
            time_window=48,
        )
        assert result["workflow_id"] == "wf-1"
        assert result["time_window_hours"] == 48
        assert result["health_score"] == 100

    def test_extract_integrations(self, optimizer):
        nodes = [{"config": {"integration": "slack"}}, {"config": {}}, {"config": {"integration": "slack"}}]
        assert sorted(optimizer._extract_integrations(nodes)) == ["slack"]

    def test_calculate_complexity_score(self, optimizer, workflow):
        score = optimizer._calculate_complexity_score(workflow)
        assert score == 39.0

    def test_estimate_execution_time_fallback(self, optimizer):
        assert optimizer._estimate_execution_time({}) == 0.0

    def test_estimate_execution_time_condition_and_other(self, optimizer):
        workflow = {"nodes": [
            {"type": "condition", "config": {}},
            {"type": "mystery", "config": {}},
            {"type": "action", "config": {}},
        ]}
        assert optimizer._estimate_execution_time(workflow) == 0.1 + 0.5 + 1.0

    def test_identify_failure_points_low_rate_limit(self, optimizer):
        workflow = {"nodes": [
            {"id": "n1", "type": "action", "config": {"integration": "openai", "error_handling": True, "model": "test-model"}},
        ]}
        points = optimizer._identify_failure_points(workflow)
        assert points[0]["risk_level"] == "medium"

    def test_identify_bottlenecks(self, optimizer):
        workflow = {
            "nodes": [{"id": "n1", "type": "action", "config": {"batch_size": 2000}}],
            "edges": [],
        }
        bottlenecks = optimizer._identify_bottlenecks(workflow)
        assert any(b["type"] == "data_processing" for b in bottlenecks)

    def test_identify_bottlenecks_long_path(self, optimizer):
        nodes = [{"id": f"n{i}", "config": {}} for i in range(8)]
        edges = [{"source": f"n{i}", "target": f"n{i+1}"} for i in range(7)]
        bottlenecks = optimizer._identify_bottlenecks({"nodes": nodes, "edges": edges})
        assert any(b["type"] == "sequential_depth" for b in bottlenecks)

    @pytest.mark.asyncio
    async def test_generate_recommendations_all_types(self, optimizer, workflow):
        recs = await optimizer._generate_recommendations(workflow, {})
        types = {r.type for r in recs}
        assert OptimizationType.PERFORMANCE in types
        assert OptimizationType.COST in types
        assert OptimizationType.RELIABILITY in types
        assert OptimizationType.EFFICIENCY in types

    def test_recommendation_helpers(self, optimizer, workflow):
        data = {"workflow": workflow}
        assert optimizer._count_sequential_api_calls(data) == 10
        assert optimizer._has_large_data_processing(data) is True
        assert optimizer._has_frequent_ai_calls(data) is True
        assert optimizer._has_single_points_of_failure(data) is True
        assert optimizer._lacks_error_handling(data) is True
        assert optimizer._has_manual_bottlenecks(data) is True
        assert optimizer._has_redundant_validations(data) is False
        assert optimizer._has_underutilized_premium_integrations(data) is False
        assert optimizer._has_unnecessary_transformations(data) is False

    def test_priority_score(self, optimizer):
        rec = OptimizationRecommendation(
            id="r1", type=OptimizationType.PERFORMANCE, title="t", description="d",
            impact_level=ImpactLevel.HIGH, estimated_improvement={"x": 1.0},
            implementation_effort="medium", steps=[], prerequisites=[], risks=[],
            confidence_score=50,
        )
        assert optimizer._calculate_priority_score(rec) > 0

    def test_implementation_phases(self, optimizer):
        recs = [
            OptimizationRecommendation(id=f"r{i}", type=OptimizationType.COST, title="t", description="d",
                                       impact_level=ImpactLevel.MEDIUM, estimated_improvement={},
                                       implementation_effort=effort, steps=[], prerequisites=[], risks=[])
            for i, effort in enumerate(["easy", "easy", "easy", "easy", "medium", "medium", "medium",
                                        "medium", "medium", "complex", "complex", "complex"])
        ]
        phases = optimizer._create_implementation_phases(recs, {"budget": 100})
        assert len(phases) == 3
        assert len(phases[0]["recommendations"]) == 3
        assert len(phases[1]["recommendations"]) == 5
        assert len(phases[2]["recommendations"]) == 3

    def test_total_improvement_and_timeline(self, optimizer):
        recs = [
            OptimizationRecommendation(id="r1", type=OptimizationType.COST, title="t", description="d",
                                       impact_level=ImpactLevel.MEDIUM, estimated_improvement={"cost": 10.0},
                                       implementation_effort="easy", steps=[], prerequisites=[], risks=[]),
            OptimizationRecommendation(id="r2", type=OptimizationType.COST, title="t", description="d",
                                       impact_level=ImpactLevel.MEDIUM, estimated_improvement={"cost": 5.0},
                                       implementation_effort="complex", steps=[], prerequisites=[], risks=[]),
        ]
        assert optimizer._calculate_total_improvement(recs) == {"cost": 15.0}
        assert optimizer._estimate_implementation_timeline(recs) == "8 days"

    def test_performance_trends_and_health(self, optimizer):
        trends = optimizer._analyze_performance_trends({}, 24)
        assert trends["execution_time"] == "stable"
        issues = optimizer._identify_performance_issues({"success_rate": "declining"})
        assert issues[0]["severity"] == "high"
        assert optimizer._calculate_health_score({}, [{"severity": "critical"}, {"severity": "high"}, {"severity": "medium"}]) == 40
        assert optimizer._calculate_health_score({}, [{"severity": "low"}]) == 100

    @pytest.mark.asyncio
    async def test_generate_urgent_recommendation(self, optimizer):
        rec = await optimizer._generate_urgent_recommendation({"type": "x", "description": "d"}, {})
        assert rec.id == "urgent_x"
        assert rec.impact_level == ImpactLevel.CRITICAL

    def test_singleton(self):
        optimizer = get_ai_workflow_optimizer()
        assert get_ai_workflow_optimizer() is optimizer

    @pytest.mark.asyncio
    async def test_rule_error_swallowed(self, optimizer):
        optimizer.optimization_rules[OptimizationType.PERFORMANCE] = [
            {
                "pattern": "x",
                "condition": lambda data: (_ for _ in ()).throw(RuntimeError("rule boom")),
                "recommendation": lambda data, rule: None,
                "impact": ImpactLevel.LOW,
                "improvement": {},
            }
        ]
        recs = await optimizer._generate_recommendations({}, {})
        assert isinstance(recs, list)


# ============================================================================
# ai_workflow_optimization_endpoints
# ============================================================================

class TestOptimizationEndpoints:
    @pytest.fixture
    def optimizer(self):
        optimizer = Mock()
        optimizer.analyze_workflow = AsyncMock()
        optimizer.optimize_workflow_plan = AsyncMock()
        optimizer.monitor_workflow_performance = AsyncMock()
        return optimizer

    @pytest.fixture
    def client(self, optimizer):
        app = FastAPI()
        app.include_router(optimizer_router)
        app.dependency_overrides[get_ai_workflow_optimizer] = lambda: optimizer
        return TestClient(app)

    @pytest.fixture
    def analysis(self):
        rec = Mock()
        rec.id = "parallel"
        rec.type = OptimizationType.PERFORMANCE
        rec.title = "Parallelize"
        rec.impact_level = ImpactLevel.HIGH
        rec.estimated_improvement = {"execution_time": 40}
        rec.implementation_effort = "medium"
        rec.confidence_score = 85
        rec.description = "desc"
        analysis = Mock()
        analysis.workflow_id = "wf-1"
        analysis.workflow_name = "Test"
        analysis.total_nodes = 3
        analysis.total_edges = 2
        analysis.complexity_score = 12.5
        analysis.estimated_execution_time = 3.0
        analysis.integrations_used = ["slack"]
        analysis.failure_points = [{"node_id": "n1", "issues": ["no error handling"], "risk_level": "high"}]
        analysis.bottlenecks = [{"type": "sequential_depth"}]
        analysis.optimization_opportunities = [rec]
        analysis.analysis_timestamp = datetime.now(timezone.utc)
        return analysis

    def test_analyze_workflow_endpoint(self, client, optimizer, analysis):
        optimizer.analyze_workflow.return_value = analysis
        response = client.post("/api/v1/workflows/analyze", json={"workflow_data": {"id": "wf-1"}})
        assert response.status_code == 200
        body = response.json()
        assert body["analysis"]["risk_assessment"]["risk_level"] == "high"
        assert body["analysis"]["optimization_opportunities"] == 1
        assert body["analysis"]["top_recommendations"][0]["type"] == "performance"

    def test_analyze_workflow_500(self, client, optimizer):
        optimizer.analyze_workflow.side_effect = RuntimeError("boom")
        response = client.post("/api/v1/workflows/analyze", json={"workflow_data": {}})
        assert response.status_code == 500
        assert response.json()["detail"] == "Internal error"

    def test_optimization_plan_invalid_goal_400(self, client):
        response = client.post(
            "/api/v1/workflows/optimization-plan",
            json={"workflow_data": {}, "optimization_goals": ["bogus"]},
        )
        assert response.status_code == 400

    def test_optimization_plan_success(self, client, optimizer, analysis):
        optimizer.optimize_workflow_plan.return_value = {
            "optimization_plan": {"goals": ["performance"], "phases": []},
            "workflow_analysis": {
                "workflow_id": "wf-1", "workflow_name": "Test", "complexity_score": 12.5,
                "failure_points": analysis.failure_points,
                "optimization_opportunities": analysis.optimization_opportunities,
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        response = client.post(
            "/api/v1/workflows/optimization-plan",
            json={"workflow_data": {}, "optimization_goals": ["performance"]},
        )
        assert response.status_code == 200
        body = response.json()
        assert body["recommendations_by_type"]["performance"][0]["id"] == "parallel"

    def test_optimization_plan_500(self, client, optimizer):
        optimizer.optimize_workflow_plan.side_effect = RuntimeError("boom")
        response = client.post(
            "/api/v1/workflows/optimization-plan",
            json={"workflow_data": {}, "optimization_goals": ["performance"]},
        )
        assert response.status_code == 500

    def test_monitor_endpoint_workflow_id_override(self, client, optimizer):
        optimizer.monitor_workflow_performance.return_value = {
            "health_score": 50, "urgent_recommendations": [], "identified_issues": [],
        }
        response = client.post(
            "/api/v1/workflows/path-wf/monitor",
            json={"workflow_id": "body-wf", "metrics": {}},
        )
        assert response.status_code == 200
        assert optimizer.monitor_workflow_performance.await_args.args[0] == "path-wf"
        assert response.json()["health_status"]["status"] == "critical"

    def test_monitor_endpoint_500(self, client, optimizer):
        optimizer.monitor_workflow_performance.side_effect = RuntimeError("boom")
        response = client.post("/api/v1/workflows/wf/monitor", json={"workflow_id": "wf", "metrics": {}})
        assert response.status_code == 500

    def test_recommendations_with_filters(self, client):
        response = client.get(
            "/api/v1/workflows/wf-1/recommendations",
            params={"type_filter": "performance", "impact_filter": "high"},
        )
        assert response.status_code == 200
        assert response.json()["total_recommendations"] == 1

    def test_recommendations_no_match(self, client):
        response = client.get(
            "/api/v1/workflows/wf-1/recommendations",
            params={"type_filter": "scalability"},
        )
        assert response.json()["total_recommendations"] == 0

    def test_optimization_types_endpoint(self, client):
        response = client.get("/api/v1/workflows/optimization-types")
        assert response.status_code == 200
        assert "security" in response.json()["optimization_types"]

    def test_batch_analysis(self, client, optimizer, analysis):
        optimizer.analyze_workflow.return_value = analysis
        response = client.post("/api/v1/workflows/batch-analysis", json=[
            {"id": "wf-1"}, {"id": "wf-2"},
        ])
        assert response.status_code == 200
        body = response.json()["batch_analysis"]
        assert body["summary"]["total_workflows"] == 2
        assert body["summary"]["total_recommendations"] == 2
        assert body["summary"]["common_issues"]["no error handling"] == 2
        assert body["summary"]["optimization_priorities"]["performance"] == 2
        assert body["workflow_results"][0]["top_priority"] == "performance"

    def test_batch_analysis_too_many(self, client):
        response = client.post(
            "/api/v1/workflows/batch-analysis",
            json=[{"id": f"wf-{i}"} for i in range(51)],
        )
        assert response.status_code == 400

    def test_batch_analysis_partial_failure(self, client, optimizer, analysis):
        optimizer.analyze_workflow = AsyncMock(side_effect=[analysis, RuntimeError("boom")])
        response = client.post("/api/v1/workflows/batch-analysis", json=[{"id": "wf-1"}, {"id": "wf-2"}])
        assert response.status_code == 200
        results = response.json()["batch_analysis"]["workflow_results"]
        assert results[1]["error"] == "boom"

    def test_optimization_insights_endpoint(self, client):
        response = client.get("/api/v1/workflows/optimization-insights", params={"time_range": "30d"})
        assert response.status_code == 200
        assert response.json()["time_range"] == "30d"

    def test_implement_optimization(self, client):
        with patch("core.ai_workflow_optimization_endpoints._execute_optimization_implementation") as mock_exec:
            response = client.post(
                "/api/v1/workflows/wf-1/implement-optimization",
                json={"optimization_id": "parallel"},
            )
        assert response.status_code == 200
        assert response.json()["status"] == "initiated"
        mock_exec.assert_called_once()

    @pytest.mark.asyncio
    async def test_execute_optimization_implementation(self):
        from core.ai_workflow_optimization_endpoints import _execute_optimization_implementation
        db = MagicMock()
        with patch("core.database.SessionLocal", return_value=db), \
                patch("asyncio.sleep", new=AsyncMock()):
            await _execute_optimization_implementation("job-1", "wf-1", "caching")
        db.close.assert_called_once()

    @pytest.mark.asyncio
    async def test_execute_optimization_implementation_error(self):
        from core.ai_workflow_optimization_endpoints import _execute_optimization_implementation
        with patch("core.database.SessionLocal", side_effect=RuntimeError("boom")), \
                patch("asyncio.sleep", new=AsyncMock()):
            await _execute_optimization_implementation("job-1", "wf-1", "caching")

    def test_calculate_risk_level(self):
        from core.ai_workflow_optimization_endpoints import _calculate_risk_level
        analysis = Mock()
        analysis.failure_points = [{"risk_level": "high"}]
        analysis.bottlenecks = []
        assert _calculate_risk_level(analysis) == "high"
        analysis.failure_points = [{"risk_level": "medium"}, {"risk_level": "medium"}, {"risk_level": "medium"}]
        analysis.bottlenecks = []
        assert _calculate_risk_level(analysis) == "medium"
        analysis.failure_points = [{"risk_level": "medium"}]
        assert _calculate_risk_level(analysis) == "low"
        analysis.failure_points = [{"risk_level": "medium"}]
        analysis.bottlenecks = [{"x": 1}, {"x": 2}, {"x": 3}, {"x": 4}, {"x": 5}]
        assert _calculate_risk_level(analysis) == "high"

    def test_group_recommendations_by_type(self):
        from core.ai_workflow_optimization_endpoints import _group_recommendations_by_type
        recs = [
            OptimizationRecommendation(
                id="r1", type=OptimizationType.PERFORMANCE, title="t1", description="d",
                impact_level=ImpactLevel.HIGH, estimated_improvement={},
                implementation_effort="easy", steps=[], prerequisites=[], risks=[],
            ),
            OptimizationRecommendation(
                id="r2", type=OptimizationType.COST, title="t2", description="d",
                impact_level=ImpactLevel.LOW, estimated_improvement={},
                implementation_effort="easy", steps=[], prerequisites=[], risks=[],
            ),
        ]
        grouped = _group_recommendations_by_type(recs)
        assert set(grouped.keys()) == {"performance", "cost"}
        assert grouped["performance"][0]["impact_level"] == "high"


# ============================================================================
# atom_saas_client
# ============================================================================

class FakeResponse:
    def __init__(self, status_code: int = 200, data: Any = None):
        self.status_code = status_code
        self._data = data if data is not None else {}
        self.content = b"data"

    def raise_for_status(self):
        if self.status_code >= 400:
            import httpx
            request = httpx.Request("GET", "http://test")
            raise httpx.HTTPStatusError(
                f"Error {self.status_code}", request=request, response=httpx.Response(self.status_code, request=request)
            )

    def json(self):
        return self._data


class TestAtomSaaSClient:
    @pytest.fixture
    def client(self):
        return AtomSaaSClient(AtomSaaSConfig(
            ws_url="wss://example.com/ws",
            api_url="https://example.com/api/v1/marketplace",
            api_token="tok-123",
            instance_id="inst-1",
        ))

    @pytest.fixture
    def http_mock(self):
        with patch("core.atom_saas_client.httpx.AsyncClient") as mock_cls:
            instance = MagicMock()
            instance.get = AsyncMock()
            instance.post = AsyncMock()
            instance.delete = AsyncMock()
            instance.aclose = AsyncMock()
            mock_cls.return_value = instance
            yield mock_cls, instance

    @pytest.mark.asyncio
    async def test_load_config_from_env(self):
        with patch.dict(os.environ, {
            "ATOM_SAAS_URL": "wss://x/ws",
            "ATOM_SAAS_API_URL": "https://x/api",
            "ATOM_SAAS_API_TOKEN": "secret-token",
        }, clear=True):
            from core.atom_saas_client import AtomAgentOSMarketplaceClient
            c = AtomAgentOSMarketplaceClient()
        assert c.config.api_token == "secret-token"
        assert c.config.instance_id is not None

    def test_load_config_no_token_warns(self):
        with patch.dict(os.environ, {"ATOM_SAAS_API_TOKEN": ""}, clear=True):
            from core.atom_saas_client import AtomAgentOSMarketplaceClient
            c = AtomAgentOSMarketplaceClient()
        assert c.config.api_token == ""

    @pytest.mark.asyncio
    async def test_fetch_skills_success(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"skills": [{"id": "s1"}]})
        result = await client.fetch_skills(query="q", category="cat", skill_type="official", page=2, page_size=10)
        assert result["skills"][0]["id"] == "s1"
        assert instance.get.call_args.args[0] == "/skills/marketplace/skills"
        params = instance.get.call_args.kwargs["params"]
        assert params["category"] == "cat" and params["skill_type"] == "official"

    @pytest.mark.asyncio
    async def test_fetch_skills_error(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(500)
        result = await client.fetch_skills()
        assert result == {"skills": [], "total": 0, "page": 1, "page_size": 20}

    @pytest.mark.asyncio
    async def test_get_skill_by_id(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"id": "s1"})
        assert await client.get_skill_by_id("s1") == {"id": "s1"}
        instance.get.return_value = FakeResponse(500)
        assert await client.get_skill_by_id("s2") is None

    @pytest.mark.asyncio
    async def test_get_categories(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, [{"name": "c1"}])
        assert await client.get_categories() == [{"name": "c1"}]
        instance.get.return_value = FakeResponse(500)
        assert await client.get_categories() == []

    @pytest.mark.asyncio
    async def test_rate_skill_invalid_rating(self, client, http_mock):
        result = await client.rate_skill("s1", "u1", 9)
        assert result["success"] is False

    @pytest.mark.asyncio
    async def test_rate_skill_success_and_error(self, client, http_mock):
        _, instance = http_mock
        instance.post.return_value = FakeResponse(200, {"success": True})
        result = await client.rate_skill("s1", "u1", 5, "great")
        assert result["success"] is True
        payload = instance.post.call_args.kwargs["json"]
        assert payload["rating"] == 5 and payload["comment"] == "great"
        instance.post.return_value = FakeResponse(500)
        result = await client.rate_skill("s1", "u1", 5)
        assert result["success"] is False

    @pytest.mark.asyncio
    async def test_install_uninstall_skill(self, client, http_mock):
        _, instance = http_mock
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert (await client.install_skill("s1", "a1", auto_install_deps=False))["success"] is True
        instance.delete.return_value = FakeResponse(200, {"success": True})
        assert (await client.uninstall_skill("s1", "a1"))["success"] is True
        instance.delete.return_value = FakeResponse(500)
        assert (await client.uninstall_skill("s1", "a1"))["success"] is False

    @pytest.mark.asyncio
    async def test_fetch_agents(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"agents": [{"id": "a1"}], "total": 1})
        result = await client.fetch_agents(query="q", category="c", page=1, page_size=10)
        assert result["agents"][0]["id"] == "a1"
        instance.get.return_value = FakeResponse(500)
        result = await client.fetch_agents()
        assert result == {"agents": [], "total": 0, "page": 1, "page_size": 20}

    @pytest.mark.asyncio
    async def test_get_agent_template(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"id": "t1"})
        assert await client.get_agent_template("t1") == {"id": "t1"}
        instance.get.return_value = FakeResponse(500)
        assert await client.get_agent_template("t1") is None

    @pytest.mark.asyncio
    async def test_install_agent(self, client, http_mock):
        _, instance = http_mock
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert (await client.install_agent("t1", "tenant-1"))["success"] is True
        instance.post.return_value = FakeResponse(500)
        assert (await client.install_agent("t1", "tenant-1"))["success"] is False

    @pytest.mark.asyncio
    async def test_fetch_workflows(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"workflows": [{"id": "w1"}]})
        assert (await client.fetch_workflows(category="c"))["workflows"][0]["id"] == "w1"
        instance.get.return_value = FakeResponse(500)
        assert (await client.fetch_workflows())["workflows"] == []

    @pytest.mark.asyncio
    async def test_get_workflow_template_uses_id(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"id": "tmpl-1"})
        result = await client.get_workflow_template("tmpl-1")
        assert result == {"id": "tmpl-1"}
        assert "/tmpl-1" in instance.get.call_args.args[0]
        instance.get.return_value = FakeResponse(500)
        assert await client.get_workflow_template("tmpl-1") is None

    @pytest.mark.asyncio
    async def test_fetch_domains(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"domains": [{"id": "d1"}]})
        assert (await client.fetch_domains())["domains"][0]["id"] == "d1"
        instance.get.return_value = FakeResponse(500)
        assert (await client.fetch_domains())["domains"] == []

    @pytest.mark.asyncio
    async def test_get_domain_template_and_install(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"id": "d1"})
        assert await client.get_domain_template("d1") == {"id": "d1"}
        instance.get.return_value = FakeResponse(500)
        assert await client.get_domain_template("d1") is None
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert (await client.install_domain("d1", "tenant-1"))["success"] is True
        instance.post.return_value = FakeResponse(500)
        assert (await client.install_domain("d1", "tenant-1"))["success"] is False

    @pytest.mark.asyncio
    async def test_search_skills(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"skills": []})
        await client.search_skills("query", {"category": "cat", "skill_type": "official"})
        params = instance.get.call_args.kwargs["params"]
        assert params["category"] == "cat"
        await client.search_skills("query", None)

    @pytest.mark.asyncio
    async def test_connect_websocket_no_websockets_package(self, client):
        with patch.dict("sys.modules", {"websockets": None}):
            with pytest.raises(RuntimeError):
                await client.connect_websocket(lambda p: None)

    @pytest.mark.asyncio
    async def test_connect_websocket_no_token(self):
        c = AtomSaaSClient(AtomSaaSConfig(ws_url="wss://x", api_url="https://x", api_token=""))
        with pytest.raises(RuntimeError):
            await c.connect_websocket(lambda p: None)

    @pytest.mark.asyncio
    async def test_connect_websocket_already_connected(self, client):
        client._connected = True
        client._ws_connection = object()
        await client.connect_websocket(lambda p: None)

    @pytest.mark.asyncio
    async def test_connect_websocket_connection_failure(self, client):
        fake_ws = MagicMock()
        fake_ws.connect = AsyncMock(side_effect=RuntimeError("refused"))
        with patch.dict("sys.modules", {"websockets": fake_ws}):
            with pytest.raises(RuntimeError):
                await client.connect_websocket(lambda p: None)
        assert client._connected is False

    @pytest.mark.asyncio
    async def test_connect_websocket_dispatch_loop(self, client):
        received = []

        class FakeConnection:
            def __init__(self, messages):
                self._messages = messages

            def __aiter__(self):
                self._iter = iter(self._messages)
                return self

            async def __anext__(self):
                try:
                    return next(self._iter)
                except StopIteration:
                    raise StopAsyncIteration

            async def close(self):
                pass

        conn = FakeConnection([b'{"type": "skill_update"}', "not-json", b'{"x": 1}'])
        fake_ws = MagicMock()
        fake_ws.connect = AsyncMock(return_value=conn)
        fake_ws.exceptions.ConnectionClosed = Exception
        with patch.dict("sys.modules", {"websockets": fake_ws}):
            async def handler(payload):
                received.append(payload)
                if isinstance(payload, dict) and payload.get("x"):
                    raise RuntimeError("handler boom")
            await client.connect_websocket(handler)
        assert len(received) == 3
        assert client._connected is False

    @pytest.mark.asyncio
    async def test_connect_websocket_connection_closed(self, client):
        class ConnectionClosed(Exception):
            pass

        class ClosedConnection:
            def __aiter__(self):
                return self

            async def __anext__(self):
                raise ConnectionClosed()

        conn = ClosedConnection()
        fake_ws = MagicMock()
        fake_ws.connect = AsyncMock(return_value=conn)
        fake_ws.exceptions = MagicMock(ConnectionClosed=ConnectionClosed)
        with patch.dict("sys.modules", {"websockets": fake_ws}):
            await client.connect_websocket(lambda p: None)
        assert client._connected is False

    @pytest.mark.asyncio
    async def test_disconnect_websocket(self, client):
        ws = MagicMock()
        ws.close = AsyncMock()
        client._ws_connection = ws
        client._connected = True
        await client.disconnect_websocket()
        ws.close.assert_awaited_once()
        assert client._connected is False
        await client.disconnect_websocket()

    @pytest.mark.asyncio
    async def test_register_instance(self, client, http_mock):
        _, instance = http_mock
        instance.post.return_value = FakeResponse(200, {"instance_id": "i1"})
        result = await client.register_instance("my-instance", "2.0", "docker")
        assert result["instance_id"] == "i1"
        payload = instance.post.call_args.kwargs["json"]
        assert payload["instance_name"] == "my-instance"
        instance.post.return_value = FakeResponse(500)
        assert (await client.register_instance())["success"] is False

    @pytest.mark.asyncio
    async def test_push_analytics_empty(self, client, http_mock):
        result = await client.push_analytics("i1", [])
        assert result == {"success": True, "count": 0}

    @pytest.mark.asyncio
    async def test_push_analytics_success_and_error(self, client, http_mock):
        _, instance = http_mock
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert (await client.push_analytics("i1", [{"m": 1}]))["success"] is True
        instance.post.return_value = FakeResponse(500)
        assert (await client.push_analytics("i1", [{"m": 1}]))["success"] is False

    @pytest.mark.asyncio
    async def test_fetch_components(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"components": [{"id": "c1"}], "total": 1})
        result = await client.fetch_components(query="q", category="cat", page=2, page_size=10)
        assert result["components"][0]["id"] == "c1"
        params = instance.get.call_args.kwargs["params"]
        assert params["limit"] == 10 and params["offset"] == 10
        instance.get.return_value = FakeResponse(500)
        assert (await client.fetch_components())["components"] == []

    @pytest.mark.asyncio
    async def test_get_component_details_and_install(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"id": "c1"})
        assert await client.get_component_details("c1") == {"id": "c1"}
        instance.get.return_value = FakeResponse(500)
        assert await client.get_component_details("c1") is None
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert (await client.install_component("c1", "canvas-1"))["success"] is True
        instance.post.return_value = FakeResponse(500)
        assert (await client.install_component("c1"))["success"] is False

    @pytest.mark.asyncio
    async def test_health_check(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200)
        assert await client.health_check() is True
        instance.get.return_value = FakeResponse(500)
        assert await client.health_check() is False

    @pytest.mark.asyncio
    async def test_close(self, client, http_mock):
        _, instance = http_mock
        client._http_client = instance
        await client.close()
        instance.aclose.assert_awaited_once()
        assert client._http_client is None

    def test_sync_wrappers(self, client, http_mock):
        _, instance = http_mock
        instance.get.return_value = FakeResponse(200, {"skills": []})
        assert client.fetch_skills_sync() == {"skills": []}
        assert client.get_skill_by_id_sync("s1") is not None
        instance.get.return_value = FakeResponse(200, [])
        assert client.get_categories_sync() == []
        instance.post.return_value = FakeResponse(200, {"success": True})
        assert client.rate_skill_sync("s1", "u1", 4)["success"] is True
        assert client.install_skill_sync("s1", "a1")["success"] is True
        instance.delete.return_value = FakeResponse(200, {"success": True})
        assert client.uninstall_skill_sync("s1", "a1")["success"] is True
        instance.get.return_value = FakeResponse(200, {"skills": []})
        assert client.search_skills_sync("q") == {"skills": []}
        instance.get.return_value = FakeResponse(200, {"agents": []})
        assert client.fetch_agents_sync() == {"agents": []}
        assert client.get_agent_template_sync("t1") is not None
        assert client.install_agent_sync("t1", "tenant-1")["success"] is True
        instance.get.return_value = FakeResponse(200, {"workflows": []})
        assert client.fetch_workflows_sync()["workflows"] == []
        assert client.get_workflow_template_sync("t1") is not None
        instance.get.return_value = FakeResponse(200, {"domains": []})
        assert client.fetch_domains_sync()["domains"] == []
        assert client.get_domain_template_sync("d1") is not None
        assert client.install_domain_sync("d1", "t")["success"] is True
        instance.get.return_value = FakeResponse(200, {"components": []})
        assert client.fetch_components_sync()["components"] == []
        assert client.get_component_details_sync("c1") is not None
        assert client.install_component_sync("c1")["success"] is True
        assert client.register_instance_sync("inst")["success"] is True
        assert client.push_analytics_sync("i1", [{"a": 1}])["success"] is True
        assert client.health_check_sync() is True

    @pytest.mark.asyncio
    async def test_http_client_headers(self, client):
        import core.atom_saas_client as mod
        with patch.object(mod.httpx, "AsyncClient") as mock_cls:
            await client._get_http_client()
            mock_cls.assert_called_once()
            kwargs = mock_cls.call_args.kwargs
            assert kwargs["headers"]["X-API-Token"] == "tok-123"
            assert kwargs["base_url"] == "https://example.com/api/v1/marketplace"


# ============================================================================
# atom_saas_websocket
# ============================================================================

class TestAtomSaaSWebSocket:
    @pytest.fixture
    def ws(self):
        return AtomSaaSWebSocketClient(api_token="tok-123", ws_url="ws://localhost:5058/ws")

    @pytest.fixture
    def db_mock(self):
        with patch("core.atom_saas_websocket.SessionLocal") as sl:
            db = MagicMock()
            sl.return_value.__enter__.return_value = db
            yield sl, db

    def test_init_and_status(self, ws):
        assert ws.api_token == "tok-123"
        assert ws.is_connected is False
        status = ws.get_status()
        assert status["connected"] is False
        assert status["rate_limit_messages_per_sec"] == 100

    def test_ws_url_from_env(self):
        with patch.dict(os.environ, {"ATOM_SAAS_WS_URL": "wss://env/ws"}, clear=True):
            c = AtomSaaSWebSocketClient(api_token="t")
        assert c.ws_url == "wss://env/ws"

    @pytest.mark.asyncio
    async def test_connect_success(self, ws, db_mock):
        fake_conn = MagicMock()
        fake_conn.close = AsyncMock()
        with patch("core.atom_saas_websocket.websockets.connect", new=AsyncMock(return_value=fake_conn)) as mock_connect:
            assert await ws.connect(lambda t, d: None) is True
        assert ws.is_connected is True
        assert mock_connect.await_args.args[0] == "ws://localhost:5058/ws?token=tok-123"
        assert ws._heartbeat_task is not None
        ws._heartbeat_task.cancel()
        ws._ws_connection = None

    @pytest.mark.asyncio
    async def test_connect_already_connected(self, ws):
        ws._connected = True
        ws._ws_connection = object()
        assert await ws.connect(None) is True

    @pytest.mark.asyncio
    async def test_connect_failure(self, ws, db_mock):
        with patch("core.atom_saas_websocket.websockets.connect", new=AsyncMock(side_effect=ConnectionError("refused"))):
            with pytest.raises(WebSocketConnectionError):
                await ws.connect(lambda t, d: None)
        assert ws._connected is False
        assert ws._consecutive_failures == 1

    @pytest.mark.asyncio
    async def test_disconnect(self, ws, db_mock):
        fake_conn = MagicMock()
        fake_conn.close = AsyncMock()
        ws._ws_connection = fake_conn
        ws._connected = True
        ws._heartbeat_task = asyncio.create_task(asyncio.sleep(100))
        await ws.disconnect()
        assert ws._connected is False
        assert ws._ws_connection is None

    @pytest.mark.asyncio
    async def test_disconnect_close_error(self, ws, db_mock):
        fake_conn = MagicMock()
        fake_conn.close = AsyncMock(side_effect=RuntimeError("close boom"))
        ws._ws_connection = fake_conn
        ws._connected = True
        await ws.disconnect()
        assert ws._connected is False

    @pytest.mark.asyncio
    async def test_send_message_not_connected(self, ws):
        assert await ws.send_message({"type": "ping"}) is False

    @pytest.mark.asyncio
    async def test_send_message_success_and_error(self, ws):
        conn = MagicMock()
        conn.send = AsyncMock()
        ws._connected = True
        ws._ws_connection = conn
        assert await ws.send_message({"type": "ping"}) is True
        conn.send = AsyncMock(side_effect=RuntimeError("send boom"))
        assert await ws.send_message({"type": "ping"}) is False

    @pytest.mark.asyncio
    async def test_message_loop_closed_ok(self, ws, db_mock):
        class ConnOK:
            def __aiter__(self):
                return self

            async def __anext__(self):
                raise ConnectionClosedOK(None, None)

        ws._ws_connection = ConnOK()
        with patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_handle:
            await ws._message_loop()
        mock_handle.assert_called_once()

    @pytest.mark.asyncio
    async def test_message_loop_closed_error(self, ws, db_mock):
        class ConnErr:
            def __aiter__(self):
                return self

            async def __anext__(self):
                raise ConnectionClosedError(None, None)

        ws._ws_connection = ConnErr()
        with patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_handle:
            await ws._message_loop()
        mock_handle.assert_called_once()

    @pytest.mark.asyncio
    async def test_message_loop_generic_error(self, ws, db_mock):
        class ConnBoom:
            def __aiter__(self):
                return self

            async def __anext__(self):
                raise RuntimeError("loop boom")

        ws._ws_connection = ConnBoom()
        with patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_handle:
            await ws._message_loop()
        mock_handle.assert_called_once()

    @pytest.mark.asyncio
    async def test_handle_message_oversize(self, ws):
        await ws._handle_message("x" * (ws.MAX_MESSAGE_SIZE + 1))

    @pytest.mark.asyncio
    async def test_handle_message_rate_limit(self, ws):
        ws._message_timestamps = [time.time()] * ws.RATE_LIMIT_MESSAGES
        await ws._handle_message('{"type": "ping"}')

    @pytest.mark.asyncio
    async def test_handle_message_bad_json(self, ws):
        await ws._handle_message("{not json")

    @pytest.mark.asyncio
    async def test_handle_message_invalid_structure(self, ws):
        await ws._handle_message('["not", "a", "dict"]')
        await ws._handle_message('{"no_type": 1}')
        await ws._handle_message('{"type": "skill_update"}')

    @pytest.mark.asyncio
    async def test_handle_message_ping_pong(self, ws):
        conn = MagicMock()
        conn.send = AsyncMock()
        ws._connected = True
        ws._ws_connection = conn
        await ws._handle_message(json.dumps({"type": "pong"}))
        await ws._handle_message(json.dumps({"type": "ping"}))
        conn.send.assert_awaited_once()
        assert "pong" in conn.send.await_args.args[0]

    @pytest.mark.asyncio
    async def test_handle_message_invalid_data(self, ws):
        await ws._handle_message(json.dumps({"type": "skill_update", "data": [1, 2]}))
        await ws._handle_message(json.dumps({"type": "skill_update", "data": {"name": "x"}}))
        await ws._handle_message(json.dumps({"type": "category_update", "data": {}}))
        await ws._handle_message(json.dumps({"type": "rating_update", "data": {"skill_id": "s1"}}))
        await ws._handle_message(json.dumps({"type": "rating_update", "data": {"skill_id": "s1", "rating": 9}}))
        await ws._handle_message(json.dumps({"type": "skill_delete", "data": {}}))

    @pytest.mark.asyncio
    async def test_handle_message_dispatch(self, ws, db_mock):
        received = []
        conn = MagicMock()
        conn.send = AsyncMock()
        ws._connected = True
        ws._ws_connection = conn

        async def handler(msg_type, data):
            received.append((msg_type, data))

        ws._message_handler = handler
        with patch.object(ws, "_update_cache", new=AsyncMock()) as mock_cache:
            await ws._handle_message(json.dumps({"type": "skill_update", "data": {"skill_id": "s1", "name": "n"}}))
        assert received[0] == ("skill_update", {"skill_id": "s1", "name": "n"})
        mock_cache.assert_awaited_once()

    def test_validate_message_helpers(self, ws):
        assert ws._validate_message({"type": "ping"}) is True
        assert ws._validate_message({"type": "skill_update", "data": {}}) is True
        assert ws._validate_message("nope") is False
        assert ws._validate_message({"data": {}}) is False
        assert ws._validate_message({"type": "skill_update"}) is False
        assert ws._validate_message_data("skill_update", {"skill_id": "s", "name": "n"}) is True
        assert ws._validate_message_data("skill_update", {"skill_id": "s"}) is False
        assert ws._validate_message_data("category_update", {"name": "c"}) is True
        assert ws._validate_message_data("category_update", {}) is False
        assert ws._validate_message_data("rating_update", {"skill_id": "s", "rating": 3}) is True
        assert ws._validate_message_data("rating_update", {"skill_id": "s", "rating": 0}) is False
        assert ws._validate_message_data("skill_delete", {"skill_id": "s"}) is True
        assert ws._validate_message_data("skill_delete", {}) is False
        assert ws._validate_message_data("mystery", {"a": 1}) is True

    @pytest.mark.asyncio
    async def test_update_cache_skill_new_and_existing(self, ws, db_mock):
        sl, db = db_mock
        db.query.return_value.filter.return_value.first.side_effect = [None, MagicMock()]
        await ws._update_cache("skill_update", {"skill_id": "s1"})
        db.add.assert_called_once()
        await ws._update_cache("skill_update", {"skill_id": "s2"})
        assert db.commit.call_count == 2

    @pytest.mark.asyncio
    async def test_update_cache_category(self, ws, db_mock):
        sl, db = db_mock
        db.query.return_value.filter.return_value.first.side_effect = [None, MagicMock()]
        await ws._update_cache("category_update", {"name": "cat1"})
        db.add.assert_called_once()
        await ws._update_cache("category_update", {"category": "cat2"})
        assert db.commit.call_count == 2

    @pytest.mark.asyncio
    async def test_update_cache_skill_delete(self, ws, db_mock):
        sl, db = db_mock
        db.query.return_value.filter.return_value.delete.return_value = 1
        await ws._update_cache("skill_delete", {"skill_id": "s1"})
        db.commit.assert_called_once()

    @pytest.mark.asyncio
    async def test_update_cache_exception(self, ws, db_mock):
        sl, db = db_mock
        db.query.side_effect = RuntimeError("db boom")
        await ws._update_cache("skill_update", {"skill_id": "s1"})

    @pytest.mark.asyncio
    async def test_heartbeat_loop_healthy(self, ws, db_mock):
        ws._connected = True
        with patch.object(ws, "_wait_for_pong", new=AsyncMock(return_value=True)), \
                patch.object(ws, "send_message", new=AsyncMock(return_value=True)), \
                patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_disc, \
                patch("asyncio.sleep", new=AsyncMock()):
            ws._connected = False
            await ws._heartbeat_loop()
        mock_disc.assert_not_called()

    @pytest.mark.asyncio
    async def test_heartbeat_loop_no_pong(self, ws, db_mock):
        ws._connected = True
        with patch.object(ws, "_wait_for_pong", new=AsyncMock(return_value=False)), \
                patch.object(ws, "send_message", new=AsyncMock(return_value=True)), \
                patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_disc, \
                patch("asyncio.sleep", new=AsyncMock()):
            ws._connected = False
            await ws._heartbeat_loop()
        mock_disc.assert_not_called()

    @pytest.mark.asyncio
    async def test_heartbeat_loop_pong_timeout(self, ws, db_mock):
        ws._connected = True
        with patch.object(ws, "_wait_for_pong", new=AsyncMock(side_effect=asyncio.TimeoutError())), \
                patch.object(ws, "send_message", new=AsyncMock(return_value=True)), \
                patch.object(ws, "_handle_disconnect", new=AsyncMock()) as mock_disc, \
                patch("asyncio.sleep", new=AsyncMock()):
            ws._connected = False
            await ws._heartbeat_loop()
        mock_disc.assert_not_called()

    @pytest.mark.asyncio
    async def test_wait_for_pong(self, ws):
        with patch("asyncio.sleep", new=AsyncMock()):
            assert await ws._wait_for_pong() is True

    @pytest.mark.asyncio
    async def test_handle_disconnect_schedules_reconnect(self, ws, db_mock):
        ws._reconnect_attempts = 0
        with patch("asyncio.create_task") as mock_create:
            await ws._handle_disconnect("test_reason")
        assert ws._last_disconnect_reason == "test_reason"
        mock_create.assert_called_once()

    @pytest.mark.asyncio
    async def test_handle_disconnect_max_attempts(self, ws, db_mock):
        ws._reconnect_attempts = ws.MAX_RECONNECT_ATTEMPTS
        with patch("asyncio.create_task") as mock_create:
            await ws._handle_disconnect("exhausted")
        mock_create.assert_not_called()

    @pytest.mark.asyncio
    async def test_reconnect_success(self, ws, db_mock):
        with patch.object(ws, "connect", new=AsyncMock(return_value=True)), \
                patch("asyncio.sleep", new=AsyncMock()):
            await ws._reconnect()
        assert ws._reconnect_attempts == 1

    @pytest.mark.asyncio
    async def test_reconnect_failure(self, ws, db_mock):
        with patch.object(ws, "connect", new=AsyncMock(side_effect=WebSocketConnectionError("nope"))), \
                patch("asyncio.sleep", new=AsyncMock()), \
                patch("asyncio.create_task") as mock_create:
            await ws._reconnect()
        assert mock_create.called

    @pytest.mark.asyncio
    async def test_reconnect_failure_maxed(self, ws, db_mock):
        ws._reconnect_attempts = ws.MAX_RECONNECT_ATTEMPTS
        with patch.object(ws, "connect", new=AsyncMock(side_effect=WebSocketConnectionError("nope"))), \
                patch("asyncio.sleep", new=AsyncMock()), \
                patch("asyncio.create_task") as mock_create:
            await ws._reconnect()
        mock_create.assert_not_called()

    @pytest.mark.asyncio
    async def test_update_db_state_create_and_update(self, ws, db_mock):
        sl, db = db_mock
        db.query.return_value.first.return_value = None
        await ws._update_db_state(connected=True)
        db.add.assert_called_once()
        db.commit.assert_called_once()
        state = MagicMock()
        db.query.return_value.first.return_value = state
        await ws._update_db_state(connected=False, disconnect_reason="x", reconnect_attempts=3)
        assert state.connected is False
        assert state.disconnect_reason == "x"
        assert state.reconnect_attempts == 3

    @pytest.mark.asyncio
    async def test_update_db_state_error(self, ws, db_mock):
        sl, db = db_mock
        db.query.side_effect = RuntimeError("boom")
        await ws._update_db_state(connected=True)

    def test_on_message_registers_handler(self, ws):
        def callback(t, d):
            pass
        ws.on_message(callback)
        assert ws._message_handler is callback

    @pytest.mark.asyncio
    async def test_handle_skill_update_and_delete(self, ws):
        with patch.object(ws, "_update_cache", new=AsyncMock()) as mock_cache:
            await ws.handle_skill_update({"skill_id": "s1"})
            await ws.handle_category_update({"name": "c1"})
            await ws.handle_skill_delete({"skill_id": "s1"})
        assert mock_cache.await_count == 3

    @pytest.mark.asyncio
    async def test_handle_rating_update(self, ws, db_mock):
        sl, db = db_mock
        await ws.handle_rating_update({"skill_id": "s1", "rating": 4, "average_rating": 4.2, "rating_count": 10})
        db.query.return_value.filter.return_value.first.return_value = None
        await ws.handle_rating_update({"skill_id": "s1", "rating": 4})
        cache = MagicMock()
        cache.skill_data = {"skill_id": "s1"}
        db.query.return_value.filter.return_value.first.return_value = cache
        await ws.handle_rating_update({"skill_id": "s1", "rating": 4, "average_rating": 4.2, "rating_count": 10})
        assert cache.skill_data["average_rating"] == 4.2
        db.query.side_effect = RuntimeError("boom")
        await ws.handle_rating_update({"skill_id": "s1", "rating": 4})

    def test_get_websocket_state(self, db_mock):
        from core.atom_saas_websocket import get_websocket_state
        sl, db = db_mock
        db.query.return_value.first.return_value = "state"
        assert get_websocket_state() == "state"
        db.query.side_effect = RuntimeError("boom")
        assert get_websocket_state() is None


# ============================================================================
# auto_document_ingestion
# ============================================================================

class TestDocumentParser:
    @pytest.fixture(autouse=True)
    def reset_docling(self):
        from core.auto_document_ingestion import DocumentParser
        old = DocumentParser._docling_processor
        DocumentParser._docling_processor = None
        yield
        DocumentParser._docling_processor = old

    @pytest.mark.asyncio
    async def test_parse_txt(self):
        from core.auto_document_ingestion import DocumentParser
        text = await DocumentParser.parse_document(b"hello world", "txt", "a.txt")
        assert text == "hello world"

    @pytest.mark.asyncio
    async def test_parse_md(self):
        from core.auto_document_ingestion import DocumentParser
        text = await DocumentParser.parse_document(b"# Title", "md", "a.md")
        assert text == "# Title"

    @pytest.mark.asyncio
    async def test_parse_json(self):
        from core.auto_document_ingestion import DocumentParser
        text = await DocumentParser.parse_document(b'{"a": 1}', "json", "a.json")
        assert '"a": 1' in text

    @pytest.mark.asyncio
    async def test_parse_json_error_returns_empty(self):
        from core.auto_document_ingestion import DocumentParser
        text = await DocumentParser.parse_document(b"{bad json", "json", "a.json")
        assert text == ""

    @pytest.mark.asyncio
    async def test_parse_csv(self):
        from core.auto_document_ingestion import DocumentParser
        text = await DocumentParser.parse_document(b"a,b\n1,2\n3,4", "csv", "a.csv")
        assert "a | b" in text
        assert "1 | 2" in text

    @pytest.mark.asyncio
    async def test_parse_csv_truncation(self):
        from core.auto_document_ingestion import DocumentParser
        content = "\n".join([f"row{i},x" for i in range(1500)]).encode()
        text = await DocumentParser.parse_document(content, "csv", "a.csv")
        assert "... (truncated)" in text

    @pytest.mark.asyncio
    async def test_parse_unsupported(self):
        from core.auto_document_ingestion import DocumentParser
        assert await DocumentParser.parse_document(b"x", "exe", "a.exe") == ""

    @pytest.mark.asyncio
    async def test_parse_docling_success(self):
        from core.auto_document_ingestion import DocumentParser
        docling = MagicMock()
        docling.process_document = AsyncMock(return_value={"success": True, "content": "parsed by docling", "total_chars": 16})
        with patch("core.docling_processor.is_docling_available", return_value=True), \
                patch("core.docling_processor.get_docling_processor", return_value=docling):
            text = await DocumentParser.parse_document(b"pdf-bytes", "pdf", "a.pdf")
        assert text == "parsed by docling"

    @pytest.mark.asyncio
    async def test_parse_docling_unavailable(self):
        from core.auto_document_ingestion import DocumentParser
        with patch("core.docling_processor.is_docling_available", return_value=False):
            text = await DocumentParser.parse_document(b"# md via fallback", "md", "a.md")
        assert text == "# md via fallback"

    @pytest.mark.asyncio
    async def test_parse_docling_import_error(self):
        from core.auto_document_ingestion import DocumentParser
        with patch.dict("sys.modules", {"core.docling_processor": None}):
            text = await DocumentParser.parse_document(b"# md", "md", "a.md")
        assert text == "# md"

    @pytest.mark.asyncio
    async def test_parse_docling_error_fallback(self):
        from core.auto_document_ingestion import DocumentParser
        docling = MagicMock()
        docling.process_document = AsyncMock(side_effect=RuntimeError("docling boom"))
        with patch("core.docling_processor.is_docling_available", return_value=True), \
                patch("core.docling_processor.get_docling_processor", return_value=docling):
            text = await DocumentParser.parse_document(b"# md", "md", "a.md")
        assert text == "# md"

    @pytest.mark.asyncio
    async def test_parse_pdf_import_error(self):
        from core.auto_document_ingestion import DocumentParser
        with patch.dict("sys.modules", {"pypdf": None}):
            text = await DocumentParser.parse_document(b"%PDF", "pdf", "a.pdf")
        assert text == "[PDF content - parser not available]"

    @pytest.mark.asyncio
    async def test_parse_pdf_error(self):
        from core.auto_document_ingestion import DocumentParser
        with patch.dict("sys.modules", {"pypdf": MagicMock()}):
            text = await DocumentParser.parse_document(b"%PDF", "pdf", "a.pdf")
        assert text == ""

    @pytest.mark.asyncio
    async def test_parse_docx_import_error(self):
        from core.auto_document_ingestion import DocumentParser
        with patch.dict("sys.modules", {"docx": None}):
            text = await DocumentParser.parse_document(b"docx", "docx", "a.docx")
        assert text == "[DOCX content - parser not available]"

    @pytest.mark.asyncio
    async def test_parse_docx_success(self):
        from core.auto_document_ingestion import DocumentParser
        fake_doc = MagicMock()
        para = MagicMock()
        para.text = "para1"
        fake_doc.paragraphs = [para]
        table = MagicMock()
        row = MagicMock()
        cell = MagicMock()
        cell.text = "c1"
        row.cells = [cell]
        table.rows = [row]
        fake_doc.tables = [table]
        with patch.dict("sys.modules", {"docx": MagicMock(Document=Mock(return_value=fake_doc))}):
            text = await DocumentParser.parse_document(b"docx", "docx", "a.docx")
        assert "para1" in text and "c1" in text

    @pytest.mark.asyncio
    async def test_parse_excel_pandas(self):
        from core.auto_document_ingestion import DocumentParser
        df = MagicMock()
        df.to_string.return_value = "row data"
        xls = MagicMock()
        xls.sheet_names = ["Sheet1", "Sheet2"]
        with patch.dict("sys.modules", {"pandas": MagicMock()}):
            import pandas
            pandas.ExcelFile = Mock(return_value=xls)
            pandas.read_excel = Mock(return_value=df)
            text = await DocumentParser.parse_document(b"xlsx", "xlsx", "a.xlsx")
        assert "Sheet: Sheet1" in text

    @pytest.mark.asyncio
    async def test_parse_excel_openpyxl_fallback(self):
        from core.auto_document_ingestion import DocumentParser
        sheet = MagicMock()
        sheet.iter_rows.return_value = [
            [None, "x"], ["y", None],
        ]
        wb = MagicMock()
        wb.sheetnames = ["S1"]
        wb.__getitem__.return_value = sheet
        with patch.dict("sys.modules", {"pandas": None, "openpyxl": MagicMock()}):
            import openpyxl
            openpyxl.load_workbook = Mock(return_value=wb)
            text = await DocumentParser.parse_document(b"xlsx", "xlsx", "a.xlsx")
        assert "=== Sheet: S1 ===" in text
        assert "y | " in text

    @pytest.mark.asyncio
    async def test_parse_excel_no_parser(self):
        from core.auto_document_ingestion import DocumentParser
        with patch.dict("sys.modules", {"pandas": None, "openpyxl": None}):
            text = await DocumentParser.parse_document(b"xlsx", "xlsx", "a.xlsx")
        assert "[Excel content - parser not available]" in text

    @pytest.mark.asyncio
    async def test_parse_csv_formula_extraction(self, tmp_path):
        from core.auto_document_ingestion import DocumentParser
        extractor = MagicMock()
        extractor.extract_from_csv.return_value = [{"formula": "=A1+B1"}]
        with patch("core.formula_extractor.get_formula_extractor", return_value=extractor):
            text = await DocumentParser.parse_document(b"a,b\n1,2", "csv", "a.csv")
        assert "a | b" in text


class TestAutoDocumentIngestionService:
    @pytest.fixture
    def service(self):
        with patch("core.lancedb_handler.get_lancedb_handler") as mock_lance, \
                patch("core.secrets_redactor.get_secrets_redactor") as mock_redactor:
            from core.auto_document_ingestion import AutoDocumentIngestionService
            memory = MagicMock()
            memory.add_document = Mock(return_value=True)
            memory.embed_text = Mock(return_value=None)
            mock_lance.return_value = memory
            redactor = MagicMock()
            redactor.redact.return_value = MagicMock(has_secrets=False)
            mock_redactor.return_value = redactor
            svc = AutoDocumentIngestionService()
            svc.memory_handler = memory
            svc.redactor = redactor
            yield svc

    def test_service_no_lancedb(self):
        with patch.dict("sys.modules", {"core.lancedb_handler": None, "core.secrets_redactor": None}):
            from core.auto_document_ingestion import AutoDocumentIngestionService
            svc = AutoDocumentIngestionService()
        assert svc.memory_handler is None
        assert svc.redactor is None

    def test_get_and_update_settings(self, service):
        settings = service.get_settings("dropbox")
        assert settings.integration_id == "dropbox"
        assert settings.enabled is False
        updated = service.update_settings(
            "dropbox", enabled=True, auto_sync_new_files=False, file_types=["pdf"],
            sync_folders=["/x"], exclude_folders=["/y"], max_file_size_mb=10, sync_frequency_minutes=30,
        )
        assert updated.enabled is True
        assert updated.file_types == ["pdf"]
        assert updated.max_file_size_mb == 10
        assert service.get_settings("dropbox") is updated

    @pytest.mark.asyncio
    async def test_process_file_bytes_no_extension(self, service):
        result = await service.process_file_bytes(b"x", "noext")
        assert result["status"] == "skipped"

    @pytest.mark.asyncio
    async def test_process_file_bytes_parse_failure(self, service):
        service.parser.parse_document = AsyncMock(side_effect=RuntimeError("boom"))
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "error"

    @pytest.mark.asyncio
    async def test_process_file_bytes_no_text(self, service):
        service.parser.parse_document = AsyncMock(return_value="")
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "skipped"

    @pytest.mark.asyncio
    async def test_process_file_bytes_short_text(self, service):
        # Threshold is blank-only now (R84f): short-but-real content ingests;
        # only empty/whitespace text is dropped.
        service.parser.parse_document = AsyncMock(return_value="ab")
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "ingested"

    @pytest.mark.asyncio
    async def test_process_file_bytes_blank_text(self, service):
        service.parser.parse_document = AsyncMock(return_value="   ")
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "skipped"
        assert result["reason"] == "no_text"

    @pytest.mark.asyncio
    async def test_process_file_bytes_redacts(self, service):
        service.parser.parse_document = AsyncMock(return_value="secret content here")
        redaction = MagicMock(has_secrets=True, redacted_text="*** content here")
        service.redactor.redact.return_value = redaction
        result = await service.process_file_bytes(b"x", "a.txt", source="onedrive", user_id="u1", workspace_id="ws1")
        assert result["status"] == "ingested"
        kwargs = service.memory_handler.add_document.call_args.kwargs
        assert kwargs["text"] == "*** content here"
        assert kwargs["metadata"]["integration_id"] == "onedrive"
        assert kwargs["user_id"] == "u1"

    @pytest.mark.asyncio
    async def test_process_file_bytes_redactor_error(self, service):
        service.parser.parse_document = AsyncMock(return_value="content here")
        service.redactor.redact.side_effect = RuntimeError("redact boom")
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "ingested"

    @pytest.mark.asyncio
    async def test_process_file_bytes_add_failed(self, service):
        service.parser.parse_document = AsyncMock(return_value="valid content")
        service.memory_handler.add_document.return_value = False
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "skipped"

    @pytest.mark.asyncio
    async def test_process_file_bytes_ingest_error(self, service):
        service.parser.parse_document = AsyncMock(return_value="valid content")
        service.memory_handler.add_document.side_effect = RuntimeError("lance boom")
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "error"

    @pytest.mark.asyncio
    async def test_process_file_bytes_no_memory_handler(self, service):
        service.parser.parse_document = AsyncMock(return_value="valid content")
        service.memory_handler = None
        result = await service.process_file_bytes(b"x", "a.txt")
        assert result["status"] == "skipped"

    @pytest.mark.asyncio
    async def test_sync_integration_disabled(self, service):
        service.update_settings("dropbox", enabled=False)
        result = await service.sync_integration("dropbox")
        assert result["skipped"] is True

    @pytest.mark.asyncio
    async def test_sync_integration_recently_synced(self, service):
        service.update_settings("dropbox", enabled=True, sync_frequency_minutes=60)
        service.settings["dropbox"].last_sync = datetime.now(timezone.utc)
        result = await service.sync_integration("dropbox")
        assert result["skipped"] is True

    @pytest.mark.asyncio
    async def test_sync_integration_skips_and_ingests(self, service):
        service.update_settings("dropbox", enabled=True, file_types=["pdf", "txt"], max_file_size_mb=1)
        files = [
            {"id": "f1", "name": "a.pdf", "size": 100, "modified_at": datetime.now(timezone.utc)},
            {"id": "f2", "name": "b.txt", "size": 100, "modified_at": None},
            {"id": "f3", "name": "c.md", "size": 100, "modified_at": None},
            {"id": "f4", "name": "big.pdf", "size": 5 * 1024 * 1024, "modified_at": None},
        ]
        service._list_files = AsyncMock(return_value=files)
        service._download_file = AsyncMock(return_value=b"content here")
        service.parser.parse_document = AsyncMock(return_value="parsed text content")
        with patch("core.doc_freshness_service.hash_text", return_value="hash-1"), \
                patch("core.doc_freshness_service.extra_columns_for_ingest", return_value={}), \
                patch("core.atom_meta_agent.handle_data_event_trigger", new=AsyncMock()):
            result = await service.sync_integration("dropbox")
        assert result["files_found"] == 4
        assert result["files_ingested"] == 2
        assert result["files_skipped"] == 2
        assert "f1" in service.ingested_docs
        assert result["success"] is True
        assert service.settings["dropbox"].last_sync is not None

    @pytest.mark.asyncio
    async def test_sync_integration_duplicate_not_modified(self, service):
        service.update_settings("dropbox", enabled=True)
        modified = datetime.now(timezone.utc)
        from core.auto_document_ingestion import IngestedDocument
        service.ingested_docs["f1"] = IngestedDocument(
            id="doc_1", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
            integration_id="dropbox", workspace_id="default", file_size_bytes=10,
            content_preview="x", ingested_at=datetime.now(timezone.utc), external_id="f1",
            external_modified_at=modified,
        )
        files = [{"id": "f1", "name": "a.pdf", "size": 10, "modified_at": modified}]
        service._list_files = AsyncMock(return_value=files)
        result = await service.sync_integration("dropbox")
        assert result["files_skipped"] == 1
        assert result["files_ingested"] == 0

    @pytest.mark.asyncio
    async def test_sync_integration_stale_marking(self, service):
        service.update_settings("dropbox", enabled=True)
        old_modified = datetime.now(timezone.utc) - timedelta(days=1)
        new_modified = datetime.now(timezone.utc)
        from core.auto_document_ingestion import IngestedDocument
        service.ingested_docs["f1"] = IngestedDocument(
            id="doc_1", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
            integration_id="dropbox", workspace_id="default", file_size_bytes=10,
            content_preview="x", ingested_at=datetime.now(timezone.utc), external_id="f1",
            external_modified_at=old_modified,
        )
        files = [{"id": "f1", "name": "a.pdf", "size": 10, "modified_at": new_modified}]
        service._list_files = AsyncMock(return_value=files)
        service._download_file = AsyncMock(return_value=b"new content here")
        service.parser.parse_document = AsyncMock(return_value="new parsed text")
        with patch.object(service, "_mark_doc_stale") as mock_stale, \
                patch("core.doc_freshness_service.hash_text", return_value="h2"), \
                patch("core.doc_freshness_service.extra_columns_for_ingest", return_value={}), \
                patch("core.atom_meta_agent.handle_data_event_trigger", new=AsyncMock()):
            result = await service.sync_integration("dropbox")
        mock_stale.assert_called_once()
        assert result["files_ingested"] == 1

    @pytest.mark.asyncio
    async def test_sync_integration_download_failure(self, service):
        service.update_settings("dropbox", enabled=True)
        files = [{"id": "f1", "name": "a.pdf", "size": 10, "modified_at": None}]
        service._list_files = AsyncMock(return_value=files)
        service._download_file = AsyncMock(return_value=None)
        result = await service.sync_integration("dropbox")
        assert result["files_ingested"] == 0

    @pytest.mark.asyncio
    async def test_sync_integration_time_limit(self, service):
        service.update_settings("dropbox", enabled=True)
        service._list_files = AsyncMock(return_value=[{"id": f"f{i}", "name": "a.pdf", "size": 1, "modified_at": None} for i in range(5)])
        with patch("core.auto_document_ingestion.datetime") as mock_dt:
            now = datetime.now(timezone.utc)
            mock_dt.now.return_value = now + timedelta(minutes=11)
            mock_dt.fromisoformat.return_value = now
            result = await service.sync_integration("dropbox")
        assert any("Time limit reached" in e for e in result["errors"])

    @pytest.mark.asyncio
    async def test_sync_integration_generic_error(self, service):
        service.update_settings("dropbox", enabled=True)
        service._list_files = AsyncMock(side_effect=RuntimeError("list boom"))
        result = await service.sync_integration("dropbox")
        assert result["success"] is False
        assert "error" in result

    @pytest.mark.asyncio
    async def test_list_files_unknown_and_error(self, service):
        assert await service._list_files("mystery", service.get_settings("mystery")) == []
        with patch.object(service, "_list_google_drive_files", side_effect=RuntimeError("boom")):
            assert await service._list_files("google_drive", service.get_settings("x")) == []

    @pytest.mark.asyncio
    async def test_download_file_paths(self, service):
        assert await service._download_file("mystery", {}) is None
        assert await service._download_file("onedrive", {}) is None
        assert await service._download_file("notion", {}) is None
        with patch.object(service, "_download_google_drive_file", side_effect=RuntimeError("boom")):
            assert await service._download_file("google_drive", {}) is None

    @pytest.mark.asyncio
    async def test_list_onedrive_notion(self, service):
        assert await service._list_onedrive_files(service.get_settings("x")) == []
        assert await service._list_notion_pages(service.get_settings("x")) == []

    @pytest.mark.asyncio
    async def test_google_drive_no_token(self, service):
        with patch.dict(os.environ, {}, clear=True):
            assert await service._list_google_drive_files(service.get_settings("x")) == []
            assert await service._download_google_drive_file({"id": "f1"}) is None

    @pytest.mark.asyncio
    async def test_google_drive_list_success_and_failure(self, service):
        with patch.dict(os.environ, {"GOOGLE_DRIVE_ACCESS_TOKEN": "tok"}):
            gds = MagicMock()
            gds.list_files = AsyncMock(return_value={"status": "success", "data": {"files": [{"id": "f1"}]}})
            with patch("integrations.google_drive_service.google_drive_service", gds):
                files = await service._list_google_drive_files(service.get_settings("x"))
            assert files[0]["id"] == "f1"
            gds.list_files = AsyncMock(return_value={"status": "error", "message": "denied"})
            with patch("integrations.google_drive_service.google_drive_service", gds):
                assert await service._list_google_drive_files(service.get_settings("x")) == []

    @pytest.mark.asyncio
    async def test_google_drive_download_export(self, service):
        def make_client():
            client = MagicMock()
            client.__aenter__ = AsyncMock(return_value=client)
            client.__aexit__ = AsyncMock(return_value=False)
            client.get = AsyncMock()
            return client

        with patch.dict(os.environ, {"DROPBOX_ACCESS_TOKEN": "", "GOOGLE_DRIVE_ACCESS_TOKEN": "tok"}):
            client = make_client()
            client.get.return_value = FakeResponse(200, {})
            with patch("httpx.AsyncClient", return_value=client):
                content = await service._download_google_drive_file(
                    {"id": "f1", "mimeType": "application/vnd.google-apps.document"}
                )
            assert "export" in client.get.await_args.args[0]
            client.get = AsyncMock(return_value=FakeResponse(200, {}))
            with patch("httpx.AsyncClient", return_value=client):
                await service._download_google_drive_file({"id": "f1", "mimeType": "application/pdf"})
            assert "alt=media" in client.get.await_args.args[0]
            client.get = AsyncMock(return_value=FakeResponse(500))
            with patch("httpx.AsyncClient", return_value=client):
                assert await service._download_google_drive_file({"id": "f1"}) is None
            assert await service._download_google_drive_file({}) is None

    @pytest.mark.asyncio
    async def test_dropbox_list_and_download(self, service):
        def make_client():
            client = MagicMock()
            client.__aenter__ = AsyncMock(return_value=client)
            client.__aexit__ = AsyncMock(return_value=False)
            client.post = AsyncMock()
            client.get = AsyncMock()
            return client

        with patch.dict(os.environ, {"DROPBOX_ACCESS_TOKEN": "tok", "GOOGLE_DRIVE_ACCESS_TOKEN": ""}):
            client0 = make_client()
            client0.post.return_value = FakeResponse(200, {"entries": []})
            with patch("httpx.AsyncClient", return_value=client0):
                assert await service._list_dropbox_files(service.get_settings("x")) == []
            client = make_client()
            client.post.return_value = FakeResponse(200, {"entries": [
                {".tag": "file", "id": "1", "name": "a.pdf", "path_lower": "/a.pdf", "size": 1},
                {".tag": "folder", "id": "2"},
            ]})
            with patch("httpx.AsyncClient", return_value=client):
                files = await service._list_dropbox_files(service.get_settings("x"))
            assert len(files) == 1
            client2 = make_client()
            client2.post.return_value = FakeResponse(200, {"link": "https://dl.example/a.pdf"})
            client2.get.return_value = FakeResponse(200, b"data")
            with patch("httpx.AsyncClient", return_value=client2):
                content = await service._download_dropbox_file({"path_lower": "/a.pdf"})
            assert content == b"data"
            client2b = make_client()
            client2b.post.return_value = FakeResponse(200, {})
            with patch("httpx.AsyncClient", return_value=client2b):
                assert await service._download_dropbox_file({"path_lower": "/a.pdf"}) is None
            client3 = make_client()
            client3.post.return_value = FakeResponse(500)
            with patch("httpx.AsyncClient", return_value=client3):
                assert await service._download_dropbox_file({"path_lower": "/a.pdf"}) is None
            with patch.dict(os.environ, {"DROPBOX_ACCESS_TOKEN": ""}, clear=True):
                assert await service._list_dropbox_files(service.get_settings("x")) == []
                assert await service._download_dropbox_file({"path_lower": "/a.pdf"}) is None

    def test_get_ingested_documents_filters(self, service):
        from core.auto_document_ingestion import IngestedDocument
        doc = IngestedDocument(
            id="d1", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
            integration_id="dropbox", workspace_id="default", file_size_bytes=10,
            content_preview="x", ingested_at=datetime.now(timezone.utc), external_id="f1",
        )
        service.ingested_docs["f1"] = doc
        assert len(service.get_ingested_documents()) == 1
        assert len(service.get_ingested_documents(integration_id="google_drive")) == 0
        assert len(service.get_ingested_documents(file_type="pdf")) == 1

    @pytest.mark.asyncio
    async def test_remove_integration_documents(self, service):
        from core.auto_document_ingestion import IngestedDocument
        for i, ext in enumerate(["f1", "f2", "f3"]):
            service.ingested_docs[ext] = IngestedDocument(
                id=f"d{i}", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
                integration_id="dropbox" if i < 2 else "google_drive", workspace_id="default",
                file_size_bytes=10, content_preview="x", ingested_at=datetime.now(timezone.utc),
                external_id=ext,
            )
        result = await service.remove_integration_documents("dropbox")
        assert result["documents_removed"] == 2
        assert len(service.ingested_docs) == 1

    def test_get_all_settings(self, service):
        service.update_settings("dropbox", enabled=True)
        all_settings = service.get_all_settings()
        assert all_settings[0]["integration_id"] == "dropbox"
        assert all_settings[0]["enabled"] is True
        assert all_settings[0]["last_sync"] is None

    def test_singleton(self):
        from core.auto_document_ingestion import get_document_ingestion_service, _doc_ingestion_service
        original = _doc_ingestion_service
        try:
            import core.auto_document_ingestion as mod
            mod._doc_ingestion_service = None
            svc = get_document_ingestion_service()
            assert get_document_ingestion_service() is svc
        finally:
            mod._doc_ingestion_service = original

    @pytest.mark.asyncio
    async def test_freshness_persist(self, service):
        from core.auto_document_ingestion import IngestedDocument
        doc = IngestedDocument(
            id="d1", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
            integration_id="dropbox", workspace_id="default", file_size_bytes=10,
            content_preview="x", ingested_at=datetime.now(timezone.utc), external_id="f1",
        )
        session = MagicMock()
        freshness_svc = MagicMock()
        session.query.return_value.filter.return_value.first.return_value = None
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc):
            service._persist_freshness_on_ingest(doc, source_url="https://x", content_hash="h", source_modified_at=None)
        session.add.assert_called_once()
        freshness_svc.mark_on_ingest.assert_called_once()
        session.close.assert_called_once()
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc):
            session.query.return_value.filter.return_value.first.return_value = MagicMock()
            service._persist_freshness_on_ingest(doc, source_url="https://x", content_hash="h", source_modified_at=None)

    def test_mark_doc_stale(self, service):
        from core.auto_document_ingestion import IngestedDocument
        doc = IngestedDocument(
            id="d1", file_name="a.pdf", file_path="/a.pdf", file_type="pdf",
            integration_id="dropbox", workspace_id="default", file_size_bytes=10,
            content_preview="x", ingested_at=datetime.now(timezone.utc), external_id="f1",
        )
        session = MagicMock()
        freshness_svc = MagicMock()
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc):
            session.query.return_value.filter.return_value.first.return_value = None
            service._mark_doc_stale(doc, reason="x")
            session.query.return_value.filter.return_value.first.return_value = MagicMock()
            service._mark_doc_stale(doc, reason="x")
        assert doc.freshness_status == "stale"
        freshness_svc.mark_stale.assert_called_once()

    def test_reevaluate_workspace(self, service):
        session = MagicMock()
        freshness_svc = MagicMock()
        freshness_svc.reevaluate_workspace.return_value = MagicMock(as_dict=Mock(return_value={"tombs": 1}))
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc):
            summary = service._reevaluate_workspace({"f1"})
        assert summary == {"tombs": 1}

    def test_maybe_supersede_older_docs(self, service):
        session = MagicMock()
        row = MagicMock()
        row.id = "old-1"
        row.content_preview = "old text"
        row.ingested_at = datetime.now(timezone.utc)
        row.external_modified_at = None
        row.source_modified_at = None
        row.freshness_status = "fresh"
        session.query.return_value.filter.return_value.filter.return_value.order_by.return_value.limit.return_value.all.return_value = [row]
        freshness_svc = MagicMock()
        freshness_svc.entity_set_for_doc.return_value = {"A"}
        freshness_svc.cascade_graph_supersession = Mock()
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc), \
                patch("core.doc_freshness_service.detect_supersession", return_value=["cand-1"]):
            service._maybe_supersede_older_docs(text="new text", new_doc_id="new-1", source_modified_at=None)
        freshness_svc.apply_supersession.assert_called_once()
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc), \
                patch("core.doc_freshness_service.detect_supersession", return_value=[]):
            service._maybe_supersede_older_docs(text="new text", new_doc_id="new-1", source_modified_at=None)
        session.query.return_value.filter.return_value.filter.return_value.order_by.return_value.limit.return_value.all.return_value = []
        with patch("core.database.SessionLocal", return_value=session), \
                patch("core.doc_freshness_service.DocFreshnessService", return_value=freshness_svc):
            service._maybe_supersede_older_docs(text="new text", new_doc_id="new-1", source_modified_at=None)
