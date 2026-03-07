#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Coverage Trend Export Utility

Purpose: Export trending data for external analysis (Excel, BI tools, custom scripts).
Supports multiple formats (CSV, JSON, Excel) with date range filtering.

Usage:
    python coverage_trend_export.py [options]

Options:
    --trending-file PATH        Path to cross_platform_trend.json (default: relative path)
    --output PATH               Path to output file (default: coverage_export.csv)
    --format FORMAT             Export format: csv|json|excel (default: csv)
    --days INT                  Number of days to export (default: 30, 0 for all history)

Example:
    python coverage_trend_export.py --format csv --days 30
    python coverage_trend_export.py --format excel --days 90 --output quarterly_report.xlsx
    python coverage_trend_export.py --format json --days 0 --output full_history.json
"""

import argparse
import csv
import json
import logging
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional

# Optional dependencies for Excel export
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)

# Default paths
TREND_FILE = Path("tests/coverage_reports/metrics/cross_platform_trend.json")


def load_trending_data(trend_file: Path) -> dict:
    """
    Load trending data from cross_platform_trend.json.

    Returns empty dict with history=[] if file doesn't exist.
    """
    try:
        with open(trend_file, 'r') as f:
            data = json.load(f)
            logger.info(f"Loaded trending data from {trend_file}")
            return data
    except FileNotFoundError:
        logger.warning(f"Trending file not found: {trend_file}")
        return {"history": [], "latest": None}
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in trending file: {e}")
        return {"history": [], "latest": None}


def filter_by_date(history: list, days: int) -> list:
    """
    Filter history entries to last N days.

    Args:
        history: List of trend entries
        days: Number of days to filter (0 for all history)

    Returns:
        Filtered list of entries
    """
    if days == 0:
        return history

    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)
    cutoff_iso = cutoff_date.isoformat().replace('+00:00', 'Z')

    filtered = [entry for entry in history if entry.get('timestamp', '') >= cutoff_iso]

    logger.info(f"Filtered {len(history)} entries to {len(filtered)} entries (last {days} days)")
    return filtered


def calculate_summary_stats(trending_data: dict) -> dict:
    """
    Calculate summary statistics for trending data.

    Returns dict with per-platform min/max/avg/current stats.
    """
    history = trending_data.get('history', [])

    if not history:
        return {}

    platforms = ['backend', 'frontend', 'mobile', 'desktop']
    stats = {}

    for platform in platforms:
        coverages = [
            entry.get('platforms', {}).get(platform, 0.0)
            for entry in history
            if platform in entry.get('platforms', {})
        ]

        if coverages:
            stats[platform] = {
                'min': round(min(coverages), 2),
                'max': round(max(coverages), 2),
                'avg': round(sum(coverages) / len(coverages), 2),
                'current': round(coverages[-1], 2) if coverages else 0.0,
                'first': round(coverages[0], 2) if coverages else 0.0,
                'last': round(coverages[-1], 2) if coverages else 0.0,
                'count': len(coverages)
            }

    # Overall stats
    stats['overall'] = {
        'entry_count': len(history),
        'date_range': {
            'start': history[0].get('timestamp') if history else None,
            'end': history[-1].get('timestamp') if history else None
        }
    }

    return stats


def export_to_csv(trending_data: dict, output_file: Path, days: int = 30) -> None:
    """
    Export trending data to CSV format.

    CSV columns:
    - timestamp, overall_coverage, backend, frontend, mobile, desktop, commit_sha, branch
    """
    history = trending_data.get('history', [])
    filtered = filter_by_date(history, days)

    if not filtered:
        logger.warning("No data to export")
        return

    try:
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)

            # Header row
            writer.writerow([
                'timestamp',
                'overall_coverage',
                'backend',
                'frontend',
                'mobile',
                'desktop',
                'commit_sha',
                'branch'
            ])

            # Data rows
            for entry in filtered:
                platforms = entry.get('platforms', {})
                writer.writerow([
                    entry.get('timestamp', ''),
                    entry.get('overall_coverage', 0.0),
                    platforms.get('backend', 0.0),
                    platforms.get('frontend', 0.0),
                    platforms.get('mobile', 0.0),
                    platforms.get('desktop', 0.0),
                    entry.get('commit_sha', ''),
                    entry.get('branch', '')
                ])

        logger.info(f"Exported {len(filtered)} rows to {output_file}")

        # Log date range
        if filtered:
            start_date = filtered[0].get('timestamp', 'Unknown')
            end_date = filtered[-1].get('timestamp', 'Unknown')
            logger.info(f"Date range: {start_date} to {end_date}")

    except IOError as e:
        logger.error(f"Failed to write CSV: {e}")
        sys.exit(1)


def export_to_json(trending_data: dict, output_file: Path, days: int = 30) -> None:
    """
    Export trending data to JSON format with metadata.

    JSON structure:
    {
        "export_time": "2026-03-07T19:30:00Z",
        "total_entries": 100,
        "filtered_entries": 30,
        "date_range": {...},
        "summary_stats": {...},
        "history": [...]
    }
    """
    history = trending_data.get('history', [])
    filtered = filter_by_date(history, days)

    if not filtered:
        logger.warning("No data to export")
        return

    # Calculate summary stats
    summary_stats = calculate_summary_stats({'history': filtered})

    # Build export structure
    export_data = {
        'export_time': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'total_entries': len(history),
        'filtered_entries': len(filtered),
        'date_range': {
            'start': filtered[0].get('timestamp') if filtered else None,
            'end': filtered[-1].get('timestamp') if filtered else None
        },
        'summary_stats': summary_stats,
        'history': filtered
    }

    try:
        with open(output_file, 'w') as f:
            json.dump(export_data, f, indent=2)

        logger.info(f"Exported {len(filtered)} entries to {output_file}")

        # Log date range
        if filtered:
            start_date = filtered[0].get('timestamp', 'Unknown')
            end_date = filtered[-1].get('timestamp', 'Unknown')
            logger.info(f"Date range: {start_date} to {end_date}")

    except IOError as e:
        logger.error(f"Failed to write JSON: {e}")
        sys.exit(1)


def export_to_excel(trending_data: dict, output_file: Path, days: int = 30) -> None:
    """
    Export trending data to Excel format with multiple sheets.

    Sheets:
    1. Summary - Overall stats, platform breakdown
    2. History - Time series data with all columns
    """
    if not EXCEL_SUPPORT:
        logger.error("Excel export requires openpyxl. Install with: pip install openpyxl")
        sys.exit(1)

    history = trending_data.get('history', [])
    filtered = filter_by_date(history, days)

    if not filtered:
        logger.warning("No data to export")
        return

    try:
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Summary sheet
        ws_summary = wb.create_sheet('Summary')

        # Header
        ws_summary['A1'] = 'Coverage Trend Summary'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:B1')

        ws_summary['A3'] = 'Export Date:'
        ws_summary['B3'] = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

        ws_summary['A5'] = 'Total Entries:'
        ws_summary['B5'] = len(history)

        ws_summary['A6'] = 'Filtered Entries:'
        ws_summary['B6'] = len(filtered)

        ws_summary['A7'] = 'Date Range:'
        ws_summary['B7'] = f"{filtered[0].get('timestamp', 'Unknown')} to {filtered[-1].get('timestamp', 'Unknown')}"

        # Platform stats
        ws_summary['A9'] = 'Platform'
        ws_summary['B9'] = 'Current'
        ws_summary['C9'] = 'Min'
        ws_summary['D9'] = 'Max'
        ws_summary['E9'] = 'Average'

        # Header formatting
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws_summary[f'{col}9']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Platform data
        platforms = ['backend', 'frontend', 'mobile', 'desktop']
        stats = calculate_summary_stats({'history': filtered})

        row = 10
        for platform in platforms:
            if platform in stats:
                ws_summary[f'A{row}'] = platform.capitalize()
                ws_summary[f'B{row}'] = stats[platform]['current']
                ws_summary[f'C{row}'] = stats[platform]['min']
                ws_summary[f'D{row}'] = stats[platform]['max']
                ws_summary[f'E{row}'] = stats[platform]['avg']
                row += 1

        # Column widths
        ws_summary.column_dimensions['A'].width = 15
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 12

        # History sheet
        ws_history = wb.create_sheet('History')

        # Header row
        headers = ['Timestamp', 'Overall', 'Backend', 'Frontend', 'Mobile', 'Desktop', 'Commit SHA', 'Branch']
        for col, header in enumerate(headers, start=1):
            cell = ws_history.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Data rows
        for row_idx, entry in enumerate(filtered, start=2):
            platforms = entry.get('platforms', {})
            ws_history.cell(row=row_idx, column=1).value = entry.get('timestamp', '')
            ws_history.cell(row=row_idx, column=2).value = entry.get('overall_coverage', 0.0)
            ws_history.cell(row=row_idx, column=3).value = platforms.get('backend', 0.0)
            ws_history.cell(row=row_idx, column=4).value = platforms.get('frontend', 0.0)
            ws_history.cell(row=row_idx, column=5).value = platforms.get('mobile', 0.0)
            ws_history.cell(row=row_idx, column=6).value = platforms.get('desktop', 0.0)
            ws_history.cell(row=row_idx, column=7).value = entry.get('commit_sha', '')
            ws_history.cell(row=row_idx, column=8).value = entry.get('branch', '')

        # Column widths
        ws_history.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_history.column_dimensions[col].width = 12
        ws_history.column_dimensions['G'].width = 20
        ws_history.column_dimensions['H'].width = 15

        # Number format for coverage columns
        for row in range(2, len(filtered) + 2):
            for col in [2, 3, 4, 5, 6]:
                ws_history.cell(row=row, column=col).number_format = '0.00'

        # Save workbook
        wb.save(output_file)

        logger.info(f"Exported {len(filtered)} entries to {output_file}")

        # Log date range
        if filtered:
            start_date = filtered[0].get('timestamp', 'Unknown')
            end_date = filtered[-1].get('timestamp', 'Unknown')
            logger.info(f"Date range: {start_date} to {end_date}")

    except Exception as e:
        logger.error(f"Failed to write Excel: {e}")
        sys.exit(1)


def main():
    """Main entry point for CLI."""
    parser = argparse.ArgumentParser(
        description='Export coverage trending data for external analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export last 30 days to CSV
  python coverage_trend_export.py --format csv --days 30

  # Export all history to JSON
  python coverage_trend_export.py --format json --days 0

  # Export last 90 days to Excel
  python coverage_trend_export.py --format excel --days 90 --output quarterly_report.xlsx
        """
    )

    parser.add_argument(
        '--trending-file',
        type=Path,
        default=TREND_FILE,
        help='Path to cross_platform_trend.json'
    )

    parser.add_argument(
        '--output',
        type=Path,
        default=Path('coverage_export.csv'),
        help='Path to output file'
    )

    parser.add_argument(
        '--format',
        choices=['csv', 'json', 'excel'],
        default='csv',
        help='Export format (default: csv)'
    )

    parser.add_argument(
        '--days',
        type=int,
        default=30,
        help='Number of days to export (default: 30, 0 for all history)'
    )

    args = parser.parse_args()

    # Load trending data
    trending_data = load_trending_data(args.trending_file)

    if not trending_data.get('history'):
        logger.error("No trending data found")
        sys.exit(1)

    # Export based on format
    if args.format == 'csv':
        export_to_csv(trending_data, args.output, args.days)
    elif args.format == 'json':
        export_to_json(trending_data, args.output, args.days)
    elif args.format == 'excel':
        export_to_excel(trending_data, args.output, args.days)

    logger.info("Export complete")


if __name__ == '__main__':
    main()
