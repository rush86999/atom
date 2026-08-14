# -*- coding: utf-8 -*-
"""Coverage wave 83a — core/circuit_breaker, core/data/dataset_manager,
core/debug_insights/performance, core/formula_extractor, core/formula_memory.

EXTENDS existing suites (test_covpush_w82c-style conventions):
  - circuit_breaker: existing tests/core/test_circuit_breaker_coverage.py hits
    97%; this file closes the last gaps (no-Redis _save_stats_to_redis, and
    the _try_reenable in-memory re-enable callback-exception path).
  - dataset_manager: existing security/data suites hit 89%; this file closes
    the format-inference branches, real-file loads (json/excel/parquet),
    DuckDB success path, pandas-DataFrame result path, import fallbacks, and
    missing-dataset head().
  - debug_insights/performance: untouched (11%); full in-memory SQLite suite
    (zero LLM spend, no network).
  - formula_extractor: existing tests/unit/test_formula_extractor.py hits 36%;
    this file adds xls/ods/openpyxl/csv bodies, import-failure fallbacks,
    implicit formulas, remaining semantic-expression branches, _store_formulas.
  - formula_memory: existing tests/core/test_formula_memory_coverage.py hits
    91%; this file adds _ensure_initialized success path, dependency-name
    fetch, get_formula/delete_formula exception paths, apply_formula generic
    exception path.
"""
from __future__ import annotations

import importlib
import importlib.util
import json
import sys
import time
import types
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from unittest.mock import Mock, call, patch

import pytest
from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker

from core.circuit_breaker import CircuitBreaker, IntegrationStats
from core.database import Base
from core.data import dataset_manager as dm_mod
from core.debug_insights.performance import PerformanceInsightGenerator
from core.formula_extractor import FormulaExtractor, get_formula_extractor
from core.formula_memory import FormulaMemoryManager
from core.models import (  # noqa: F401 (register models)
    DebugEvent,
    DebugInsight,
    DebugMetric,
    DebugInsightType,
    DebugInsightSeverity,
)


# ===========================================================================
# core/circuit_breaker.py — gap closure (before: 97%)
# ===========================================================================


class TestCircuitBreakerExtend:
    async def test_save_stats_to_redis_no_redis_returns(self):
        """_save_stats_to_redis early-returns when no Redis is configured."""
        cb = CircuitBreaker()
        await cb._save_stats_to_redis("svc", IntegrationStats())

    async def test_try_reenable_in_memory_callback_raises(self):
        """_try_reenable re-enables in-memory integration after cooldown and
        swallows a raising on_reset callback."""
        cb = CircuitBreaker()
        cb.disabled.add("svc")
        cb.disabled_until["svc"] = time.time() - 10

        def boom(name):
            raise RuntimeError("callback boom")

        cb.register_on_reset(boom)

        assert await cb._try_reenable("svc") is True
        assert "svc" not in cb.disabled
        assert "svc" not in cb.disabled_until

    async def test_try_reenable_callback_async_raises(self):
        """_try_reenable awaits async on_reset callbacks and swallows raises."""
        cb = CircuitBreaker()
        cb.disabled.add("svc")
        cb.disabled_until["svc"] = time.time() - 10

        async def boom(name):
            raise RuntimeError("async callback boom")

        cb.register_on_reset(boom)

        assert await cb._try_reenable("svc") is True

    async def test_try_reenable_not_disabled_in_memory(self):
        """_try_reenable returns True when integration was never disabled."""
        cb = CircuitBreaker()
        assert await cb._try_reenable("fresh") is True

    async def test_try_reenable_still_in_cooldown_in_memory(self):
        """_try_reenable returns False while in-memory cooldown active."""
        cb = CircuitBreaker()
        cb.disabled.add("svc")
        cb.disabled_until["svc"] = time.time() + 1000
        assert await cb._try_reenable("svc") is False
        assert "svc" in cb.disabled


# ===========================================================================
# core/data/dataset_manager.py — gap closure (before: 89%)
# ===========================================================================


class _FakeDuckDB:
    """Stand-in for duckdb returning a real pandas DataFrame from .df()."""

    def __init__(self):
        self.calls = []

    def sql(self, query):
        self.calls.append(query)
        import pandas as pd

        return type("_Result", (), {"df": lambda self_: pd.DataFrame({"x": [1, 2, 3]})})()


class TestDatasetManagerExtend:
    def test_validate_source_rejects_url(self):
        with pytest.raises(ValueError, match="URL sources are not allowed"):
            dm_mod._validate_dataset_source("https://evil.example.com/data.csv")

    def test_validate_sql_empty_returns_none(self):
        assert dm_mod._validate_dataset_sql("") is None
        assert dm_mod._validate_dataset_sql(None) is None

    def test_load_runtime_error_without_pandas(self, monkeypatch):
        monkeypatch.setattr(dm_mod, "_HAS_PANDAS", False)
        with pytest.raises(RuntimeError, match="pandas is required"):
            dm_mod.DatasetManager().load('[{"a": 1}]', name="x")

    def test_load_json_file_inferred(self, tmp_path):
        p = tmp_path / "data.json"
        p.write_text('[{"a": 1}, {"a": 2}]')
        handle = dm_mod.DatasetManager().load(str(p), name="j")
        assert handle.row_count == 2
        assert handle.columns == ["a"]
        assert handle.backend == "pandas"

    def test_load_excel_file_inferred(self, tmp_path):
        import openpyxl

        p = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append([1, 2])
        ws.append([3, 4])
        wb.save(str(p))
        wb.close()

        handle = dm_mod.DatasetManager().load(str(p), name="x")
        assert handle.row_count == 2
        assert handle.columns == ["a", "b"]

    def test_load_parquet_file_inferred(self, tmp_path):
        import pandas as pd

        p = tmp_path / "data.parquet"
        pd.DataFrame({"a": [1, 2]}).to_parquet(str(p))

        handle = dm_mod.DatasetManager().load(str(p), name="pq")
        assert handle.row_count == 2
        assert handle.backend == "pandas"

    def test_load_unsupported_format_raises(self, tmp_path):
        with pytest.raises(ValueError, match="Unsupported format"):
            dm_mod.DatasetManager().load(str(tmp_path / "x.txt"), name="x", format="weird")

    def test_get_handle_returns_cached_handle(self):
        mgr = dm_mod.DatasetManager()
        mgr.load('[{"a": 1}]', name="t", session_id="s1")
        handle = mgr.get_handle("t", "s1")
        assert handle is not None
        assert handle.name == "t"
        assert handle.session_id == "s1"
        assert mgr.get_handle("nope", "s1") is None

    def test_query_duckdb_success_path(self, monkeypatch):
        mgr = dm_mod.DatasetManager()
        mgr.load('[{"a": 1}, {"a": 2}]', name="t", session_id="s")
        fake = _FakeDuckDB()
        monkeypatch.setattr(dm_mod, "_HAS_DUCKDB", True)
        monkeypatch.setattr(dm_mod, "_duckdb", fake, raising=False)

        result = mgr.query("t", "select * from t", session_id="s")
        assert result["success"] is True
        assert result["data"] == [{"x": 1}, {"x": 2}, {"x": 3}]
        assert result["columns"] == ["x"]
        assert result["row_count"] == 3
        assert fake.calls == ["select * from t"]

    def test_query_pandas_dataframe_result(self):
        mgr = dm_mod.DatasetManager()
        mgr.load('[{"a": 1}, {"a": 2}, {"a": 3}]', name="t", session_id="s")
        result = mgr.query("t", 'df[df["a"] > 1]', session_id="s")
        assert result["success"] is True
        assert result["data"] == [{"a": 2}, {"a": 3}]
        assert result["columns"] == ["a"]

    def test_head_missing_dataset(self):
        mgr = dm_mod.DatasetManager()
        result = mgr.head("missing")
        assert result["success"] is False
        assert "not loaded" in result["error"]

    def test_dollar_quote_sql_with_restricted_name_allowed(self):
        """B17: dollar-quoted strings containing a restricted fn name must not
        be blocked (regression guard for the $tag$...$tag$ strip)."""
        assert dm_mod._validate_dataset_sql(
            "select $$read_csv$$ as x"
        ) is None
        assert dm_mod._validate_dataset_sql(
            "select $tag$read_csv$tag$ as x"
        ) is None
        assert dm_mod._validate_dataset_sql(
            "select 'read_csv' as x"
        ) is None


class TestDatasetManagerImportFallbacks:
    def test_duckdb_import_error_fallback(self):
        with patch.dict(sys.modules, {"duckdb": None}):
            importlib.reload(dm_mod)
            assert dm_mod._HAS_DUCKDB is False
        importlib.reload(dm_mod)
        assert dm_mod._HAS_DUCKDB == (importlib.util.find_spec("duckdb") is not None)

    def test_duckdb_import_success_sets_flag(self):
        fake_duckdb = types.ModuleType("duckdb")
        with patch.dict(sys.modules, {"duckdb": fake_duckdb}):
            importlib.reload(dm_mod)
            assert dm_mod._HAS_DUCKDB is True
        importlib.reload(dm_mod)
        assert dm_mod._HAS_DUCKDB == (importlib.util.find_spec("duckdb") is not None)

    def test_pandas_import_error_fallback(self):
        with patch.dict(sys.modules, {"pandas": None}):
            importlib.reload(dm_mod)
            assert dm_mod._HAS_PANDAS is False
        importlib.reload(dm_mod)
        assert dm_mod._HAS_PANDAS is True


# ===========================================================================
# core/debug_insights/performance.py (before: 11%)
# ===========================================================================


@pytest.fixture()
def db():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()
    engine.dispose()


def _event(db, eid, *, component_type="agent", component_id="agent-1",
           correlation_id="corr-1", data=None, message="msg", ts=None):
    event = DebugEvent(
        id=eid,
        event_type="log",
        component_type=component_type,
        component_id=component_id,
        correlation_id=correlation_id,
        data=data,
        message=message,
        timestamp=ts if ts is not None else datetime.now(timezone.utc),
    )
    db.add(event)
    db.commit()
    return event


def _metric(db, mid, *, metric_name, component_type="agent",
            component_id="agent-1", value=0.0, ts=None):
    metric = DebugMetric(
        id=mid,
        metric_name=metric_name,
        component_type=component_type,
        component_id=component_id,
        value=value,
        timestamp=ts if ts is not None else datetime.now(timezone.utc),
    )
    db.add(metric)
    db.commit()
    return metric


class _BadSession:
    def query(self, model):
        raise RuntimeError("db down")


class TestAnalyzeComponentLatency:
    async def test_insufficient_data_returns_none(self, db):
        for i in range(9):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.analyze_component_latency("agent", "agent-1") is None

    async def test_events_without_duration_returns_none(self, db):
        for i in range(12):
            _event(db, f"e{i}", data={"note": "no duration"})
        gen = PerformanceInsightGenerator(db)
        assert await gen.analyze_component_latency("agent", "agent-1") is None

    async def test_high_p95_returns_warning(self, db):
        for i in range(9):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        _event(db, "e9", data={"duration_ms": 6000.0})
        gen = PerformanceInsightGenerator(db)
        insight = await gen.analyze_component_latency("agent", "agent-1")
        assert insight is not None
        assert insight.severity == DebugInsightSeverity.WARNING.value
        assert insight.insight_type == DebugInsightType.PERFORMANCE.value
        assert "High latency" in insight.title
        assert insight.evidence["p95_ms"] == 6000
        assert insight.confidence_score == 0.90
        assert insight.affected_components == [{"type": "agent", "id": "agent-1"}]
        assert len(insight.suggestions) == 4

    async def test_acceptable_latency_returns_info(self, db):
        for i in range(10):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        gen = PerformanceInsightGenerator(db)
        insight = await gen.analyze_component_latency("agent", "agent-1")
        assert insight is not None
        assert insight.severity == DebugInsightSeverity.INFO.value
        assert "acceptable" in insight.title
        assert insight.evidence["p50_ms"] == 100
        assert insight.confidence_score == 0.95

    async def test_db_exception_returns_none(self):
        gen = PerformanceInsightGenerator(_BadSession())
        assert await gen.analyze_component_latency("agent", "agent-1") is None


class TestIdentifyBottlenecks:
    async def test_no_events_returns_none(self, db):
        gen = PerformanceInsightGenerator(db)
        assert await gen.identify_bottlenecks("corr-missing") is None

    async def test_events_without_duration_returns_none(self, db):
        _event(db, "e1", data={"note": "x"})
        _event(db, "e2", data=None)
        gen = PerformanceInsightGenerator(db)
        assert await gen.identify_bottlenecks("corr-1") is None

    async def test_dominant_step_returns_warning(self, db):
        _event(db, "e1", data={"duration_ms": 10.0}, message="step1")
        _event(db, "e2", data={"duration_ms": 80.0}, message="step2")
        _event(db, "e3", data={"duration_ms": 10.0}, message="step3")
        gen = PerformanceInsightGenerator(db)
        insight = await gen.identify_bottlenecks("corr-1")
        assert insight is not None
        assert insight.title == "Performance bottleneck identified"
        assert insight.evidence["slowest_percentage"] == 80.0
        assert insight.evidence["total_duration_ms"] == 100.0
        assert insight.evidence["all_steps"][0]["component"] == "agent/agent-1"
        assert insight.confidence_score == 0.88

    async def test_no_dominant_step_returns_none(self, db):
        _event(db, "e1", data={"duration_ms": 40.0})
        _event(db, "e2", data={"duration_ms": 30.0})
        _event(db, "e3", data={"duration_ms": 30.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.identify_bottlenecks("corr-1") is None

    async def test_zero_total_duration_returns_none(self, db):
        _event(db, "e1", data={"duration_ms": 0.0})
        _event(db, "e2", data={"duration_ms": 0.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.identify_bottlenecks("corr-1") is None

    async def test_db_exception_returns_none(self):
        gen = PerformanceInsightGenerator(_BadSession())
        assert await gen.identify_bottlenecks("corr-1") is None


class TestTrackResourceUtilization:
    async def test_no_metrics_returns_empty(self, db):
        gen = PerformanceInsightGenerator(db)
        assert await gen.track_resource_utilization() == []

    async def test_high_cpu_and_memory_warnings(self, db):
        _metric(db, "m1", metric_name="cpu_usage", value=90.0)
        _metric(db, "m2", metric_name="cpu_usage", value=95.0)
        _metric(db, "m3", metric_name="memory_usage", value=85.0)
        gen = PerformanceInsightGenerator(db)
        insights = await gen.track_resource_utilization()
        assert len(insights) == 2
        titles = [i.title for i in insights]
        assert any("High CPU" in t for t in titles)
        assert any("High memory" in t for t in titles)
        cpu = next(i for i in insights if "CPU" in i.title)
        assert cpu.evidence["avg_cpu_percent"] == 92.5
        assert cpu.evidence["max_cpu_percent"] == 95.0
        mem = next(i for i in insights if "memory" in i.title)
        assert mem.evidence["avg_memory_percent"] == 85.0
        assert mem.confidence_score == 0.92

    async def test_below_threshold_returns_empty(self, db):
        _metric(db, "m1", metric_name="cpu_usage", value=50.0)
        _metric(db, "m2", metric_name="memory_usage", value=40.0)
        gen = PerformanceInsightGenerator(db)
        assert await gen.track_resource_utilization() == []

    async def test_db_exception_returns_empty(self):
        gen = PerformanceInsightGenerator(_BadSession())
        assert await gen.track_resource_utilization() == []


class TestDetectPerformanceDegradation:
    async def test_insufficient_events_returns_none(self, db):
        for i in range(12):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.detect_performance_degradation("agent", "agent-1") is None

    async def test_degradation_detected_returns_warning(self, db):
        for i in range(10):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        for i in range(10, 20):
            _event(db, f"e{i}", data={"duration_ms": 200.0})
        gen = PerformanceInsightGenerator(db)
        insight = await gen.detect_performance_degradation("agent", "agent-1")
        assert insight is not None
        assert "degradation" in insight.title
        assert insight.evidence["first_half_avg_ms"] == 100.0
        assert insight.evidence["second_half_avg_ms"] == 200.0
        assert insight.evidence["degradation_percent"] == pytest.approx(100.0)
        assert insight.confidence_score == 0.85
        assert len(insight.suggestions) == 5

    async def test_no_degradation_returns_none(self, db):
        for i in range(20):
            _event(db, f"e{i}", data={"duration_ms": 100.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.detect_performance_degradation("agent", "agent-1") is None

    async def test_first_half_missing_durations_returns_none(self, db):
        for i in range(10):
            _event(db, f"e{i}", data={"note": "no duration"})
        for i in range(10, 20):
            _event(db, f"e{i}", data={"duration_ms": 200.0})
        gen = PerformanceInsightGenerator(db)
        assert await gen.detect_performance_degradation("agent", "agent-1") is None

    async def test_db_exception_returns_none(self):
        gen = PerformanceInsightGenerator(_BadSession())
        assert await gen.detect_performance_degradation("agent", "agent-1") is None


class TestAnalyzeThroughput:
    async def test_no_events_returns_none(self, db):
        gen = PerformanceInsightGenerator(db)
        assert await gen.analyze_throughput("agent") is None

    async def test_low_throughput_returns_info(self, db):
        for i in range(5):
            _event(db, f"e{i}")
        gen = PerformanceInsightGenerator(db)
        insight = await gen.analyze_throughput("agent")
        assert insight is not None
        assert "Low throughput" in insight.title
        assert insight.severity == DebugInsightSeverity.INFO.value
        assert insight.evidence["avg_throughput_per_min"] == 5
        assert insight.confidence_score == 0.80

    async def test_high_throughput_returns_info(self, db):
        for i in range(1100):
            _event(db, f"e{i}")
        gen = PerformanceInsightGenerator(db)
        insight = await gen.analyze_throughput("agent")
        assert insight is not None
        assert "High throughput" in insight.title
        assert insight.evidence["avg_throughput_per_min"] == 1100
        assert insight.evidence["min_throughput_per_min"] == 1100
        assert insight.evidence["max_throughput_per_min"] == 1100

    async def test_mid_throughput_returns_none(self, db):
        for i in range(50):
            _event(db, f"e{i}")
        gen = PerformanceInsightGenerator(db)
        assert await gen.analyze_throughput("agent") is None

    async def test_db_exception_returns_none(self):
        gen = PerformanceInsightGenerator(_BadSession())
        assert await gen.analyze_throughput("agent") is None


class TestParseTimeRange:
    def test_last_1h(self):
        gen = PerformanceInsightGenerator(_BadSession())
        result = gen._parse_time_range("last_1h")
        assert abs((datetime.now(timezone.utc) - result).total_seconds() - 3600) < 5

    def test_last_24h(self):
        gen = PerformanceInsightGenerator(_BadSession())
        result = gen._parse_time_range("last_24h")
        assert abs((datetime.now(timezone.utc) - result).total_seconds() - 86400) < 5

    def test_last_7d(self):
        gen = PerformanceInsightGenerator(_BadSession())
        result = gen._parse_time_range("last_7d")
        assert abs((datetime.now(timezone.utc) - result).total_seconds() - 604800) < 5

    def test_unknown_defaults_to_1h(self):
        gen = PerformanceInsightGenerator(_BadSession())
        result = gen._parse_time_range("last_30m")
        assert abs((datetime.now(timezone.utc) - result).total_seconds() - 3600) < 5


# ===========================================================================
# core/formula_extractor.py — gap closure (before: 36%)
# ===========================================================================


class _FakeCell:
    def __init__(self, value, column, coordinate):
        self.value = value
        self.column = column
        self.coordinate = coordinate


class _FakeSheet:
    def __init__(self, header_cells, rows):
        self._header = header_cells
        self._rows = rows

    def iter_rows(self):
        yield from self._rows

    def __getitem__(self, idx):
        return self._header


class _FakeXlsCell:
    def __init__(self, ctype, value):
        self.ctype = ctype
        self.value = value


class _FakeXlsSheet:
    def __init__(self, name, rows):
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell(self, row, col):
        if row < self.nrows and col < len(self._rows[row]):
            return self._rows[row][col]
        return _FakeXlsCell(0, "")


class _FakeXlsWorkbook:
    def __init__(self, sheets):
        self._sheets = sheets

    def sheets(self):
        return self._sheets


class _FakeOdfEl:
    def __init__(self, attrs=None, children=None):
        self._attrs = attrs or {}
        self._children = children or []

    def getAttribute(self, name):
        return self._attrs.get(name)

    def getElementsByType(self, t):
        return [c for c in self._children if isinstance(c, t)]


def _build_fake_odf_modules():
    """Build fake odf / odf.table / odf.text / odf.opendocument modules."""
    odf_root = types.ModuleType("odf")
    odf_table = types.ModuleType("odf.table")
    odf_text = types.ModuleType("odf.text")
    odf_opendoc = types.ModuleType("odf.opendocument")

    class Table(_FakeOdfEl):
        pass

    class TableRow(_FakeOdfEl):
        pass

    class TableCell(_FakeOdfEl):
        pass

    class P(_FakeOdfEl):
        def __init__(self, text):
            super().__init__()
            self._text = text

        def __str__(self):
            return self._text

    odf_table.Table = Table
    odf_table.TableRow = TableRow
    odf_table.TableCell = TableCell
    odf_text.P = P

    doc = types.SimpleNamespace(spreadsheet=_FakeOdfEl())
    odf_opendoc.load = Mock(return_value=doc)

    return {
        "odf": odf_root,
        "odf.table": odf_table,
        "odf.text": odf_text,
        "odf.opendocument": odf_opendoc,
    }


def _make_odf_sheet(odf_table, odf_text, name, header, formula_cells):
    """header: list of str; formula_cells: list of col formula strings."""
    header_row = odf_table.TableRow(children=[
        odf_table.TableCell(children=[odf_text.P(h)]) for h in header
    ])
    data_row = odf_table.TableRow(children=[
        odf_table.TableCell(attrs={"formula": f}) for f in formula_cells
    ])
    return odf_table.Table(attrs={"name": name}, children=[header_row, data_row])


class TestGetFormulaManagerAndFactory:
    def test_get_formula_manager_lazy_loads_once(self):
        ex = FormulaExtractor("ws-lazy")
        fake = Mock()
        with patch("core.formula_memory.get_formula_manager", return_value=fake) as getter:
            assert ex._get_formula_manager() is fake
            assert ex._get_formula_manager() is fake
        assert getter.call_count == 1
        assert getter.call_args == call("ws-lazy")

    def test_get_formula_extractor_factory(self):
        ex = get_formula_extractor("ws-factory")
        assert isinstance(ex, FormulaExtractor)
        assert ex.workspace_id == "ws-factory"


class TestExtractFromFileRouting:
    def test_xlsx_routes_to_extract_from_excel(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_excel", lambda *a, **k: ["xlsx-formula"])
        assert ex.extract_from_file("book.xlsx") == ["xlsx-formula"]

    def test_csv_routes_to_extract_from_csv(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_csv", lambda *a, **k: ["csv-formula"])
        assert ex.extract_from_file("book.csv") == ["csv-formula"]

    def test_xls_routes_to_extract_from_xls(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_xls", lambda *a, **k: ["xls-formula"])
        assert ex.extract_from_file("book.xls") == ["xls-formula"]

    def test_ods_routes_to_extract_from_ods(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_ods", lambda *a, **k: ["ods-formula"])
        assert ex.extract_from_file("book.ods") == ["ods-formula"]

    def test_numbers_routes_to_extract_from_excel(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_excel", lambda *a, **k: ["numbers-formula"])
        assert ex.extract_from_file("book.numbers") == ["numbers-formula"]

    def test_unsupported_extension_returns_empty(self):
        ex = FormulaExtractor()
        assert ex.extract_from_file("book.txt") == []


class TestExtractFromExcel:
    def test_real_workbook_multiple_sheets(self, tmp_path, monkeypatch):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Revenue", "Cost"])
        ws1.append(["=SUM(A2:B2)", "=AVERAGE(A2:B2)"])
        ws2 = wb.create_sheet("Ops")
        ws2.append(["A", "B"])
        ws2.append(["=A2+B2", 5])
        p = tmp_path / "book.xlsx"
        wb.save(str(p))
        wb.close()

        ex = FormulaExtractor("ws")
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        formulas = ex.extract_from_excel(str(p), user_id="u1", auto_store=True)
        assert len(formulas) == 3
        assert {f["source_sheet"] for f in formulas} == {"Sales", "Ops"}
        by_cell = {(f["source_sheet"], f["source_cell"]): f for f in formulas}
        assert by_cell[("Sales", "A2")]["expression"] == "sum(Revenue, Cost)"
        assert by_cell[("Sales", "A2")]["name"] == "Revenue"
        assert by_cell[("Sales", "B2")]["expression"] == "average(Revenue, Cost)"
        assert by_cell[("Sales", "B2")]["name"] == "Cost"
        assert by_cell[("Ops", "A2")]["expression"] == "A + B"
        assert by_cell[("Ops", "A2")]["name"] == "A"

    def test_auto_store_disabled_skips_store(self, tmp_path, monkeypatch):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append(["=A2*2"])
        p = tmp_path / "nostore.xlsx"
        wb.save(str(p))
        wb.close()

        ex = FormulaExtractor()
        store = Mock()
        monkeypatch.setattr(ex, "_store_formulas", store)
        formulas = ex.extract_from_excel(str(p), auto_store=False)
        assert len(formulas) == 1
        store.assert_not_called()

    def test_openpyxl_missing_returns_empty(self):
        with patch.dict(sys.modules, {"openpyxl": None}):
            ex = FormulaExtractor()
            assert ex.extract_from_excel("book.xlsx") == []

    def test_load_failure_returns_empty(self, monkeypatch):
        import openpyxl

        monkeypatch.setattr(openpyxl, "load_workbook", Mock(side_effect=RuntimeError("corrupt")))
        ex = FormulaExtractor()
        assert ex.extract_from_excel("book.xlsx") == []


class TestExtractFromXls:
    def _ex(self, monkeypatch=None, store_patch=True):
        ex = FormulaExtractor()
        if store_patch:
            monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        return ex

    def test_real_xlrd_flow(self, monkeypatch):
        header = [_FakeXlsCell(1, "Revenue"), _FakeXlsCell(1, "Cost")]
        row1 = [_FakeXlsCell(1, "=SUM(A2:B2)"), _FakeXlsCell(1, "text")]
        sheet = _FakeXlsSheet("Sheet1", [header, row1])
        wb = _FakeXlsWorkbook([sheet])
        xlrd_mod = types.ModuleType("xlrd")
        xlrd_mod.XL_CELL_TEXT = 1
        xlrd_mod.open_workbook = Mock(return_value=wb)

        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, {"xlrd": xlrd_mod}):
            formulas = ex.extract_from_xls("book.xls", auto_store=True)
        assert len(formulas) == 1
        f = formulas[0]
        assert f["expression"] == "sum(Revenue, Cost)"
        assert f["source_sheet"] == "Sheet1"
        assert f["source_cell"] == "R2C1"
        assert f["original_formula"] == "=SUM(A2:B2)"
    def test_non_formula_text_ignored(self, monkeypatch):
        header = [_FakeXlsCell(1, "A")]
        row1 = [_FakeXlsCell(1, "plain text")]
        sheet = _FakeXlsSheet("S", [header, row1])
        xlrd_mod = types.ModuleType("xlrd")
        xlrd_mod.XL_CELL_TEXT = 1
        xlrd_mod.open_workbook = Mock(return_value=_FakeXlsWorkbook([sheet]))

        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, {"xlrd": xlrd_mod}):
            formulas = ex.extract_from_xls("book.xls")
        assert formulas == []

    def test_empty_sheet_no_headers(self, monkeypatch):
        sheet = _FakeXlsSheet("Empty", [])
        xlrd_mod = types.ModuleType("xlrd")
        xlrd_mod.XL_CELL_TEXT = 1
        xlrd_mod.open_workbook = Mock(return_value=_FakeXlsWorkbook([sheet]))

        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, {"xlrd": xlrd_mod}):
            assert ex.extract_from_xls("book.xls") == []

    def test_xlrd_missing_falls_back_to_excel(self, monkeypatch):
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "extract_from_excel", lambda *a, **k: ["fallback"])
        with patch.dict(sys.modules, {"xlrd": None}):
            assert ex.extract_from_xls("book.xls") == ["fallback"]

    def test_open_workbook_failure_returns_empty(self, monkeypatch):
        xlrd_mod = types.ModuleType("xlrd")
        xlrd_mod.XL_CELL_TEXT = 1
        xlrd_mod.open_workbook = Mock(side_effect=RuntimeError("bad xls"))
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, {"xlrd": xlrd_mod}):
            assert ex.extract_from_xls("book.xls") == []

    def test_get_xls_headers(self):
        ex = FormulaExtractor()
        sheet = _FakeXlsSheet("S", [
            [_FakeXlsCell(1, " Revenue "), _FakeXlsCell(0, ""), _FakeXlsCell(1, "Cost")],
        ])
        assert ex._get_xls_headers(sheet) == {1: "Revenue", 3: "Cost"}

    def test_parse_xls_formula(self):
        ex = FormulaExtractor()
        result = ex._parse_xls_formula(
            formula_str="=SUM(A2:B2)",
            row=1,
            col=0,
            headers={1: "Sales", 2: "Revenue"},
            sheet_name="Sheet1",
        )
        assert result is not None
        assert result["expression"] == "sum(Sales, Revenue)"
        assert result["name"] == "Sales"
        assert result["domain"] == "finance"
        assert result["source_cell"] == "R2C1"
        assert {p["name"] for p in result["parameters"]} == {"Sales", "Revenue"}
        assert all(p["type"] == "number" for p in result["parameters"])


class TestExtractFromCsv:
    def test_real_csv_with_formula(self, tmp_path, monkeypatch):
        p = tmp_path / "data.csv"
        p.write_text(
            "Revenue,Cost\n=SUM(A2:B2),100\nnot-a-formula,200\n",
            encoding="utf-8-sig",
        )
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        formulas = ex.extract_from_csv(str(p), auto_store=True)
        assert len(formulas) == 1
        f = formulas[0]
        assert f["expression"] == "sum(Revenue, Cost)"
        assert f["source_sheet"] == "CSV"
        assert f["source_cell"] == "R2C1"
        assert f["name"] == "Revenue"

    def test_implicit_formula_detected(self, tmp_path, monkeypatch):
        p = tmp_path / "implicit.csv"
        p.write_text(
            "Item,Price,Quantity,Total\n"
            "A,10,2,20\n"
            "B,20,3,60\n"
            "C,30,4,120\n",
            encoding="utf-8-sig",
        )
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        formulas = ex.extract_from_csv(str(p))
        assert len(formulas) == 1
        f = formulas[0]
        assert f["original_formula"] == "(implicit)"
        assert f["expression"] == "Price * Quantity"
        assert f["source_cell"] == "implicit"

    def test_read_failure_returns_empty(self, tmp_path):
        ex = FormulaExtractor()
        assert ex.extract_from_csv(str(tmp_path / "missing.csv")) == []

    def test_empty_file_returns_empty(self, tmp_path):
        p = tmp_path / "empty.csv"
        p.write_text("", encoding="utf-8-sig")
        ex = FormulaExtractor()
        assert ex.extract_from_csv(str(p)) == []

    def test_parse_csv_formula(self):
        ex = FormulaExtractor()
        result = ex._parse_csv_formula(
            formula_str="=SUM(B2:C2)",
            row=2,
            col=0,
            headers={2: "Sales", 3: "Revenue"},
        )
        assert result["expression"] == "sum(Sales, Revenue)"
        assert result["name"] == "Column_1"
        assert result["source_sheet"] == "CSV"
        assert result["source_cell"] == "R2C1"

    def test_detect_implicit_formulas_too_few_rows(self):
        ex = FormulaExtractor()
        assert ex._detect_implicit_formulas([["A", "B"]], {1: "A"}) == []

    def test_detect_implicit_formulas_no_headers(self):
        ex = FormulaExtractor()
        rows = [["Price", "Qty", "Total"], ["1", "2", "2"], ["3", "4", "12"]]
        assert ex._detect_implicit_formulas(rows, {}) == []

    def test_detect_implicit_formulas_non_numeric_column(self):
        ex = FormulaExtractor()
        rows = [
            ["Name", "Price", "Qty", "Total"],
            ["A", "10", "2", "20"],
            ["B", "20", "3", "60"],
            ["C", "30", "4", "120"],
        ]
        headers = {1: "Name", 2: "Price", 3: "Qty", 4: "Total"}
        implicit = ex._detect_implicit_formulas(rows, headers)
        assert len(implicit) == 1
        assert implicit[0]["expression"] == "Price * Qty"
        assert implicit[0]["use_case"] == "Calculate Total from Price and Qty"

    def test_detect_implicit_formulas_no_calculated_column(self):
        ex = FormulaExtractor()
        rows = [
            ["Price", "Qty"],
            ["10", "2"],
            ["20", "3"],
            ["30", "4"],
        ]
        headers = {1: "Price", 2: "Qty"}
        assert ex._detect_implicit_formulas(rows, headers) == []


class TestExtractFromOds:
    def test_real_odf_flow(self, monkeypatch):
        odf_mods = _build_fake_odf_modules()
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, odf_mods):
            sheet = _make_odf_sheet(
                odf_mods["odf.table"], odf_mods["odf.text"],
                "Sheet1", ["Revenue", "Cost"],
                ["of:SUM(A1:B1)", "=AVERAGE(A1:B1)"],
            )
            odf_mods["odf.opendocument"].load.return_value.spreadsheet = _FakeOdfEl(
                children=[sheet]
            )
            formulas = ex.extract_from_ods("book.ods", auto_store=True)
        assert len(formulas) == 2
        by_cell = {(f["source_sheet"], f["source_cell"]): f for f in formulas}
        sum_f = by_cell[("Sheet1", "R2C1")]
        assert sum_f["expression"] == "sum(Revenue, Cost)"
        assert sum_f["original_formula"] == "=SUM(A1:B1)"
        avg_f = by_cell[("Sheet1", "R2C2")]
        assert avg_f["expression"] == "average(Revenue, Cost)"

    def test_odf_sheet_without_rows(self, monkeypatch):
        odf_mods = _build_fake_odf_modules()
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, odf_mods):
            empty_sheet = odf_mods["odf.table"].Table(attrs={"name": "Empty"})
            odf_mods["odf.opendocument"].load.return_value.spreadsheet = _FakeOdfEl(
                children=[empty_sheet]
            )
            assert ex.extract_from_ods("book.ods") == []

    def test_odf_missing_returns_empty(self):
        with patch.dict(sys.modules, {"odf": None}):
            ex = FormulaExtractor()
            assert ex.extract_from_ods("book.ods") == []

    def test_odf_load_failure_returns_empty(self, monkeypatch):
        odf_mods = _build_fake_odf_modules()
        odf_mods["odf.opendocument"].load = Mock(side_effect=RuntimeError("bad ods"))
        ex = FormulaExtractor()
        monkeypatch.setattr(ex, "_store_formulas", lambda *a, **k: None)
        with patch.dict(sys.modules, odf_mods):
            assert ex.extract_from_ods("book.ods") == []

    def test_parse_ods_formula_of_prefix(self):
        ex = FormulaExtractor()
        result = ex._parse_ods_formula(
            formula_str="of:SUM(A1:B1)",
            row=1,
            col=0,
            headers={1: "Sales", 2: "Revenue"},
            sheet_name="S",
        )
        assert result["expression"] == "sum(Sales, Revenue)"
        assert result["original_formula"] == "=SUM(A1:B1)"

    def test_parse_ods_formula_no_prefix(self):
        ex = FormulaExtractor()
        result = ex._parse_ods_formula(
            formula_str="=AVERAGE(A1:B1)",
            row=1,
            col=1,
            headers={1: "Sales", 2: "Revenue"},
            sheet_name="S",
        )
        assert result["expression"] == "average(Sales, Revenue)"
        assert result["source_cell"] == "R1C2"


class TestExtractFromSheetAndParse:
    def test_extract_from_sheet(self):
        header = [_FakeCell("Revenue", 1, "A1"), _FakeCell("Cost", 2, "B1")]
        rows = [
            [_FakeCell("=SUM(A2:B2)", 1, "A2"), _FakeCell("=AVERAGE(A2:B2)", 2, "B2")],
            [_FakeCell("=A3*B3", 1, "A3"), _FakeCell(100, 2, "B3")],
        ]
        ex = FormulaExtractor()
        formulas = ex._extract_from_sheet(_FakeSheet(header, rows), "Data")
        assert len(formulas) == 3
        by_cell = {f["source_cell"]: f for f in formulas}
        assert by_cell["A2"]["expression"] == "sum(Revenue, Cost)"
        assert by_cell["A2"]["name"] == "Revenue"
        assert by_cell["B2"]["expression"] == "average(Revenue, Cost)"
        assert by_cell["A3"]["expression"] == "Revenue * Cost"

    def test_get_column_headers_empty_header(self):
        header = [_FakeCell(None, 1, "A1"), _FakeCell("Cost", 2, "B1")]
        ex = FormulaExtractor()
        assert ex._get_column_headers(_FakeSheet(header, [])) == {2: "Cost"}

    def test_parse_formula_maps_cell_to_headers(self):
        ex = FormulaExtractor()
        cell = _FakeCell("=SUM(A2:B2)", 2, "B2")
        result = ex._parse_formula(
            cell=cell,
            formula_str="=SUM(A2:B2)",
            headers={1: "Sales", 2: "Revenue"},
            sheet_name="Sheet1",
        )
        assert result["expression"] == "sum(Sales, Revenue)"
        assert result["name"] == "Revenue"
        assert result["source_cell"] == "B2"
        assert result["source_sheet"] == "Sheet1"
        assert result["domain"] == "finance"


class TestSemanticExpressionRemainingBranches:
    def test_arithmetic_subtraction(self):
        ex = FormulaExtractor()
        result = ex._create_semantic_expression(
            "A1 - B1", "ARITHMETIC", [{"name": "a"}, {"name": "b"}]
        )
        assert result == "a - b"

    def test_arithmetic_division(self):
        ex = FormulaExtractor()
        result = ex._create_semantic_expression(
            "A1 / B1", "ARITHMETIC", [{"name": "a"}, {"name": "b"}]
        )
        assert result == "a / b"

    def test_custom_default_expression(self):
        ex = FormulaExtractor()
        result = ex._create_semantic_expression(
            "=FOO(A1)", "CUSTOM", [{"name": "a"}]
        )
        assert result == "custom(a)"

    def test_arithmetic_single_param_falls_to_default(self):
        ex = FormulaExtractor()
        result = ex._create_semantic_expression(
            "A1 + 1", "ARITHMETIC", [{"name": "a"}]
        )
        assert result == "arithmetic(a)"


class TestGenerateUseCaseRemainingBranches:
    def test_custom_use_case(self):
        ex = FormulaExtractor()
        result = ex._generate_use_case("GrossMargin", "CUSTOM", [{"name": "a"}])
        assert result == "Calculate GrossMargin using CUSTOM formula"


class TestStoreFormulas:
    def test_store_counts_successes_and_swallows_errors(self, monkeypatch):
        ex = FormulaExtractor("ws-store")
        fake_mgr = Mock()
        fake_mgr.add_formula.side_effect = [None, "f2", RuntimeError("boom")]
        monkeypatch.setattr(ex, "_get_formula_manager", lambda: fake_mgr)

        formulas = [
            {"expression": "sum(a)", "name": "T1", "domain": "general",
             "use_case": "u1", "parameters": []},
            {"expression": "avg(b)", "name": "T2", "domain": "finance",
             "use_case": "u2", "parameters": []},
            {"expression": "a*b", "name": "T3", "domain": "general",
             "use_case": "u3", "parameters": []},
        ]
        ex._store_formulas(formulas, user_id="user1", source_file="/tmp/Book1.xlsx")
        assert fake_mgr.add_formula.call_count == 3
        add_call = fake_mgr.add_formula.call_args_list[0]
        assert add_call.kwargs["source"] == "excel:Book1.xlsx"
        assert add_call.kwargs["user_id"] == "user1"

    def test_detect_formula_type_custom_with_ops_and_known(self):
        ex = FormulaExtractor()
        assert ex._detect_formula_type("=VLOOKUP(A1, B1:C1, 2, FALSE)") == "VLOOKUP"
        assert ex._detect_formula_type("=COUNT(A1:A2)") == "COUNT"
        assert ex._detect_formula_type("=MAX(A1)") == "MAX"
        assert ex._detect_formula_type("=MIN(A1)") == "MIN"
        assert ex._detect_formula_type("=SUMIF(A1:A2, 'x', B1:B2)") == "SUM"
        assert ex._detect_formula_type("=CONCATENATE(A1, B1)") == "CONCATENATE"
        assert ex._detect_formula_type("=A1+1") == "ARITHMETIC"
        assert ex._detect_formula_type("=FOO(A1)") == "CUSTOM"
        assert ex._detect_formula_type("=IF(A1>1, 1, 0)") == "IF"

    def test_extract_cell_references_dedupes(self):
        ex = FormulaExtractor()
        refs = ex._extract_cell_references("=SUM($A$1:B2)+A1")
        assert set(refs) == {("A", 1), ("B", 2)}

    def test_extract_cell_references_preserves_first_occurrence_order(self):
        """Regression: refs were deduped via a set, so parameter order of the
        semantic expression varied with PYTHONHASHSEED across runs."""
        ex = FormulaExtractor()
        assert ex._extract_cell_references("=SUM(A2:B2)") == [("A", 1), ("B", 2)]
        assert ex._extract_cell_references("=SUM($B$2:A2)") == [("B", 2), ("A", 1)]
        assert ex._extract_cell_references("=AVERAGE(C2:C2)") == [("C", 3)]


# ===========================================================================
# core/formula_memory.py — gap closure (before: 91%)
# ===========================================================================


@contextmanager
def _db_session_context(session):
    yield session


class TestFormulaMemoryExtend:
    def test_ensure_initialized_success_creates_table(self):
        lancedb = Mock()
        lancedb.get_table.return_value = None
        manager = FormulaMemoryManager("ws-init")
        with patch("core.lancedb_handler.get_lancedb_handler", return_value=lancedb):
            manager._ensure_initialized()
        assert manager._lancedb is lancedb
        assert manager._initialized is True
        lancedb.create_table.assert_called_once_with("formula_cards")

    def test_ensure_formulas_table_no_lancedb(self):
        manager = FormulaMemoryManager()
        manager._ensure_formulas_table()

    def test_ensure_formulas_table_get_table_raises(self):
        lancedb = Mock()
        lancedb.get_table.side_effect = Exception("table error")
        manager = FormulaMemoryManager()
        manager._lancedb = lancedb
        manager._initialized = True
        manager._ensure_formulas_table()

    def test_add_formula_fetches_dependency_names(self):
        lancedb = Mock()
        lancedb.get_table.return_value = None
        manager = FormulaMemoryManager("ws-deps")
        manager._lancedb = lancedb
        manager._initialized = True

        db = Mock()
        dep = Mock()
        dep.name = "Net Revenue"
        db.query.return_value.filter.return_value.all.return_value = [dep]

        with patch(
            "core.database.get_db_session",
            side_effect=lambda: _db_session_context(db),
        ):
            formula_id = manager.add_formula(
                expression="a + b",
                name="Total",
                domain="finance",
                dependencies=["dep-1"],
                example_input={"a": 1},
                example_output=3,
            )
        assert formula_id is not None
        card_text = lancedb.add_document.call_args[1]["text"]
        assert "Requires: Net Revenue" in card_text
        assert "dep-1" not in card_text
        metadata = lancedb.add_document.call_args[1]["metadata"]
        assert metadata["formula_id"] == formula_id
        assert metadata["type"] == "formula_card"

    def test_get_formula_database_error_returns_none(self):
        db = Mock()
        db.query.return_value.filter.return_value.first.side_effect = Exception("db down")
        manager = FormulaMemoryManager()
        with patch("core.database.get_db_session", return_value=_db_session_context(db)):
            assert manager.get_formula("f1") is None

    def test_apply_formula_unexpected_exception(self):
        manager = FormulaMemoryManager()
        formula = {
            "id": "f1",
            "name": "F",
            "expression": "a + b",
            "domain": "math",
            "parameters": [],
            "dependencies": [],
        }
        with patch.object(manager, "get_formula", return_value=formula), patch(
            "core.safe_evaluator.safe_eval_with_math", side_effect=RuntimeError("boom")
        ):
            result = manager.apply_formula("f1", {"a": 1})
        assert result["success"] is False
        assert result["error"] == "boom"

    def test_delete_formula_sql_error_returns_false(self):
        lancedb = Mock()
        lancedb.get_table.return_value = None
        manager = FormulaMemoryManager()
        manager._lancedb = lancedb
        manager._initialized = True

        db = Mock()
        db.query.return_value.filter.return_value.first.side_effect = Exception("db down")
        with patch("core.database.get_db_session", return_value=_db_session_context(db)):
            assert manager.delete_formula("f1") is False
