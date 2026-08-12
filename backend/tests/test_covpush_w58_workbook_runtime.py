"""Coverage wave 58 — core/workbook_runtime.py (0% → 90%+).

Engine detection (soffice/formulas/openpyxl), basic HTML render (success +
error), insert rows/cols (success + missing-sheet), evaluated range reads
(single + range + missing-sheet), formula result alias, singleton. Real
openpyxl workbooks on temp files; recalculate mocked for engine variance.
"""
import asyncio
import tempfile
from pathlib import Path
from unittest.mock import Mock, patch
from types import SimpleNamespace

import openpyxl
import pytest

from core.workbook_runtime import WorkbookRuntime, get_workbook_runtime


@pytest.fixture
def wb_path(tmp_path):
    p = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"
    wb.save(p)
    return p


def await_h(coro):
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


class TestEngineDetection:
    def test_libreoffice_engine(self):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        r._has_formulas = False
        assert r.engine == "libreoffice"
        assert r.can_evaluate is True
        assert r.can_render is True

    def test_formulas_engine(self):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = True
        assert r.engine == "formulas"
        assert r.can_evaluate is True
        assert r.can_render is False

    def test_openpyxl_fallback(self):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = False
        assert r.engine == "openpyxl"
        assert r.can_evaluate is False
        assert r.can_render is False

    def test_find_soffice(self):
        from core.workbook_runtime import _find_soffice
        with patch("shutil.which", return_value="/usr/bin/soffice"):
            assert _find_soffice() == "/usr/bin/soffice"
        with patch("shutil.which", return_value=None), \
             patch("os.path.exists", return_value=True):
            assert _find_soffice() is not None
        with patch("shutil.which", return_value=None), \
             patch("os.path.exists", return_value=False):
            assert _find_soffice() is None

    def test_check_formulas(self):
        r = WorkbookRuntime()
        with patch.dict("sys.modules", {"formulas": Mock()}):
            assert r._check_formulas() is True
        with patch("builtins.__import__", side_effect=ImportError("no")):
            assert r._check_formulas() is False


class TestRenderBasic:
    def test_render_success(self, wb_path):
        r = WorkbookRuntime()
        html = r._render_html_basic(wb_path)
        assert "Sheet1" in html
        assert "10" in html
        assert "<table" in html

    def test_render_error(self):
        r = WorkbookRuntime()
        html = r._render_html_basic(Path("/nonexistent.xlsx"))
        assert "Error rendering" in html


class TestStructuralOps:
    async def test_insert_rows_success(self, wb_path):
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.insert_rows(wb_path, "Sheet1", 1, 2)
        assert result["success"] is True
        assert result["rows_inserted"] == 2
        wb = openpyxl.load_workbook(wb_path)
        assert wb["Sheet1"].max_row >= 4

    async def test_insert_rows_missing_sheet(self, wb_path):
        r = WorkbookRuntime()
        result = await r.insert_rows(wb_path, "Nope", 1)
        assert result["success"] is False

    async def test_insert_cols_success_and_missing(self, wb_path):
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.insert_cols(wb_path, "Sheet1", 1, 3)
        assert result["success"] is True
        missing = await r.insert_cols(wb_path, "Nope", 1)
        assert missing["success"] is False


class TestEvaluatedReads:
    async def test_single_cell(self, wb_path):
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.get_evaluated_range(wb_path, "Sheet1", "A1")
        assert result["success"] is True
        assert result["values"][0][0]["value"] == 10

    async def test_range(self, wb_path):
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.get_evaluated_range(wb_path, "Sheet1", "A1", "A2")
        assert len(result["values"]) == 2

    async def test_missing_sheet(self, wb_path):
        r = WorkbookRuntime()
        result = await r.get_evaluated_range(wb_path, "Nope", "A1")
        assert result["success"] is False

    async def test_formula_result_alias(self, wb_path):
        r = WorkbookRuntime()
        with patch.object(r, "get_evaluated_range",
                          new=Mock(return_value=awaitable({"success": True}))):
            result = await r.get_formula_result(wb_path, "Sheet1", "A3")
        assert result["success"] is True


class TestSingleton:
    def test_singleton(self):
        with patch.object(WorkbookRuntime, "__new__",
                          return_value=object.__new__(WorkbookRuntime)):
            r1 = get_workbook_runtime()
            assert r1 is get_workbook_runtime()


def awaitable(value=None):
    async def _a():
        return value
    return _a()


class TestRecalculateBranches:
    async def test_missing_file_raises(self, tmp_path):
        r = WorkbookRuntime()
        p = tmp_path / "missing.xlsx"
        with pytest.raises(FileNotFoundError):
            await r.recalculate(p)

    async def test_soffice_recalc(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        with patch.object(r, "_recalc_with_soffice",
                          new=Mock(return_value=awaitable(wb_path))):
            assert await r.recalculate(wb_path) == wb_path

    async def test_formulas_recalc(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = True
        with patch.object(r, "_recalc_with_formulas",
                          new=Mock(return_value=awaitable(wb_path))):
            assert await r.recalculate(wb_path) == wb_path

    async def test_openpyxl_fallback_returns_path(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = False
        assert await r.recalculate(wb_path) == wb_path

    async def test_soffice_recalc_exec(self, tmp_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        p = tmp_path / "a.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        with patch("asyncio.create_subprocess_exec",
                   new=Mock(return_value=awaitable(
                       SimpleNamespace(returncode=0, communicate=Mock(
                           return_value=awaitable((b"", b""))))))):
            out = await r._recalc_with_soffice(p)
        assert out is not None

    async def test_formulas_recalc_exec(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = True
        with patch("builtins.__import__", side_effect=ImportError("no formulas")):
            assert await r._recalc_with_formulas(wb_path) == wb_path

    async def test_run_macro_gates(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = None
        missing = await r.run_macro(tmp_path_join(wb_path, "nope.xlsx"), "M1")
        assert missing["success"] is False
        result = await r.run_macro(wb_path, "M1")
        assert result["success"] is False  # no soffice -> macro unsupported

    async def test_run_macro_with_soffice(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        with patch("core.firecracker_sandbox.get_sandbox") as gs:
            sandbox = gs.return_value
            sandbox.execute_in_sandbox = Mock(return_value=awaitable(True))
            result = await r.run_macro(wb_path, "M1")
        assert result["success"] is True
        sandbox.execute_in_sandbox = Mock(return_value=awaitable(False))
        with patch("core.firecracker_sandbox.get_sandbox") as gs2:
            gs2.return_value = sandbox
            failed = await r.run_macro(wb_path, "M1")
        assert failed["success"] is False

    async def test_render_to_html_dispatches(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        with patch.object(r, "_render_html_with_soffice",
                          new=Mock(return_value=awaitable("<h1>x</h1>"))):
            assert await r.render_to_html(wb_path) == "<h1>x</h1>"
        r._soffice = None
        assert "<table" in await r.render_to_html(wb_path)


def tmp_path_join(wb_path, name):
    return wb_path.parent / name


class TestPivotAndRender:
    async def test_add_pivot_table_success(self, tmp_path):
        import pandas as pd
        p = tmp_path / "pivot.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["region", "product", "amount"])
        ws.append(["north", "a", 10])
        ws.append(["north", "b", 20])
        ws.append(["south", "a", 30])
        wb.save(p)
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.add_pivot_table(
                p, "Data", "Pivot", "A1:C4",
                ["region"], ["product"], [{"field": "amount", "function": "sum"}])
        assert result["success"] is True
        wb2 = openpyxl.load_workbook(p)
        assert "Pivot" in wb2.sheetnames

    async def test_add_pivot_table_missing_sheet(self, wb_path):
        r = WorkbookRuntime()
        result = await r.add_pivot_table(wb_path, "Nope", "P", "A1:A2",
                                         ["x"], ["y"], [{"field": "z"}])
        assert result["success"] is False

    async def test_add_pivot_table_empty_data(self, tmp_path):
        p = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        wb.save(p)
        r = WorkbookRuntime()
        result = await r.add_pivot_table(p, "Data", "P", "A1",
                                         ["x"], ["y"], [{"field": "z"}])
        assert result["success"] is False

    async def test_render_to_html_missing_file(self, tmp_path):
        r = WorkbookRuntime()
        assert await r.render_to_html(tmp_path / "nope.xlsx") == "<p>File not found</p>"

    async def test_render_soffice(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        html = r._render_html_basic  # keep ref
        with patch.object(r, "_render_html_with_soffice",
                          new=Mock(return_value=awaitable("<html>LO</html>"))):
            assert await r.render_to_html(wb_path) == "<html>LO</html>"

    async def test_render_soffice_fallback(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        with patch("asyncio.create_subprocess_exec",
                   new=Mock(side_effect=RuntimeError("no soffice"))):
            html = await r._render_html_with_soffice(wb_path)
        assert "<table" in html  # fell back to basic


class TestSofficeAndFormulasImpl:
    async def test_soffice_recalc_success(self, tmp_path):
        import shutil
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        p = tmp_path / "a.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)

        real_mkdtemp = tempfile.mkdtemp
        created = {}

        def fake_mkdtemp(*a, **k):
            d = real_mkdtemp(*a, **k)
            # pre-place the "output" file that soffice would produce
            shutil.copy2(str(p), str(Path(d) / p.name))
            created["dir"] = d
            return d

        with patch("tempfile.mkdtemp", side_effect=fake_mkdtemp), \
             patch("asyncio.create_subprocess_exec",
                   new=Mock(return_value=awaitable(
                       SimpleNamespace(returncode=0,
                                       communicate=Mock(
                                           return_value=awaitable((b"", b""))))))):
            out = await r._recalc_with_soffice(p)
        assert out == p

    async def test_soffice_recalc_no_output(self, tmp_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        p = tmp_path / "a.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        with patch("asyncio.create_subprocess_exec",
                   new=Mock(return_value=awaitable(
                       SimpleNamespace(returncode=0,
                                       communicate=Mock(
                                           return_value=awaitable((b"", b"no out"))))))):
            out = await r._recalc_with_soffice(p)
        assert out == p  # degrades gracefully

    async def test_soffice_recalc_timeout(self, tmp_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        p = tmp_path / "a.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        with patch("asyncio.wait_for", side_effect=asyncio.TimeoutError()):
            out = await r._recalc_with_soffice(p)
        assert out == p

    async def test_formulas_recalc_real(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = None
        r._has_formulas = True
        formulas_mod = Mock()
        xl_model = formulas_mod.ExcelModel.return_value
        xl_model.loads.return_value = xl_model
        xl_model.finish.return_value = xl_model
        xl_model.calculate.return_value = {
            f"'[{wb_path.name}]Sheet1'!A3": SimpleNamespace(value=30)}
        with patch.dict("sys.modules", {"formulas": formulas_mod}):
            out = await r._recalc_with_formulas(wb_path)
        assert out == wb_path

    async def test_pivot_overwrite_existing_sheet(self, tmp_path):
        p = tmp_path / "p2.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["region", "amount"])
        ws.append(["north", 5])
        ws.append(["south", 7])
        wb.create_sheet("Pivot")  # pre-existing target
        wb.save(p)
        r = WorkbookRuntime()
        with patch.object(r, "recalculate", new=Mock(return_value=awaitable())):
            result = await r.add_pivot_table(
                p, "Data", "Pivot", "A1:B3", ["region"], [],
                [{"field": "amount", "function": "sum"}])
        assert result["success"] is True

    async def test_render_soffice_success(self, wb_path):
        r = WorkbookRuntime()
        r._soffice = "/usr/bin/soffice"
        import tempfile as tf
        real_mkdtemp = tf.mkdtemp
        created = {}

        def fake_mkdtemp(*a, **k):
            d = real_mkdtemp(*a, **k)
            Path(d, wb_path.stem + ".html").write_text("<html>LO output</html>")
            created["dir"] = d
            return d

        with patch("tempfile.mkdtemp", side_effect=fake_mkdtemp), \
             patch("asyncio.create_subprocess_exec",
                   new=Mock(return_value=awaitable(
                       SimpleNamespace(communicate=Mock(
                           return_value=awaitable((b"", b""))))))):
            html = await r._render_html_with_soffice(wb_path)
        assert "LO output" in html
