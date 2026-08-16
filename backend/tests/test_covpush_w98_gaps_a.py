# -*- coding: utf-8 -*-
"""Coverage wave 98 — verified gap batch.

Targets (verified under 80% by existing suites):
1.  core/workbook_runtime.py            (60%)
2.  core/webhook_monitoring.py           (54%)
3.  core/communication/adapters/teams.py (40%)
4.  core/entity_type_service.py          (78%)
5.  core/atom_meta_agent.py              (29%)

No network, no real LLM, no real Redis — every external boundary
(httpx, DB sessions, subprocess/LibreOffice, sandbox, governance) is mocked.
Plain pytest + unittest.mock (asyncio_mode=auto).
"""
import asyncio
import base64
import hashlib
import hmac as hmac_mod
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace as NS
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import pytest

# --------------------------------------------------------------------------- #
# Workbook runtime
# --------------------------------------------------------------------------- #
class TestWorkbookRuntime:
    def _runtime(self, soffice=None, has_formulas=False):
        import core.workbook_runtime as wr
        rt = wr.WorkbookRuntime.__new__(wr.WorkbookRuntime)
        rt._soffice = soffice
        rt._has_formulas = has_formulas
        return rt

    def test_find_soffice_via_which(self):
        import core.workbook_runtime as wr
        with patch("shutil.which", side_effect=lambda n: f"/bin/{n}" if n == "soffice" else None):
            assert wr._find_soffice() == "/bin/soffice"

    def test_find_soffice_mac_path(self):
        import core.workbook_runtime as wr
        with patch("shutil.which", return_value=None), \
             patch("os.path.exists", return_value=True):
            assert wr._find_soffice() == "/Applications/LibreOffice.app/Contents/MacOS/soffice"

    def test_find_soffice_missing(self):
        import core.workbook_runtime as wr
        with patch("shutil.which", return_value=None), \
             patch("os.path.exists", return_value=False):
            assert wr._find_soffice() is None

    def test_engine_properties(self):
        assert self._runtime("/bin/soffice").engine == "libreoffice"
        assert self._runtime(None, True).engine == "formulas"
        rt = self._runtime(None, False)
        assert rt.engine == "openpyxl"
        assert not rt.can_evaluate
        assert not rt.can_render
        assert self._runtime("/bin/soffice").can_render
        assert self._runtime(None, True).can_evaluate

    async def test_recalculate_not_found(self):
        rt = self._runtime()
        with pytest.raises(FileNotFoundError):
            await rt.recalculate("/nonexistent/x.xlsx")

    async def test_recalculate_no_engine(self, tmp_path):
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"x")
        rt = self._runtime()
        assert await rt.recalculate(f) == f

    async def test_recalc_soffice_success(self, tmp_path):
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"old")

        async def fake_exec(*args, **kwargs):
            # args: soffice ... --outdir tmp file
            outdir = Path(args[args.index("--outdir") + 1])
            src = Path(args[-1])
            (outdir / src.name).write_bytes(b"recalced")
            proc = AsyncMock()
            proc.communicate = AsyncMock(return_value=(b"", b""))
            return proc

        async def fake_waitfor(c, timeout=None):
            return await c

        rt = self._runtime("/bin/soffice")
        with patch("asyncio.create_subprocess_exec", side_effect=fake_exec), \
             patch("asyncio.wait_for", side_effect=fake_waitfor):
            result = await rt.recalculate(f)
        assert result == f
        assert f.read_bytes() == b"recalced"

    async def test_recalc_soffice_no_output(self, tmp_path):
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"old")
        proc = AsyncMock()
        proc.communicate = AsyncMock(return_value=(b"", b"nope"))
        rt = self._runtime("/bin/soffice")
        with patch("asyncio.create_subprocess_exec", AsyncMock(return_value=proc)):
            result = await rt.recalculate(f)
        assert result == f
        assert f.read_bytes() == b"old"

    async def test_recalc_soffice_timeout_and_error(self, tmp_path):
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"x")
        rt = self._runtime("/bin/soffice")
        with patch("asyncio.create_subprocess_exec",
                   AsyncMock(side_effect=asyncio.TimeoutError())):
            assert await rt.recalculate(f) == f
        with patch("asyncio.create_subprocess_exec",
                   AsyncMock(side_effect=RuntimeError("boom"))):
            assert await rt.recalculate(f) == f

    async def test_recalc_formulas_hit_and_miss(self, tmp_path, monkeypatch):
        from openpyxl import Workbook
        f = tmp_path / "b.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["A2"] = "=A1+1"
        ws["A3"] = "=A1+2"
        wb.save(f)

        formulas = MagicMock()
        val = NS(value=42)
        plain = 7
        sol = {
            f"'[b.xlsx]Sheet'!A2": val,
            f"[b.xlsx]Sheet!A3": plain,
        }
        formulas.ExcelModel.return_value.loads.return_value.finish.return_value.calculate.return_value = sol
        import sys
        monkeypatch.setitem(sys.modules, "formulas", formulas)

        rt = self._runtime(None, True)
        assert await rt.recalculate(f) == f
        from openpyxl import load_workbook
        wb2 = load_workbook(f)
        assert wb2.active["A2"].value == 42
        assert wb2.active["A3"].value == 7

    async def test_recalc_formulas_exception(self, tmp_path, monkeypatch):
        f = tmp_path / "c.xlsx"
        f.write_bytes(b"x")
        formulas = MagicMock()
        formulas.ExcelModel.return_value.loads.side_effect = RuntimeError("bad")
        import sys
        monkeypatch.setitem(sys.modules, "formulas", formulas)
        rt = self._runtime(None, True)
        assert await rt.recalculate(f) == f

    async def test_run_macro_missing_and_no_soffice(self, tmp_path):
        rt = self._runtime()
        res = await rt.run_macro("/nope/x.xlsx", "M")
        assert res["success"] is False
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"x")
        res = await rt.run_macro(f, "M")
        assert "LibreOffice" in res["error"]

    async def test_run_macro_sandbox_paths(self, tmp_path):
        import core.workbook_runtime as wr
        f = tmp_path / "a.xlsx"
        f.write_bytes(b"x")
        rt = self._runtime("/bin/soffice")
        rt.recalculate = AsyncMock(return_value=f)

        sandbox = MagicMock()
        mod = MagicMock()
        mod.get_sandbox.return_value = sandbox
        mod.SandboxUnavailableError = type("SandboxUnavailableError", (Exception,), {})
        with patch.dict("sys.modules", {"core.firecracker_sandbox": mod}):
            sandbox.execute_in_sandbox = AsyncMock(side_effect=mod.SandboxUnavailableError("no fc"))
            res = await rt.run_macro(f, "M")
            assert res["success"] is False

            sandbox.execute_in_sandbox = AsyncMock(return_value=False)
            res = await rt.run_macro(f, "M")
            assert res["success"] is False

            sandbox.execute_in_sandbox = AsyncMock(return_value=True)
            res = await rt.run_macro(f, "M")
            assert res["success"] is True and res["macro"] == "M"

    def _make_book(self, tmp_path):
        from openpyxl import Workbook
        f = tmp_path / "p.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["region", "amount"])
        ws.append(["east", 10])
        ws.append(["west", 20])
        ws.append(["east", 30])
        wb.save(f)
        return f

    async def test_add_pivot_table_paths(self, tmp_path):
        rt = self._runtime()
        rt.recalculate = AsyncMock()
        f = self._make_book(tmp_path)
        res = await rt.add_pivot_table(
            f, "Sheet", "Pivot", "A1:B4", ["region"], [],
            [{"field": "amount", "function": "sum"}])
        assert res["success"] is True
        # source sheet missing
        res = await rt.add_pivot_table(f, "Nope", "P", None, [], [], [])
        assert res["success"] is False
        # exception path
        res = await rt.add_pivot_table(f, "Sheet", "P", None, None, [], [])
        assert res["success"] is False

    async def test_add_pivot_table_empty_sheet(self, tmp_path):
        from openpyxl import Workbook
        f = tmp_path / "e.xlsx"
        wb = Workbook()
        wb.create_sheet("Empty")
        wb.save(f)
        rt = self._runtime()
        res = await rt.add_pivot_table(f, "Empty", "P", None, [], [], [])
        assert res["success"] is False and "no data" in res["error"]

    async def test_render_to_html(self, tmp_path):
        rt = self._runtime()
        assert await rt.render_to_html("/nope/x.xlsx") == "<p>File not found</p>"
        f = self._make_book(tmp_path)
        html = await rt.render_to_html(f)
        assert "<table" in html and "east" in html
        # corrupt file -> error branch
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"notxlsx")
        assert await rt.render_to_html(bad) == "<p>Error rendering workbook</p>"

    async def test_render_html_with_soffice_fallback(self, tmp_path):
        f = self._make_book(tmp_path)
        proc = AsyncMock()
        proc.communicate = AsyncMock(return_value=(b"", b""))
        rt = self._runtime("/bin/soffice")
        with patch("asyncio.create_subprocess_exec", AsyncMock(return_value=proc)):
            html = await rt.render_to_html(f)
        assert "<table" in html  # no html file produced -> basic fallback
        with patch("asyncio.create_subprocess_exec",
                   AsyncMock(side_effect=RuntimeError("x"))):
            html = await rt.render_to_html(f)
        assert "<table" in html

    async def test_insert_rows_cols(self, tmp_path):
        rt = self._runtime()
        rt.recalculate = AsyncMock()
        f = self._make_book(tmp_path)
        res = await rt.insert_rows(f, "Sheet", 2, 2)
        assert res["success"] is True and res["rows_inserted"] == 2
        res = await rt.insert_rows(f, "Nope", 2)
        assert res["success"] is False
        res = await rt.insert_cols(f, "Sheet", 1)
        assert res["success"] is True
        res = await rt.insert_cols(f, "Nope", 1)
        assert res["success"] is False

    async def test_get_evaluated_range(self, tmp_path):
        rt = self._runtime()
        f = self._make_book(tmp_path)
        res = await rt.get_evaluated_range(f, "Sheet", "A1", "B2")
        assert res["success"] is True
        assert res["values"][0][0]["cell"] == "A1"
        res = await rt.get_evaluated_range(f, "Nope", "A1")
        assert res["success"] is False
        res = await rt.get_formula_result(f, "Sheet", "A1")
        assert res["success"] is True
        assert res["values"][0][0]["value"] == "region"

    def test_get_workbook_runtime_singleton(self):
        import core.workbook_runtime as wr
        wr._workbook_runtime = None
        rt = wr.get_workbook_runtime()
        assert wr.get_workbook_runtime() is rt
        wr._workbook_runtime = None


# --------------------------------------------------------------------------- #
# Webhook monitoring
# --------------------------------------------------------------------------- #
class TestWebhookMonitoring:
    @pytest.fixture(autouse=True)
    def _patch_deps(self, monkeypatch):
        import core.webhook_monitoring as wm
        monkeypatch.setattr(wm, "RedisCacheService", MagicMock)
        monkeypatch.setattr(wm, "CircuitBreaker", MagicMock)
        monkeypatch.setattr(wm, "WebhookMetrics", MagicMock)
        # reset singletons
        monkeypatch.setattr(wm, "_subscription_monitor", None)
        monkeypatch.setattr(wm, "_rate_limit_tracker", None)
        monkeypatch.setattr(wm, "_monitoring_service", None)
        self.wm = wm

    def test_make_key(self):
        m = self.wm.WebhookSubscriptionMonitor()
        assert m._make_key("tenant123456789", "slack") == "webhook:subscription:tenant12:slack"

    def test_track_and_get_status(self):
        m = self.wm.WebhookSubscriptionMonitor()
        m._cache = MagicMock()
        m.track_subscription("t" * 20, "slack", "sub-1",
                             datetime.now(timezone.utc) + timedelta(hours=10))
        st = m.get_subscription_status("t" * 20, "slack")
        assert st["subscription_id"] == "sub-1"
        assert st["is_expired"] is False
        assert st["hours_remaining"] > 9
        assert m.get_subscription_status("t" * 20, "github") is None

    def test_track_redis_failure_tolerated(self):
        m = self.wm.WebhookSubscriptionMonitor()
        cache = MagicMock()
        cache.set_async.side_effect = RuntimeError("redis down")
        m._cache = cache
        m.track_subscription("t", "slack", "s", datetime.now(timezone.utc) + timedelta(days=1))
        assert m.get_subscription_status("t", "slack") is not None

    def test_is_expired_and_hours(self):
        m = self.wm.WebhookSubscriptionMonitor()
        m.track_subscription("t", "a", "s", datetime.now(timezone.utc) - timedelta(hours=1))
        assert m.is_expired("t", "a") is True
        assert m.get_hours_remaining("t", "a") == 0
        assert m.is_expired("t", "zz") is False
        assert m.get_hours_remaining("t", "zz") == 0

    def test_check_expiration_alerts(self):
        m = self.wm.WebhookSubscriptionMonitor()
        m.track_subscription("t", "near72", "s", datetime.now(timezone.utc) + timedelta(hours=72))
        m.track_subscription("t", "near24", "s", datetime.now(timezone.utc) + timedelta(hours=24))
        m.track_subscription("t", "far", "s", datetime.now(timezone.utc) + timedelta(days=30))
        alerts = m.check_expiration_alerts()
        assert {a["connector_id"] for a in alerts} == {"near72", "near24"}
        assert all(a["alert_type"] == "subscription_expiring" for a in alerts)

    def test_rate_limit_tracker_update(self):
        t = self.wm.RateLimitTracker()
        t._cache = MagicMock()
        assert t._make_key("slack", "tenant1") == "webhook:ratelimit:tenant1:slack"
        t.update_from_headers("slack", "tenant1", {
            "X-RateLimit-Remaining": "10", "X-RateLimit-Limit": "100",
            "X-RateLimit-Reset": "1700000000"})
        st = t.get_rate_limit_status("slack", "tenant1")
        assert st["remaining"] == 10 and st["limit"] == 100
        assert st["percentage_remaining"] == pytest.approx(10.0)
        assert t.get_percentage_remaining("slack", "tenant1") == pytest.approx(10.0)
        assert t.is_quota_low("slack", "tenant1") is True
        assert t.get_rate_limit_status("slack", "other") is None
        assert t.get_percentage_remaining("slack", "other") == 100.0
        assert t.is_quota_low("slack", "other") is False

    def test_rate_limit_parse_variants(self):
        t = self.wm.RateLimitTracker()
        t._cache = MagicMock()
        t.update_from_headers("sh", "t1", {"ratelimit-remaining": "5", "ratelimit-limit": "0"})
        assert t.get_rate_limit_status("sh", "t1")["percentage_remaining"] == 0
        # missing headers -> no-op
        t.update_from_headers("sh", "t2", {"foo": "bar"})
        assert t.get_rate_limit_status("sh", "t2") is None
        # unparsable ints
        t.update_from_headers("sh", "t3", {"x-ratelimit-remaining": "abc", "x-ratelimit-limit": "10"})
        assert t.get_rate_limit_status("sh", "t3") is None
        # redis failure tolerated
        t2 = self.wm.RateLimitTracker()
        cache = MagicMock()
        cache.set_async.side_effect = RuntimeError("x")
        t2._cache = cache
        t2.update_from_headers("a", "b", {"x-ratelimit-remaining": "1", "x-ratelimit-limit": "2"})
        assert t2.get_rate_limit_status("a", "b") is not None
        # parse helper miss
        assert t2._parse_header({}, "remaining") is None
        assert t2._parse_header({"X-RateLimit-Limit": "9"}, "limit") == "9"
        assert t2._parse_header({}, "bogus") is None

    def test_monitoring_service(self):
        svc = self.wm.WebhookMonitoringService()
        svc.metrics = MagicMock()
        svc.record_webhook_failure("slack", "t", "transformation_error")
        svc.metrics.record_processing_error.assert_called_once()
        svc.update_rate_limits_from_response("slack", "t", {
            "x-ratelimit-remaining": "1", "x-ratelimit-limit": "4"})
        assert svc.get_rate_limit_status("slack", "t")["remaining"] == 1
        svc.track_subscription("t", "gh", "s", datetime.now(timezone.utc) + timedelta(days=5))
        assert svc.check_subscription_expirations() == []
        assert svc.get_circuit_state("slack") == "unknown"
        summary = svc.get_health_summary()
        assert summary["subscriptions_tracked"] == 1
        assert summary["rate_limits_tracked"] == 1

    def test_module_helpers_and_singletons(self):
        wm = self.wm
        mon = wm.get_subscription_monitor()
        assert wm.get_subscription_monitor() is mon
        trk = wm.get_rate_limit_tracker()
        assert wm.get_rate_limit_tracker() is trk
        svc = wm.get_monitoring_service()
        assert wm.get_monitoring_service() is svc
        mon.track_subscription("t9", "slack", "s", datetime.now(timezone.utc) + timedelta(days=2))
        assert wm.get_subscription_status("t9", "slack")["subscription_id"] == "s"
        healthy = wm.check_rate_limit_health("none", "none")
        assert healthy["healthy"] is True and healthy["percentage_remaining"] == 100.0
        trk.update_from_headers("slack", "t9", {"x-ratelimit-remaining": "5", "x-ratelimit-limit": "100"})
        h = wm.check_rate_limit_health("slack", "t9")
        assert h["healthy"] is False and h["remaining"] == 5


# --------------------------------------------------------------------------- #
# Teams adapter
# --------------------------------------------------------------------------- #
class TestTeamsAdapter:
    @pytest.fixture(autouse=True)
    def _reset(self, monkeypatch):
        import core.communication.adapters.teams as teams_mod
        monkeypatch.setattr(teams_mod, "_seen_jwt_ids", {})
        monkeypatch.delenv("MICROSOFT_APP_ID", raising=False)
        monkeypatch.delenv("MICROSOFT_APP_PASSWORD", raising=False)
        monkeypatch.delenv("ENVIRONMENT", raising=False)
        monkeypatch.delenv("BYPASS_WEBHOOK_SIGNATURE", raising=False)
        self.mod = teams_mod

    def _adapter(self, app_id="app-1", app_password="secret"):
        return self.mod.TeamsAdapter(app_id=app_id, app_password=app_password)

    def _req(self, auth=None):
        req = MagicMock()
        headers = MagicMock()
        headers.get = lambda k, d=None: (auth or "") if k == "Authorization" else (d or "")
        req.headers = headers
        return req

    async def test_jwks_cached_and_fetch(self):
        a = self._adapter()
        a.jwks_keys = ["k1"]
        a.jwks_expiry = time.time() + 1000
        assert await a._get_jwks_keys() == ["k1"]

        client = AsyncMock()
        client.get = AsyncMock(side_effect=[
            NS(raise_for_status=lambda: None, json=lambda: {"jwks_uri": "https://k"}),
            NS(raise_for_status=lambda: None, json=lambda: {"keys": [{"kid": "k"}]}),
        ])
        cm = MagicMock()
        cm.__aenter__ = AsyncMock(return_value=client)
        cm.__aexit__ = AsyncMock(return_value=False)
        a.jwks_keys = None
        with patch.object(self.mod.httpx, "AsyncClient", MagicMock(return_value=cm)):
            keys = await a._get_jwks_keys()
        assert keys == [{"kid": "k"}]

        # failure -> []
        a2 = self._adapter()
        bad = MagicMock(return_value=cm)
        cm.__aenter__ = AsyncMock(side_effect=RuntimeError("net"))
        with patch.object(self.mod.httpx, "AsyncClient", bad):
            assert await a2._get_jwks_keys() == []

    async def test_bot_token_paths(self):
        a = self._adapter()
        a._access_token = "tok"
        a._token_expiry = time.time() + 100
        assert await a._get_bot_access_token() == "tok"

        a2 = self._adapter(app_id=None, app_password=None)
        assert await a2._get_bot_access_token() is None

        client = AsyncMock()
        client.post = AsyncMock(return_value=NS(
            raise_for_status=lambda: None,
            json=lambda: {"access_token": "NEW", "expires_in": 3600}))
        cm = MagicMock()
        cm.__aenter__ = AsyncMock(return_value=client)
        cm.__aexit__ = AsyncMock(return_value=False)
        a3 = self._adapter()
        with patch.object(self.mod.httpx, "AsyncClient", MagicMock(return_value=cm)):
            assert await a3._get_bot_access_token() == "NEW"

        cm.__aenter__ = AsyncMock(side_effect=RuntimeError("net"))
        a3._access_token = None
        with patch.object(self.mod.httpx, "AsyncClient", MagicMock(return_value=cm)):
            assert await a3._get_bot_access_token() is None

    async def test_verify_request_guards(self, monkeypatch):
        a = self._adapter(app_id=None)
        assert await a.verify_request(self._req(), b"") is False
        monkeypatch.setenv("ENVIRONMENT", "development")
        monkeypatch.setenv("BYPASS_WEBHOOK_SIGNATURE", "true")
        assert await a.verify_request(self._req(), b"") is True

        b = self._adapter()
        assert await b.verify_request(self._req(), b"") is False  # no bearer
        assert await b.verify_request(self._req("Basic x"), b"") is False

        with patch.object(self.mod.jwt, "get_unverified_header", return_value={}):
            assert await b.verify_request(self._req("Bearer t"), b"") is False
        with patch.object(self.mod.jwt, "get_unverified_header", return_value={"kid": "k"}), \
             patch.object(self.mod.TeamsAdapter, "_get_jwks_keys", AsyncMock(return_value=[])):
            assert await b.verify_request(self._req("Bearer t"), b"") is False
        with patch.object(self.mod.jwt, "get_unverified_header", return_value={"kid": "k"}), \
             patch.object(self.mod.TeamsAdapter, "_get_jwks_keys",
                          AsyncMock(return_value=[{"kid": "other"}])):
            assert await b.verify_request(self._req("Bearer t"), b"") is False
        with patch.object(self.mod.jwt, "get_unverified_header", return_value={"kid": "k"}), \
             patch.object(self.mod.TeamsAdapter, "_get_jwks_keys",
                          AsyncMock(return_value=[{"kid": "k"}])), \
             patch.object(self.mod.jwk, "construct", MagicMock(side_effect=RuntimeError("bad jwk"))):
            assert await b.verify_request(self._req("Bearer t"), b"") is False

    def _good_key(self):
        return [{"kid": "k"}]

    async def test_verify_request_success_and_failures(self):
        a = self._adapter()
        now = time.time()
        good_claims = {"aud": "app-1", "exp": now + 600, "iat": now, "jti": "j1"}
        key = self._good_key()

        async def run(claims, header_kid="k"):
            with patch.object(self.mod.jwt, "get_unverified_header",
                              return_value={"kid": header_kid}), \
                 patch.object(self.mod.TeamsAdapter, "_get_jwks_keys",
                              AsyncMock(return_value=key)), \
                 patch.object(self.mod.jwk, "construct",
                              MagicMock(return_value=MagicMock(to_pem=lambda: b"pem"))), \
                 patch.object(self.mod.jwt, "decode", AsyncMock(return_value=claims)) \
                 if False else patch.object(self.mod.jwt, "decode",
                                            MagicMock(return_value=claims)):
                return await a.verify_request(self._req("Bearer tok"), b"")

        assert await run(dict(good_claims)) is True
        # expired claim
        assert await run({"aud": "app-1", "exp": now - 10}) is False
        # iat too old
        assert await run({"aud": "app-1", "exp": now + 600, "iat": now - 1000}) is False
        # replay: same jti seen already
        self.mod._seen_jwt_ids["j9"] = now
        assert await run({"aud": "app-1", "exp": now + 600, "iat": now, "jti": "j9"}) is False
        # jose errors
        with patch.object(self.mod.jwt, "get_unverified_header", return_value={"kid": "k"}), \
             patch.object(self.mod.TeamsAdapter, "_get_jwks_keys",
                          AsyncMock(return_value=key)), \
             patch.object(self.mod.jwk, "construct",
                          MagicMock(return_value=MagicMock(to_pem=lambda: b"pem"))), \
             patch.object(self.mod.jwt, "decode",
                          MagicMock(side_effect=self.mod.jwt.ExpiredSignatureError())):
            assert await a.verify_request(self._req("Bearer tok"), b"") is False
        with patch.object(self.mod.jwt, "get_unverified_header", return_value={"kid": "k"}), \
             patch.object(self.mod.TeamsAdapter, "_get_jwks_keys",
                          AsyncMock(return_value=key)), \
             patch.object(self.mod.jwk, "construct",
                          MagicMock(return_value=MagicMock(to_pem=lambda: b"pem"))), \
             patch.object(self.mod.jwt, "decode",
                          MagicMock(side_effect=self.mod.jwt.JWTError("bad"))):
            assert await a.verify_request(self._req("Bearer tok"), b"") is False
        with patch.object(self.mod.jwt, "get_unverified_header",
                          MagicMock(side_effect=ValueError("boom"))):
            assert await a.verify_request(self._req("Bearer tok"), b"") is False

    def test_verify_webhook_signature(self):
        a = self._adapter()

        def sig_for(payload, ts, password="secret"):
            msg = f"{ts}.{payload.decode()}"
            return base64.b64encode(
                hmac_mod.new(password.encode(), msg.encode(), hashlib.sha256).digest()
            ).decode()

        now = str(time.time())
        payload = b'{"a":1}'
        good = sig_for(payload, now)
        assert a.verify_webhook_signature(payload, good, now) is True
        assert a.verify_webhook_signature(payload, sig_for(payload, str(time.time() - 999)), now) is False
        assert a.verify_webhook_signature(payload, good, "not-a-float") is False
        assert a.verify_webhook_signature(payload, "", now) is False
        assert a.verify_webhook_signature(payload, good, now, ) is True
        no_pwd = self._adapter(app_password=None)
        assert no_pwd.verify_webhook_signature(payload, "x", now) is False
        # prefixes
        assert a.verify_webhook_signature(payload, "Bearer " + good, now) is True
        assert a.verify_webhook_signature(payload, "HMAC " + good, now) is True
        assert a.verify_webhook_signature(payload, "!!!bad-b64!!!", now) is False
        # wrong hmac
        assert a.verify_webhook_signature(payload, sig_for(payload, now, "other"), now) is False

    def test_normalize_payload(self):
        a = self._adapter()
        assert a.normalize_payload("notadict") == {}
        assert a.normalize_payload({"type": "typing"}) == {}
        out = a.normalize_payload({
            "type": "message", "text": "hi",
            "from": {"id": "u1", "name": "User"},
            "conversation": {"id": "c1"}, "serviceUrl": "https://s",
            "id": "act-1",
        })
        assert out["platform"] == "teams"
        assert out["user_id"] == "u1"
        assert out["channel_id"] == "c1"
        assert out["content"] == "hi"
        assert out["metadata"]["serviceUrl"] == "https://s"

    async def test_send_message(self):
        a = self._adapter()
        assert await a.send_message("c1", "hi", None) is False
        assert await a.send_message("c1", "hi", {}) is False

        a._access_token = "tok"
        a._token_expiry = time.time() + 100
        client = AsyncMock()
        client.post = AsyncMock(return_value=NS(raise_for_status=lambda: None, json=lambda: {}))
        cm = MagicMock()
        cm.__aenter__ = AsyncMock(return_value=client)
        cm.__aexit__ = AsyncMock(return_value=False)
        with patch.object(self.mod.httpx, "AsyncClient", MagicMock(return_value=cm)):
            ok = await a.send_message("c1", "hi", {"serviceUrl": "https://s/"})
        assert ok is True

        client.post = AsyncMock(side_effect=RuntimeError("net"))
        with patch.object(self.mod.httpx, "AsyncClient", MagicMock(return_value=cm)):
            ok = await a.send_message("c1", "hi", {"serviceUrl": "https://s"})
        assert ok is False


# --------------------------------------------------------------------------- #
# Entity type service
# --------------------------------------------------------------------------- #
class TestEntityTypeGaps:
    @pytest.fixture()
    def service(self):
        from core.entity_type_service import EntityTypeService
        validator = Mock()
        validator.validate_schema = Mock(return_value=(True, ""))
        factory = Mock()
        factory.invalidate_cache = Mock(return_value=0)
        db = Mock()
        db.add = Mock()
        db.flush = Mock()
        db.commit = Mock()
        db.refresh = Mock()
        db.delete = Mock()
        db.rollback = Mock()
        db.close = Mock()
        return EntityTypeService(db=db, schema_validator=validator, model_factory=factory)

    @staticmethod
    def _etype(**kw):
        defaults = dict(
            id="et-1", tenant_id="t1", slug="customer", display_name="Customer",
            json_schema={"type": "object",
                         "properties": {"name": {"type": "string"},
                                        "age": {"type": "integer"}},
                         "required": ["name"]},
            is_system=False, is_active=True, version=1, metadata_json=None,
            available_skills=None, description=None,
        )
        defaults.update(kw)
        return NS(**defaults)

    @staticmethod
    def _chain(db, first=None, all_result=None, count=0):
        q = Mock()
        q.filter = Mock(return_value=q)
        q.order_by = Mock(return_value=q)
        q.limit = Mock(return_value=q)
        q.offset = Mock(return_value=q)
        q.first = Mock(return_value=first)
        q.all = Mock(return_value=all_result or [])
        q.count = Mock(return_value=count)
        db.query = Mock(return_value=q)
        return q

    def test_merge_metadata_none_branch(self, service):
        src = self._etype(id="src", slug="draft_type", metadata_json={"discovery_reasoning": "r"})
        tgt = self._etype(slug="customer", metadata_json=None)
        service.get_entity_type = Mock(side_effect=[src, tgt])
        q = self._chain(service.db)
        q.update = Mock(return_value=3)
        assert service.merge_entity_types("t1", "src", "customer") is True
        assert tgt.metadata_json["merges"][0]["nodes_count"] == 3
        assert src.is_active is False

    def test_create_entity_type(self, service):
        # invalid slug
        with pytest.raises(ValueError, match="Invalid slug"):
            service.create_entity_type("t1", "bad slug!", "Name", {})
        # invalid schema
        service.validator.validate_schema = Mock(return_value=(False, "bad schema"))
        with pytest.raises(ValueError, match="Invalid JSON Schema"):
            service.create_entity_type("t1", "ok-slug", "Name", {})
        service.validator.validate_schema = Mock(return_value=(True, ""))
        # duplicate
        self._chain(service.db, first=self._etype())
        with pytest.raises(ValueError, match="already exists"):
            service.create_entity_type("t1", "customer", "Customer", {"type": "object"})
        # success
        self._chain(service.db, first=None)
        et = service.create_entity_type("t1", "invoice", "Invoice", {"type": "object"},
                                        description="d", available_skills=["s1"])
        assert et.slug == "invoice" and et.version == 1
        # commit failure
        service.db.commit = Mock(side_effect=RuntimeError("db"))
        with pytest.raises(RuntimeError):
            service.create_entity_type("t1", "invoice2", "I2", {"type": "object"})
        service.db.commit = Mock()

    def test_get_entity_type_branches(self, service):
        with pytest.raises(ValueError):
            service.get_entity_type("t1")
        et = self._etype()
        q = self._chain(service.db, first=et)
        assert service.get_entity_type("t1", entity_type_id="et-1") is et
        assert service.get_entity_type("t1", slug="customer") is et
        self._chain(service.db, first=None)
        assert service.get_entity_type("t1", slug="nope") is None

    def test_list_entity_types(self, service):
        ets = [self._etype(), self._etype(slug="other")]
        q = self._chain(service.db, all_result=ets)
        out = service.list_entity_types("t1", include_system=True, is_active=None,
                                        search="cust", limit=10, offset=5)
        assert out == ets
        assert q.limit.call_args == ((10,),)

    def test_update_entity_type(self, service):
        # not found
        service.get_entity_type = Mock(return_value=None)
        with pytest.raises(ValueError, match="not found"):
            service.update_entity_type("t1", "zz", json_schema={"type": "object"})
        # system type
        service.get_entity_type = Mock(return_value=self._etype(is_system=True))
        with pytest.raises(ValueError, match="read-only"):
            service.update_entity_type("t1", "et-1", description="x")
        # invalid schema
        et = self._etype()
        service.get_entity_type = Mock(return_value=et)
        service.validator.validate_schema = Mock(return_value=(False, "nope"))
        with pytest.raises(ValueError, match="Invalid JSON Schema"):
            service.update_entity_type("t1", "et-1", json_schema={"type": "object"})
        # success with schema change -> snapshot + invalidate + version bump
        service.validator.validate_schema = Mock(return_value=(True, ""))
        out = service.update_entity_type(
            "t1", "et-1", display_name="Cust2", json_schema={
                "type": "object", "properties": {"name": {"type": "string"}}},
            description="d", available_skills=["s"],
            changed_by="me", change_summary="cs")
        assert out.version == 2 and out.display_name == "Cust2"
        service.db.add.assert_called()
        service.model_factory.invalidate_cache.assert_called_once()
        # commit failure
        service.db.commit = Mock(side_effect=RuntimeError("x"))
        with pytest.raises(RuntimeError):
            service.update_entity_type("t1", "et-1", description="y")
        service.db.commit = Mock()

    def test_delete_entity_type(self, service):
        service.get_entity_type = Mock(return_value=None)
        assert service.delete_entity_type("t1", "zz") is False
        service.get_entity_type = Mock(return_value=self._etype(is_system=True))
        with pytest.raises(ValueError, match="read-only"):
            service.delete_entity_type("t1", "et-1")
        # hard delete success/failure
        et = self._etype()
        service.get_entity_type = Mock(return_value=et)
        assert service.delete_entity_type("t1", "et-1", hard_delete=True) is True
        service.db.commit = Mock(side_effect=RuntimeError("x"))
        with pytest.raises(RuntimeError):
            service.delete_entity_type("t1", "et-1", hard_delete=True)
        # soft delete success/failure
        service.db.commit = Mock()
        et2 = self._etype()
        service.get_entity_type = Mock(return_value=et2)
        assert service.delete_entity_type("t1", "et-1") is True
        assert et2.is_active is False
        service.db.commit = Mock(side_effect=RuntimeError("x"))
        with pytest.raises(RuntimeError):
            service.delete_entity_type("t1", "et-1")
        service.db.commit = Mock()

    def test_count_entity_types(self, service):
        q = self._chain(service.db, count=7)
        assert service.count_entity_types("t1") == 7
        assert service.count_entity_types("t1", include_system=True) == 7

    def test_detect_breaking_optional_removal(self, service):
        et = self._etype(json_schema={
            "properties": {"a": {"type": "string"}, "b": {"type": "string"}},
            "required": ["a"]})
        service.get_entity_type = Mock(return_value=et)
        res = service.detect_breaking_changes(
            "t1", "et-1",
            {"properties": {"a": {"type": "string"}}, "required": ["a"]})
        kinds = {c["type"]: c["severity"] for c in res["changes"]}
        assert kinds.get("property_removed") == "info"

    def test_close_with_owned_session(self):
        import core.entity_type_service as ets
        with patch.object(ets, "SessionLocal", return_value=MagicMock()) as sl, \
             patch.object(ets, "get_schema_validator", return_value=Mock()), \
             patch.object(ets, "get_model_factory", return_value=Mock()):
            svc = ets.EntityTypeService()
            svc.close()
            svc.db.close.assert_called_once()
        with svc:
            pass


# --------------------------------------------------------------------------- #
# Atom meta agent
# --------------------------------------------------------------------------- #
@pytest.fixture
def meta_agent(monkeypatch):
    import core.atom_meta_agent as ama
    monkeypatch.setattr(ama, "WorldModelService", MagicMock())
    monkeypatch.setattr(ama, "CapabilityGraduationService", MagicMock())
    monkeypatch.setattr(ama, "get_canvas_provider", MagicMock(return_value=MagicMock()))
    monkeypatch.setattr(ama, "mcp_service", MagicMock())
    monkeypatch.setattr(ama, "AgentGovernanceService", MagicMock())
    monkeypatch.setattr(ama, "AgentFleetService", MagicMock())
    monkeypatch.setattr(ama, "FleetOptimizationService", MagicMock())
    monkeypatch.setattr(ama, "_TURN_FACT_VECTOR_RECALL_ENABLED", False)
    monkeypatch.setattr(ama, "_TURN_FACT_EXTRACTION_ENABLED", False)

    sl = MagicMock()
    sl.return_value.__enter__.return_value = MagicMock()
    monkeypatch.setattr(ama, "SessionLocal", sl)

    agent = ama.AtomMetaAgent()
    agent.llm = MagicMock()
    agent.world_model = MagicMock()
    agent.graduation_service = MagicMock()
    return agent, sl, ama


class TestAtomMetaAgentGaps:
    def test_retrieve_skill_instructions(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        assert agent._retrieve_skill_instructions("r") == ""  # flag off (default)
        import core.hallucination_config as hc
        monkeypatch.setattr(hc, "is_skill_injection_enabled", lambda: True, raising=False)
        import core.skill_retrieval_service as srs
        monkeypatch.setattr(
            srs, "get_skill_retrieval_service",
            lambda: MagicMock(retrieve_top_skills=lambda *a, **k: "SKILLS"), raising=False)
        assert agent._retrieve_skill_instructions("r") == "SKILLS"
        monkeypatch.setattr(
            srs, "get_skill_retrieval_service",
            lambda: MagicMock(retrieve_top_skills=Mock(side_effect=RuntimeError("db"))))
        assert agent._retrieve_skill_instructions("r") == ""

    async def test_check_budget_before_react(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        import core.budget_enforcement_service as bes
        svc = MagicMock()
        svc.check_budget_before_action = AsyncMock(return_value={"allowed": False})
        cm = MagicMock()
        cm.__enter__ = lambda s: svc
        cm.__exit__ = lambda *a: False
        monkeypatch.setattr(bes, "BudgetEnforcementService", lambda: cm, raising=False)
        assert await agent._check_budget_before_react() == {"allowed": False}
        monkeypatch.setattr(bes, "BudgetEnforcementService",
                            MagicMock(side_effect=RuntimeError("x")), raising=False)
        res = await agent._check_budget_before_react()
        assert res["allowed"] is True and res["reason"] == "budget-check-error"

    async def test_react_step_structured(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        from core.react_models import ReActStep
        step = ReActStep(thought="t", final_answer="done")
        agent.llm.generate_structured_response = AsyncMock(return_value=step)
        agent._get_communication_instruction = lambda c: ""
        agent._retrieve_skill_instructions = lambda r: ""
        monkeypatch.setattr(ama, "_get_active_facts_for_prompt",
                            lambda db, w, limit=5: ["f"])
        out = await agent._react_step(
            "req", {"experiences": [NS(input_summary="sum", outcome="ok")],
                    "knowledge": [{"text": "k"}], "formulas": [{"name": "F", "description": "d"}],
                    "business_facts": [NS(verification_status="verified", fact="f",
                                          metadata={"source": "s"})],
                    "canvas_episodes": [{"canvas_id": "c12345678", "task_description": "td",
                                          "outcome": "ok", "canvas_boost": 0.5}]},
            "TOOLS", "HIST", {}, canvas_text="CANVAS")
        assert out is step
        agent.llm.generate_structured_response.assert_awaited_once()
        prompt = agent.llm.generate_structured_response.await_args.kwargs["prompt"]
        assert "PAST EXPERIENCES" in prompt and "CANVAS EPISODES" in prompt

    async def test_react_step_fallback(self, meta_agent):
        agent, sl, ama = meta_agent
        agent._get_communication_instruction = lambda c: ""
        agent._retrieve_skill_instructions = lambda r: ""
        agent.llm.generate_structured_response = AsyncMock(return_value=None)
        from core.react_models import ReActStep
        # error content -> restriction step
        agent.llm.generate_completion = AsyncMock(return_value={"content": "LLM not initialized"})
        out = await agent._react_step("req", {}, "T", "", {})
        assert out.final_answer == "LLM not initialized"
        # None content
        agent.llm.generate_completion = AsyncMock(return_value={"content": None})
        out = await agent._react_step("req", {}, "T", "", {})
        assert "AI provider unavailable" in out.final_answer
        # normal content
        agent.llm.generate_completion = AsyncMock(return_value={"content": "all good"})
        out = await agent._react_step("req", {}, "T", "", {})
        assert out.final_answer == "all good"

    async def test_trigger_workflow(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        assert "required" in await agent._trigger_workflow(None, {}, {})
        import core.workflow_engine as we
        eng = MagicMock()
        eng.start_workflow = AsyncMock(return_value="exec-1")
        monkeypatch.setattr(we, "get_workflow_engine", lambda: eng, raising=False)
        assert "exec-1" in await agent._trigger_workflow("wf-1", {"a": 1}, {})
        monkeypatch.setattr(we, "get_workflow_engine",
                            MagicMock(side_effect=RuntimeError("x")), raising=False)
        assert "Error" in await agent._trigger_workflow("wf-1", {}, {})

    async def test_execute_delegation(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        import core.business_agents as ba
        monkeypatch.setattr(ba, "get_specialized_agent", lambda n, w: None, raising=False)
        assert "not found" in await agent._execute_delegation("nope", "t", {})
        sub = MagicMock()
        sub.name = "SalesAgent"
        sub.execute = AsyncMock(return_value={"final_output": "done"})
        monkeypatch.setattr(ba, "get_specialized_agent", lambda n, w: sub, raising=False)
        out = await agent._execute_delegation("sales", "t", {})
        assert "SalesAgent" in out and "done" in out
        monkeypatch.setattr(ba, "get_specialized_agent",
                            MagicMock(side_effect=RuntimeError("x")), raising=False)
        assert "failed" in await agent._execute_delegation("sales", "t", {})

    async def test_execute_tool_governance_special_tools(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        agent._trigger_workflow = AsyncMock(return_value="wf ok")
        assert await agent._execute_tool_with_governance(
            "trigger_workflow", {"workflow_id": "w"}, {}, None, pre_approved=True) == "wf ok"

        agent._execute_delegation = AsyncMock(return_value="del ok")
        assert "del ok" == await agent._execute_tool_with_governance(
            "delegate_task", {"agent_name": "sales", "task": "t"}, {}, None, pre_approved=True)

        agent._recruit_fleet = AsyncMock(return_value="fleet ok")
        assert "fleet ok" == await agent._execute_tool_with_governance(
            "recruit_fleet", {"sub_tasks": []}, {}, None, pre_approved=True)

        # invoke_capability: student blocked
        agent.graduation_service.get_maturity.return_value = "student"
        out = await agent._execute_tool_with_governance(
            "invoke_capability", {"capability_name": "cap"}, {}, None, pre_approved=True)
        assert "STUDENT" in out
        # invoke_capability: executes + graduation record
        agent.graduation_service.get_maturity.return_value = "autonomous"
        agent.mcp.call_tool = AsyncMock(return_value={"success": True})
        out = await agent._execute_tool_with_governance(
            "invoke_capability", {"capability_name": "cap"}, {}, None, pre_approved=True)
        agent.graduation_service.record_usage.assert_called_once()
        # graduation parse failure tolerated
        agent.mcp.call_tool = AsyncMock(return_value="plain string")
        await agent._execute_tool_with_governance(
            "invoke_capability", {"capability_name": "cap"}, {}, None, pre_approved=True)
        assert agent.graduation_service.record_usage.call_count == 2

    async def test_execute_tool_governance_sandbox_and_judge(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        agent.mcp.call_tool = AsyncMock(return_value="res")
        # enforced sandbox review blocks
        dec = MagicMock(requires_review=True, enforced=True, decision="blocked",
                        violation_detail="v")
        monkeypatch.setattr(ama, "_meta_agent_sandbox_check", lambda *a, **k: dec)
        out = await agent._execute_tool_with_governance("t", {}, {}, None, pre_approved=True)
        assert "Sandbox blocked" in out
        # shadow mode proceeds
        dec2 = MagicMock(requires_review=True, enforced=False, decision="review",
                         violation_detail="v", violation_type="vt")
        monkeypatch.setattr(ama, "_meta_agent_sandbox_check", lambda *a, **k: dec2)
        out = await agent._execute_tool_with_governance("t", {}, {}, None, pre_approved=True)
        assert out == "res"
        # generic exception -> tool error string
        agent.mcp.call_tool = AsyncMock(side_effect=RuntimeError("x"))
        out = await agent._execute_tool_with_governance("t", {}, {}, None, pre_approved=True)
        assert out == "Tool error. Please try again."
        # KillRun re-raised
        from core.sandbox_killrun import KillRunAborted
        agent.mcp.call_tool = AsyncMock(side_effect=KillRunAborted("kill"))
        with pytest.raises(KillRunAborted):
            await agent._execute_tool_with_governance("t", {}, {}, None, pre_approved=True)

    async def test_execute_tool_governance_hitl(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        db = sl.return_value
        gov = ama.AgentGovernanceService.return_value
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": True, "action_complexity": 5, "reason": "complex"})
        gov.request_approval = MagicMock(return_value="act-1")
        agent._wait_for_approval = AsyncMock(return_value=False)
        cb = AsyncMock()
        out = await agent._execute_tool_with_governance("big_tool", {}, {}, cb)
        assert "REJECTED" in out
        hitl_calls = [c.args[0] for c in cb.await_args_list
                      if c.args and isinstance(c.args[0], dict)
                      and c.args[0].get("type") == "hitl_paused"]
        assert hitl_calls and hitl_calls[0]["action_id"] == "act-1"
        assert "Propose-Only" in hitl_calls[0]["reason"]
        # allowed simple tool flows to MCP
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": True, "action_complexity": 1})
        agent.mcp.call_tool = AsyncMock(return_value="ok")
        assert await agent._execute_tool_with_governance("small", {}, {}, None) == "ok"
        # governance disallows
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": False, "reason": "no", "action_complexity": 1})
        out = await agent._execute_tool_with_governance("small", {}, {}, None)
        assert "Governance blocked" in out

    async def test_recruit_fleet(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        db = sl.return_value.__enter__.return_value
        fleet = ama.AgentFleetService.return_value
        chain = MagicMock()
        chain.id = "chain-1"
        fleet.initialize_fleet = MagicMock(return_value=chain)
        fleet.recruit_member = MagicMock(return_value=MagicMock())
        opt = ama.FleetOptimizationService.return_value
        opt.get_optimization_parameters = MagicMock(
            return_value={"optimization_reason": "r"})
        import core.business_agents as ba
        specialist = MagicMock()
        specialist.name = "SalesAgent"
        specialist.id = "sp-1"
        monkeypatch.setattr(ba, "get_specialized_agent", lambda d, w: specialist, raising=False)
        import core.agent_radio.radio_adapter as ra
        monkeypatch.setattr(ra, "attach_thread_for_chain",
                            lambda *a, **k: (_ for _ in ()).throw(RuntimeError("no radio")),
                            raising=False)
        db.query.return_value.filter.return_value.all.return_value = []
        cb = AsyncMock()
        out = await agent._recruit_fleet("goal", [
            {"domain": "sales", "task": "do sales", "use_optimizer": True},
            {"domain": "ops", "task": "do ops", "use_optimizer": False}], {}, cb)
        assert "chain-1" in out and "SalesAgent" in out
        cb.assert_awaited_once()
        # failure path
        ama.AgentFleetService.side_effect = RuntimeError("db")
        out = await agent._recruit_fleet("g", [], {})
        assert "failed" in out

    async def test_spawn_agent(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        with pytest.raises(ValueError):
            await agent.spawn_agent("unknown-template")
        # ephemeral template spawn
        agent.spawned_agents.clear()
        spawned = await agent.spawn_agent("sales_assistant")
        assert spawned.workspace_id == agent.workspace_id
        assert list(agent.spawned_agents.values())
        # custom
        custom = await agent.spawn_agent("custom",
                                         custom_params={"name": "MyAgent"})
        assert custom.name == "MyAgent"
        # persist with provided db
        gov = ama.AgentGovernanceService.return_value
        gov.register_or_update_agent = MagicMock(return_value=NS(id="reg-1"))
        persisted = await agent.spawn_agent("custom", custom_params={"name": "P"},
                                            persist=True, db=MagicMock())
        assert persisted.id == "reg-1"
        # persist without db -> fresh session
        persisted = await agent.spawn_agent("custom", custom_params={"name": "P2"},
                                            persist=True, db=None)
        assert persisted.id == "reg-1"

    async def test_query_memory_scopes(self, meta_agent):
        agent, sl, ama = meta_agent
        agent.world_model.recall_experiences = AsyncMock(return_value={
            "experiences": [1], "knowledge": [2]})
        assert (await agent.query_memory("q", "experiences")) == {"experiences": [1]}
        assert (await agent.query_memory("q", "knowledge")) == {"knowledge": [2]}
        assert (await agent.query_memory("q")) == {"experiences": [1], "knowledge": [2]}

    async def test_generate_mentorship_guidance(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        db = sl.return_value.__enter__.return_value
        # no student found -> interim supervisor
        db.query.return_value.filter.return_value.first.return_value = None
        db.query.return_value.filter.return_value.count.return_value = 0
        agent.llm.generate_response = AsyncMock(return_value="guide")
        out = await agent.generate_mentorship_guidance("stu-1", "act", {}, "why")
        assert out == "guide"
        assert "Interim Supervisor" in agent.llm.generate_response.await_args.kwargs["system_instruction"]
        # llm returns None
        agent.llm.generate_response = AsyncMock(return_value=None)
        out = await agent.generate_mentorship_guidance("stu-1", "act", {}, "why")
        assert "unable to provide guidance" in out
        # db exception inside thread -> fail open to interim
        def boom(*a, **k):
            raise RuntimeError("db")
        sl.side_effect = RuntimeError("db")
        agent.llm.generate_response = AsyncMock(return_value="g2")
        out = await agent.generate_mentorship_guidance("stu-1", "act", {}, "why")
        assert out == "g2"

    async def test_wait_for_approval(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        gov = ama.AgentGovernanceService.return_value
        gov.get_approval_status = MagicMock(return_value={"status": "approved"})
        with patch.object(ama.asyncio, "sleep", AsyncMock()):
            assert await agent._wait_for_approval("a") is True
        from core.models import HITLActionStatus
        gov.get_approval_status = MagicMock(
            return_value={"status": HITLActionStatus.REJECTED.value})
        assert await agent._wait_for_approval("a") is False

    async def test_wait_for_all_approvals(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        from core.models import HITLActionStatus
        gov = ama.AgentGovernanceService.return_value
        gov.get_approval_status = MagicMock(
            return_value={"status": HITLActionStatus.APPROVED.value})
        assert await agent._wait_for_all_approvals(["a", "b"]) is True
        gov.get_approval_status = MagicMock(
            return_value={"status": HITLActionStatus.REJECTED.value})
        assert await agent._wait_for_all_approvals(["a"]) is False
        # pending forever -> timeout returns False (sleep patched so it
        # exhausts max_wait instantly without real delay is not possible;
        # instead shrink elapsed growth is impractical — simulate via
        # rejection check after one sleep round)
        gov.get_approval_status = MagicMock(return_value={"status": "pending"})
        with patch.object(ama.asyncio, "sleep", AsyncMock(return_value=None)) as sl_mock:
            sl_mock.side_effect = RuntimeError("stop looping")
            with pytest.raises(RuntimeError):
                await agent._wait_for_all_approvals(["a"])

    async def test_execute_parallel_tools_disabled(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        import core.hallucination_config as hc
        monkeypatch.setattr(hc, "is_parallel_tools_enabled", lambda: False, raising=False)
        monkeypatch.setattr(hc, "get_max_parallel_tools", lambda: 4, raising=False)
        agent._execute_tool_with_governance = AsyncMock(return_value="obs")
        acts = [ama.ToolCall(tool="t1", params={}), ama.ToolCall(tool="t2", params={})]
        recs = await agent._execute_parallel_tools(acts, {}, None)
        assert [r["output"] for r in recs] == ["obs", "obs"]

    async def test_execute_parallel_tools_enabled(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        import core.hallucination_config as hc
        monkeypatch.setattr(hc, "is_parallel_tools_enabled", lambda: True, raising=False)
        monkeypatch.setattr(hc, "get_max_parallel_tools", lambda: 4, raising=False)
        db = sl.return_value.__enter__.return_value
        gov = ama.AgentGovernanceService.return_value
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": True, "action_complexity": 1})

        agent.mcp.call_tool = AsyncMock(return_value={"success": True})
        agent.mcp.search_tools = AsyncMock(return_value=[
            {"name": "new_tool", "description": "d", "parameters": {}}])
        acts = [
            ama.ToolCall(tool="tool_a", params={}),
            ama.ToolCall(tool="mcp_tool_search", params={"query": "q"}),
        ]
        recs = await agent._execute_parallel_tools(acts, {}, None)
        by_tool = {r["tool_name"] for r in recs}
        assert by_tool == {"tool_a", "mcp_tool_search"}
        assert any("new tool" in r["output"] or "Found 1" in r["output"] for r in recs)

        # blocked batch
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": False, "action_complexity": 1})
        recs = await agent._execute_parallel_tools(
            [ama.ToolCall(tool="tool_a", params={})], {}, None)
        assert "Governance blocked" in recs[0]["output"]

        # requires approval + rejected
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": True, "action_complexity": 5, "reason": "complex"})
        gov.request_approval = MagicMock(return_value="act-x")
        agent._wait_for_all_approvals = AsyncMock(return_value=False)
        cb = AsyncMock()
        recs = await agent._execute_parallel_tools(
            [ama.ToolCall(tool="tool_a", params={})], {}, cb)
        assert "REJECTED" in recs[0]["output"]

        # tool exception inside gather (exception escapes the gov path itself)
        from core.sandbox_killrun import KillRunAborted
        gov.can_perform_action_async = AsyncMock(return_value={
            "allowed": True, "action_complexity": 1})
        inner = AsyncMock(side_effect=RuntimeError("boom"))
        agent._execute_tool_with_governance = inner
        recs = await agent._execute_parallel_tools(
            [ama.ToolCall(tool="tool_a", params={})], {}, None)
        assert "Tool error" in recs[0]["output"]
        assert recs[0]["verified_kind"] == "error"
        # KillRun inside the batch aborts the whole run
        agent._execute_tool_with_governance = AsyncMock(side_effect=KillRunAborted("kill"))
        with pytest.raises(KillRunAborted):
            await agent._execute_parallel_tools(
                [ama.ToolCall(tool="tool_a", params={})], {}, None)

        # search failure
        agent.mcp.call_tool = AsyncMock(return_value="ok")
        agent.mcp.search_tools = AsyncMock(side_effect=RuntimeError("search fail"))
        recs = await agent._execute_parallel_tools(
            [ama.ToolCall(tool="mcp_tool_search", params={"query": "q"})], {}, None)
        assert "Tool search failed" in recs[0]["output"]

    def test_persist_reasoning_step(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        db = sl.return_value.__enter__.return_value
        out = agent._persist_reasoning_step(
            execution_id="e", step_number=1, step_type="action", thought="t",
            action_dict={"tool": "x"}, observation="o", confidence=0.9,
            verified_kind="verified", verification_evidence="ev", duration_ms=1.0,
            request="r", final_answer=None, context={})
        assert out  # uuid returned
        db.add.assert_called_once()
        # failure path
        sl.return_value.__enter__ = MagicMock(side_effect=RuntimeError("db"))
        assert agent._persist_reasoning_step(
            "e", 1, "action", "t", None, None, 0.9, "u", None, 0.0, "r", None, {}) == ""

    async def test_record_execution(self, meta_agent):
        agent, sl, ama = meta_agent
        agent.world_model.record_experience = AsyncMock()
        gov = ama.AgentGovernanceService.return_value
        gov.record_outcome = AsyncMock()
        await agent._record_execution("req", {
            "status": "success", "final_output": "ans", "actions_executed": []},
            ama.AgentTriggerMode.MANUAL)
        agent.world_model.record_experience.assert_awaited_once()
        gov.record_outcome.assert_awaited_once()
        gov.record_outcome = AsyncMock(side_effect=RuntimeError("x"))
        await agent._record_execution("req", {"status": "failed", "final_output": None},
                                      ama.AgentTriggerMode.MANUAL)

    def test_get_communication_instruction(self, meta_agent):
        agent, sl, ama = meta_agent
        assert agent._get_communication_instruction({}) == ""
        db = sl.return_value
        user = NS(metadata_json={"communication_style": {
            "enable_personalization": True, "style_guide": "Be terse."}})
        db.query.return_value.filter.return_value.first.return_value = user
        out = agent._get_communication_instruction({"user_id": "u1"})
        assert "Be terse." in out
        user2 = NS(metadata_json=None)
        db.query.return_value.filter.return_value.first.return_value = user2
        assert agent._get_communication_instruction({"user_id": "u1"}) == ""
        sl.side_effect = RuntimeError("no db")
        assert agent._get_communication_instruction({"user_id": "u1"}) == ""
        sl.side_effect = None

    async def test_route_with_governance_all_paths(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        from core.intent_classifier import IntentCategory, IntentClassification

        def intent(cat):
            return IntentClassification(
                category=cat, confidence=0.9, requires_execution=True,
                reasoning="r", suggested_handler="llm_service")

        # CHAT bypass
        agent._route_to_chat = AsyncMock(return_value={"route": "CHAT", "x": 1})
        res = await agent.route_with_governance("hi", intent(IntentCategory.CHAT), "u1")
        assert res["governance_checked"] is False
        # WORKFLOW allowed
        agent._check_governance = AsyncMock(return_value=(True, None))
        agent._route_to_workflow = AsyncMock(return_value={"route": "WORKFLOW"})
        res = await agent.route_with_governance("run payroll",
                                                intent(IntentCategory.WORKFLOW), "u1")
        assert res["governance_allowed"] is True
        # TASK allowed
        agent._route_to_task = AsyncMock(return_value={"route": "TASK"})
        res = await agent.route_with_governance("do thing",
                                                intent(IntentCategory.TASK), "u1")
        assert res["governance_allowed"] is True
        # denied -> chat alternative proposal
        agent._check_governance = AsyncMock(return_value=(False, "maturity"))
        agent._propose_chat_alternative = AsyncMock(
            return_value={"route": "CHAT", "status": "auto_takeover_proposal"})
        res = await agent.route_with_governance("run payroll",
                                                intent(IntentCategory.WORKFLOW), "u1")
        assert res["governance_allowed"] is False

    async def test_internal_routing_methods(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        agent.llm.generate_response = AsyncMock(return_value="answer")
        res = await agent._route_to_chat("hi", "u1")
        assert res == {"route": "CHAT", "handler": "LLMService",
                       "response": "answer", "status": "chat_complete"}

        queen = MagicMock()
        queen.generate_blueprint = AsyncMock(return_value={
            "blueprint_id": "bp", "architecture_name": "arch", "nodes": [1, 2]})
        with patch.object(ama, "QueenAgent", MagicMock(return_value=queen)):
            res = await agent._route_to_workflow("make workflow", "u1")
        assert res["blueprint_id"] == "bp" and res["node_count"] == 2

        admiral = MagicMock()
        admiral.recruit_and_execute = AsyncMock(
            return_value={"chain_id": "c", "specialists_count": 2})
        import core.fleet_admiral as fa
        monkeypatch.setattr(fa, "FleetAdmiral", MagicMock(return_value=admiral), raising=False)
        res = await agent._route_to_task("do task", "u1")
        assert res["chain_id"] == "c" and res["status"] == "task_routed"

        agent.llm.generate_response = AsyncMock(return_value="proposal")
        res = await agent._propose_chat_alternative("req", "workflow", "why", "u1")
        assert res["auto_takeover"] is True and res["proposal"] == "proposal"

    async def test_check_governance_method(self, meta_agent):
        agent, sl, ama = meta_agent
        db = sl.return_value.__enter__.return_value
        gov = ama.AgentGovernanceService.return_value
        gov.canPerformAction = AsyncMock(return_value=NS(allowed=True))
        assert await agent._check_governance("u1", "atom", "workflow") == (True, None)
        gov.canPerformAction = AsyncMock(return_value=NS(allowed=False, reason="no"))
        assert await agent._check_governance("u1", "atom", "workflow") == (False, "no")

    async def test_handle_data_event_trigger(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        # queue enabled
        import core.task_queue as tq
        queue = MagicMock()
        queue.enabled = True
        queue.enqueue_job = MagicMock(return_value="job-1")
        import core.agent_worker_wrapper as aww
        monkeypatch.setattr(tq, "get_task_queue", lambda: queue, raising=False)
        res = await ama.handle_data_event_trigger("invoice.created", {"id": 1})
        assert res["status"] == "queued" and res["task_id"] == "job-1"
        # queue import fails -> inline execution
        queue.enabled = False
        with patch.object(ama, "AtomMetaAgent") as cls:
            inst = cls.return_value
            inst.execute = AsyncMock(return_value={"status": "success"})
            res = await ama.handle_data_event_trigger("evt", {})
        assert res["status"] == "success"
        # enqueue import error path
        import builtins
        real_import = builtins.__import__

        def fake_import(name, *a, **k):
            if name == "core.agent_worker_wrapper":
                raise ImportError("nope")
            return real_import(name, *a, **k)
        with patch.object(builtins, "__import__", side_effect=fake_import), \
             patch.object(ama, "AtomMetaAgent") as cls:
            inst = cls.return_value
            inst.execute = AsyncMock(return_value={"status": "inline"})
            res = await ama.handle_data_event_trigger("evt", {})
        assert res["status"] == "inline"

    async def test_handle_manual_trigger(self, meta_agent, monkeypatch):
        agent, sl, ama = meta_agent
        user = NS(id="u1", email="u@x.io")
        with patch.object(ama, "AtomMetaAgent") as cls:
            inst = cls.return_value
            inst.execute = AsyncMock(return_value={"status": "success"})

            async def run_callback(execute_mock):
                # grab the streaming callback passed to execute and exercise it
                cb = execute_mock.await_args.kwargs["step_callback"]
                import core.websockets as wsmod
                tracker = MagicMock()
                # ws_manager was bound when handle_manual_trigger imported it,
                # so patch broadcast on the bound (original) manager object.
                with patch.object(wsmod.manager, "broadcast", AsyncMock()) as bc, \
                     patch("core.reasoning_chain.get_reasoning_tracker",
                           lambda: tracker), \
                     patch("core.reasoning_chain.ReasoningStep", MagicMock()), \
                     patch("core.reasoning_chain.ReasoningStepType") as rst:
                    rst.ACTION = "action"
                    rst.CONCLUSION = "conclusion"
                    rst.INTENT_ANALYSIS = "planning"
                    rst.DECISION = "decision"
                    await cb({"execution_id": "e1", "step_type": "final_answer",
                              "thought": "t", "action": None, "output": None,
                              "confidence": 0.9, "duration_ms": 1.0, "step": 1})
                bc.assert_awaited_once()
                tracker.persist_step_to_db.assert_called_once()
                # no execution_id branch
                await cb({"execution_id": None})

            res = await ama.handle_manual_trigger("do it", user, "ws1",
                                                  additional_context={"k": "v"},
                                                  execution_id="e1")
            assert res["status"] == "success"
            await run_callback(inst.execute)

    async def test_module_level_router_functions(self, meta_agent):
        agent, sl, ama = meta_agent
        fake = NS(llm=MagicMock())
        fake.llm.generate_response = AsyncMock(return_value="ans")
        res = await ama._route_to_chat(fake, "req", "u1")
        assert res["response"] == "ans"

        queen = MagicMock()
        queen.generate_blueprint = AsyncMock(return_value={
            "blueprint_id": "b", "architecture_name": "a", "nodes": []})
        fake.queen = None
        import core.agents.queen_agent as qa
        with patch.object(qa, "QueenAgent", MagicMock(return_value=queen)):
            res = await ama._route_to_workflow(fake, "req", "u1")
        assert res["blueprint_id"] == "b"

        fake.queen = None
        admiral = MagicMock()
        admiral.recruit_and_execute = AsyncMock(return_value={"chain_id": "c"})
        import core.fleet_admiral as fa
        with patch.object(fa, "FleetAdmiral", MagicMock(return_value=admiral)):
            res = await ama._route_to_task(fake, "req", "u1")
        assert res["result"]["chain_id"] == "c"

        fake.llm.generate_response = AsyncMock(return_value="prop")
        res = await ama._propose_chat_alternative(fake, "r", "workflow", "why", "u1")
        assert res["proposal"] == "prop"

        # module-level _check_governance
        gov = ama.AgentGovernanceService.return_value
        gov.canPerformAction = AsyncMock(return_value=NS(allowed=False, reason="no"))
        assert await ama._check_governance(fake, "u1", "atom", "workflow") == (False, "no")
        gov.canPerformAction = AsyncMock(return_value=NS(allowed=True))
        assert await ama._check_governance(fake, "u1", "atom", "workflow") == (True, None)

    def test_get_atom_agent_singleton(self, meta_agent):
        agent, sl, ama = meta_agent
        ama._atom_instance = None
        a1 = ama.get_atom_agent("ws1")
        assert ama.get_atom_agent("ws1") is a1
        a2 = ama.get_atom_agent("ws2")
        assert a2 is not a1 and a2.workspace_id == "ws2"
        ama._atom_instance = None


class TestAtomExecutePaths:
    """execute() ReAct-loop branches not covered by w32."""

    @pytest.fixture
    def agent(self, meta_agent):
        agent, sl, ama = meta_agent
        from ai.nlp_engine import RouteCategory, RouteClassification

        workspace = NS(tenant_id="default")
        db = sl.return_value.__enter__.return_value
        db.query.return_value.filter.return_value.first.return_value = workspace

        nlu = MagicMock()
        nlu.classify_route = AsyncMock(return_value=RouteClassification(
            category=RouteCategory.ONE_OFF, reasoning="r", confidence=0.9))
        ama.NaturalLanguageEngine = MagicMock(return_value=nlu)

        agent.world_model.recall_experiences = AsyncMock(return_value={})
        agent.world_model.recall_episodes = AsyncMock(return_value=[{"canvas_id": "c1"}])
        agent.mcp.get_all_tools = AsyncMock(return_value=[
            {"name": "trigger_workflow", "description": "d", "parameters": {}}])
        import core.field_guide_service as fgs
        fgs.get_field_guide_service = lambda: MagicMock(
            get_field_guide_context=lambda w: "guide")
        agent._check_budget_before_react = AsyncMock(return_value={"allowed": True})
        agent._persist_reasoning_step = MagicMock(return_value="step-id")
        agent._record_execution = AsyncMock()
        agent._react_step = AsyncMock()
        return agent, sl, ama

    def _step(self, ama, **kw):
        return ama.ReActStep(thought="t", **kw)

    async def test_execute_canvas_context(self, agent):
        a, sl, ama = agent
        canvas_state = MagicMock()
        canvas_state.canvas_id = "c1"
        canvas_state.comments = [NS(content="hello")]
        canvas_state.artifact_count = 2
        a.canvas_provider.get_canvas_context = AsyncMock(return_value=canvas_state)
        a.canvas_provider.format_for_agent = MagicMock(return_value="CANVAS TEXT")
        a._react_step = AsyncMock(return_value=self._step(
            ama, final_answer="done", confidence=0.9))
        res = await a.execute("short req", canvas_context={"canvas_id": "c1"})
        assert res["status"] == "success"
        a.canvas_provider.get_canvas_context.assert_awaited_once()

    async def test_execute_canvas_failure_raises(self, agent):
        a, sl, ama = agent
        a.canvas_provider.get_canvas_context = AsyncMock(side_effect=RuntimeError("x"))
        with pytest.raises(RuntimeError):
            await a.execute("short req", canvas_context={"canvas_id": "c1"})

    async def test_execute_tool_search_then_final(self, agent):
        a, sl, ama = agent
        a.mcp.search_tools = AsyncMock(return_value=[
            {"name": "new_tool", "description": "d", "parameters": {}}])
        a._react_step = AsyncMock(side_effect=[
            self._step(ama, action=ama.ToolCall(tool="mcp_tool_search",
                                                params={"query": "q"})),
            self._step(ama, final_answer="done"),
        ])
        cb = AsyncMock()
        res = await a.execute("list things", step_callback=cb)
        assert res["status"] == "success"
        assert len(a.session_tools) == 1

    async def test_execute_delegation_step(self, agent):
        a, sl, ama = agent
        a._execute_delegation = AsyncMock(return_value="delegated output")
        a._react_step = AsyncMock(side_effect=[
            self._step(ama, action=ama.ToolCall(tool="delegate_task",
                                                params={"agent_name": "sales",
                                                        "task": "do"})),
            self._step(ama, final_answer="done"),
        ])
        res = await a.execute("delegate things")
        assert res["status"] == "success"

    async def test_execute_generic_tool_with_critiques(self, agent, monkeypatch):
        a, sl, ama = agent
        a._execute_tool_with_governance = AsyncMock(return_value="Tool error. Please try again.")
        a._react_step = AsyncMock(side_effect=[
            self._step(ama, action=ama.ToolCall(tool="call_integration", params={})),
            self._step(ama, final_answer="done"),
        ])
        res = await a.execute("call tool")
        assert res["status"] == "success"

    async def test_execute_parallel_actions_branch(self, agent, monkeypatch):
        a, sl, ama = agent
        import core.hallucination_config as hc
        monkeypatch.setattr(hc, "is_parallel_tools_enabled", lambda: True, raising=False)
        monkeypatch.setattr(hc, "get_max_parallel_tools", lambda: 4, raising=False)
        a._execute_parallel_tools = AsyncMock(return_value=[{
            "tool_name": "t1", "params": {}, "output": "obs",
            "verified_kind": "failed_verification", "verified_evidence": "no proof"}])
        a._react_step = AsyncMock(side_effect=[
            self._step(ama, actions=[ama.ToolCall(tool="t1", params={})]),
            self._step(ama, final_answer="done"),
        ])
        cb = AsyncMock()
        res = await a.execute("parallel run", step_callback=cb)
        assert res["status"] == "success"

    async def test_execute_parallel_actions_disabled_promotes_first(self, agent, monkeypatch):
        a, sl, ama = agent
        import core.hallucination_config as hc
        monkeypatch.setattr(hc, "is_parallel_tools_enabled", lambda: False, raising=False)
        a._execute_tool_with_governance = AsyncMock(return_value="obs")
        a._react_step = AsyncMock(side_effect=[
            self._step(ama, actions=[ama.ToolCall(tool="t1", params={})]),
            self._step(ama, final_answer="done"),
        ])
        res = await a.execute("serial run")
        assert res["status"] == "success"

    async def test_execute_no_action_converts_thought(self, agent):
        a, sl, ama = agent
        a._react_step = AsyncMock(return_value=ama.ReActStep(thought="just thinking"))
        res = await a.execute("idle")
        assert res["final_output"] == "just thinking"

    async def test_execute_budget_halt(self, agent):
        a, sl, ama = agent
        a._check_budget_before_react = AsyncMock(return_value={
            "allowed": False, "reason": "over budget", "enforcement_mode": "hard_stop"})
        res = await a.execute("spendy")
        assert res["status"] == "budget_exceeded"
        assert res["failure_reason"] == "over budget"
        assert res["failure_mode"] == "hard_stop"

    async def test_execute_max_steps_timeout(self, agent):
        a, sl, ama = agent
        # every step wants an action -> runs out of steps
        a._execute_tool_with_governance = AsyncMock(return_value="ok")
        a._react_step = AsyncMock(return_value=self._step(
            ama, action=ama.ToolCall(tool="call_integration", params={})))
        res = await a.execute("loop forever")
        assert res["status"] == "timeout"

    async def test_execute_radio_inbox_drain(self, agent, monkeypatch):
        a, sl, ama = agent
        import core.agent_radio.radio_service as rs
        monkeypatch.setattr(rs, "inbox_drain_text",
                            lambda aid, tid: "@atom mention", raising=False)
        a._react_step = AsyncMock(return_value=self._step(ama, final_answer="done"))
        res = await a.execute("radio", context={"radio_thread_id": "th1"})
        assert res["status"] == "success"
