"""
Bug-hunt tests for the Office subsystem (TDD red-green).

Covers:
  A. Path containment (R53 pattern) gaps: write_cell / modify_document /
     modify_slides and the workbook-runtime wrappers (get_evaluated_range,
     insert_rows, insert_cols, recalculate, add_pivot_table, run_macro)
     never validate file_path — arbitrary .xlsx/.docx/.pptx read+write
     through POST /excel, POST /word, POST /pptx, /excel/formula-result,
     /excel/insert-rows, /excel/insert-columns, /excel/recalculate,
     /excel/pivot-table, /excel/run-macro.
  B. CWE-1236 CSV-injection bypass: leading whitespace before a formula
     prefix (= + - @) is not sanitized by _sanitize_csv_cell.
  C. Formula injection into xlsx: write_cell with is_formula=False stores a
     string starting with '=' as a live formula (openpyxl data_type 'f').
  D. Dead WebSocket push: broadcast_file_update calls ws_manager.broadcast
     with a single dict argument (signature is (channel, message)) — the
     canvas:update frame is never delivered, co-editing live-update is dead.
"""

from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from core.auth import get_current_user as auth_get_current_user
from core.database import get_db_session


def _make_client(monkeypatch):
    from api.office_routes import router

    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[auth_get_current_user] = lambda: MagicMock(
        id="u-bh", email="u@example.com"
    )
    app.dependency_overrides[get_db_session] = lambda: MagicMock()
    return TestClient(app, raise_server_exceptions=False)


def _office_env(monkeypatch, tmp_path):
    monkeypatch.setenv("ATOM_OFFICE_DIR", str(tmp_path / "office"))
    (tmp_path / "office").mkdir(exist_ok=True)
    outside = tmp_path / "outside"
    outside.mkdir(exist_ok=True)
    return outside


def _xlsx(path, cells):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for coord, value in cells.items():
        ws[coord] = value
    wb.save(path)


def _docx(path, text):
    import docx

    doc = docx.Document()
    doc.add_paragraph(text)
    doc.save(path)


# ---------------------------------------------------------------------------
# A. Path containment on write paths
# ---------------------------------------------------------------------------


class TestWritePathsContained:
    def test_write_cell_outside_dir_rejected(self, monkeypatch, tmp_path):
        """POST /excel must not modify .xlsx files outside the office dir."""
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "victim.xlsx"
        _xlsx(victim, {"A1": "ORIGINAL"})
        original_bytes = victim.read_bytes()

        client = _make_client(monkeypatch)
        resp = client.post(
            "/excel",
            json={
                "file_path": str(victim),
                "cell_path": "/Sheet1/A1",
                "value": "ATTACKER",
                "is_formula": False,
            },
        )

        assert resp.status_code == 400, (
            f"POST /excel accepted an out-of-scope file_path ({resp.status_code})"
        )
        assert victim.read_bytes() == original_bytes, (
            "victim .xlsx was modified through POST /excel (arbitrary file write)"
        )

    def test_write_cell_outside_dir_rejected_service(self, monkeypatch, tmp_path):
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "v2.xlsx"
        _xlsx(victim, {"A1": "ORIGINAL"})
        before = victim.read_bytes()

        from core.office_service import ExcelManager

        res = ExcelManager().write_cell(str(victim), "/Sheet1/A1", "EVIL")

        assert res.get("success") is False
        assert victim.read_bytes() == before

    def test_write_cell_new_file_outside_dir_rejected(self, monkeypatch, tmp_path):
        """write_cell creates missing files — that creation must be contained too."""
        outside = _office_env(monkeypatch, tmp_path)
        new_file = outside / "brand_new.xlsx"

        from core.office_service import ExcelManager

        res = ExcelManager().write_cell(str(new_file), "/Sheet1/A1", "x")

        assert res.get("success") is False, (
            "write_cell created a file outside the office dir"
        )
        assert not new_file.exists()

    def test_modify_word_outside_dir_rejected(self, monkeypatch, tmp_path):
        """POST /word must not modify .docx files outside the office dir."""
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "victim.docx"
        _docx(victim, "ORIGINAL CONTENT")
        original_bytes = victim.read_bytes()

        client = _make_client(monkeypatch)
        resp = client.post(
            "/word",
            json={
                "file_path": str(victim),
                "action": "append",
                "content": "ATTACKER",
            },
        )

        assert resp.status_code == 400, (
            f"POST /word accepted an out-of-scope file_path ({resp.status_code})"
        )
        assert victim.read_bytes() == original_bytes, (
            "victim .docx was modified through POST /word (arbitrary file write)"
        )

    def test_modify_word_new_file_outside_dir_rejected(self, monkeypatch, tmp_path):
        outside = _office_env(monkeypatch, tmp_path)
        new_file = outside / "new.docx"

        from core.office_service import WordManager

        res = WordManager().modify_document(str(new_file), "append", "x")

        assert res.get("success") is False
        assert not new_file.exists(), (
            "modify_document created a .docx outside the office dir"
        )

    def test_modify_pptx_outside_dir_rejected(self, monkeypatch, tmp_path):
        pytest.importorskip("pptx")
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "victim.pptx"
        import pptx

        prs = pptx.Presentation()
        prs.save(victim)
        original_bytes = victim.read_bytes()

        client = _make_client(monkeypatch)
        resp = client.post(
            "/pptx",
            json={
                "file_path": str(victim),
                "action": "add_slide",
                "options": {"title": "EVIL", "layout_idx": 1},
            },
        )

        assert resp.status_code == 400, (
            f"POST /pptx accepted an out-of-scope file_path ({resp.status_code})"
        )
        assert victim.read_bytes() == original_bytes, (
            "victim .pptx was modified through POST /pptx"
        )

    def test_formula_result_outside_dir_rejected(self, monkeypatch, tmp_path):
        """GET /excel/formula-result must not read .xlsx outside the office dir."""
        outside = _office_env(monkeypatch, tmp_path)
        secret = outside / "secret.xlsx"
        _xlsx(secret, {"A1": "SECRET_VALUE"})

        client = _make_client(monkeypatch)
        resp = client.get(
            "/excel/formula-result",
            params={"file_path": str(secret), "cell_path": "/Sheet1/A1"},
        )

        assert resp.status_code == 400, (
            f"GET /excel/formula-result leaked an out-of-scope file ({resp.status_code})"
        )
        assert "SECRET_VALUE" not in resp.text

    def test_insert_rows_outside_dir_rejected(self, monkeypatch, tmp_path):
        """POST /excel/insert-rows must not modify .xlsx outside the office dir."""
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "ins.xlsx"
        _xlsx(victim, {"A1": 1, "B1": 2})
        original_bytes = victim.read_bytes()

        client = _make_client(monkeypatch)
        resp = client.post(
            "/excel/insert-rows",
            params={"file_path": str(victim), "sheet_name": "Sheet1", "row": 2},
        )

        assert resp.status_code == 400, (
            f"POST /excel/insert-rows accepted an out-of-scope path ({resp.status_code})"
        )
        assert victim.read_bytes() == original_bytes

    def test_insert_columns_outside_dir_rejected(self, monkeypatch, tmp_path):
        outside = _office_env(monkeypatch, tmp_path)
        victim = outside / "insc.xlsx"
        _xlsx(victim, {"A1": 1, "A2": 2})
        original_bytes = victim.read_bytes()

        client = _make_client(monkeypatch)
        resp = client.post(
            "/excel/insert-columns",
            params={"file_path": str(victim), "sheet_name": "Sheet1", "column": 2},
        )

        assert resp.status_code == 400, (
            f"POST /excel/insert-columns accepted out-of-scope path ({resp.status_code})"
        )
        assert victim.read_bytes() == original_bytes

    def test_recalculate_outside_dir_rejected(self, monkeypatch, tmp_path):
        outside = _office_env(monkeypatch, tmp_path)
        secret = outside / "r.xlsx"
        _xlsx(secret, {"A1": "=1+1"})

        client = _make_client(monkeypatch)
        resp = client.post("/excel/recalculate", params={"file_path": str(secret)})

        assert resp.status_code == 400, (
            f"POST /excel/recalculate accepted an out-of-scope path ({resp.status_code})"
        )

    def test_pivot_table_outside_dir_rejected(self, monkeypatch, tmp_path):
        """pivot-table reads the whole source sheet — must not read out-of-scope files."""
        outside = _office_env(monkeypatch, tmp_path)
        secret = outside / "pivot_secret.xlsx"
        _xlsx(secret, {"A1": "SECRET_HEADER", "A2": "SECRET_ROW"})

        client = _make_client(monkeypatch)
        resp = client.post(
            "/excel/pivot-table",
            json={
                "file_path": str(secret),
                "sheet_name": "Sheet1",
                "pivot_sheet_name": "Pivot",
                "data_range": "A1:A2",
                "rows": [],
                "columns": [],
                "values": [{"field": "SECRET_HEADER", "function": "sum"}],
            },
        )

        assert resp.status_code == 400, (
            f"POST /excel/pivot-table read an out-of-scope file ({resp.status_code})"
        )

    def test_run_macro_outside_dir_rejected(self, monkeypatch, tmp_path):
        outside = _office_env(monkeypatch, tmp_path)
        secret = outside / "macro_secret.xlsx"
        _xlsx(secret, {"A1": 1})

        client = _make_client(monkeypatch)
        resp = client.post(
            "/excel/run-macro",
            json={"file_path": str(secret), "macro_name": "Eval"},
        )

        assert resp.status_code == 400, (
            f"POST /excel/run-macro accepted an out-of-scope path ({resp.status_code})"
        )
        assert "outside the allowed office directory" in resp.text, (
            "run-macro must fail with the path-containment error, not an engine error"
        )

    def test_runtime_endpoints_not_dead_code(self, monkeypatch, tmp_path):
        """OfficeService has no ExcelManager attribute — all six runtime
        endpoints used to 500 with AttributeError on EVERY request (dead-code
        seam: routes/tools call office_service.ExcelManager.X)."""
        from core.office_service import OfficeService

        svc = OfficeService()
        assert not hasattr(svc, "ExcelManager")
        for name in (
            "get_evaluated_range", "insert_rows", "insert_columns",
            "recalculate", "add_pivot_table", "run_excel_macro",
        ):
            assert hasattr(svc.excel, name), (
                f"ExcelManager lacks runtime method {name} — endpoints call "
                "office_service.ExcelManager.<method> which raises AttributeError"
            )

    def test_runtime_endpoints_return_json_not_500(self, monkeypatch, tmp_path):
        """In-scope missing file: runtime endpoints must answer 400 with a
        service error, not crash with a 500 (unhandled FileNotFoundError)."""
        _office_env(monkeypatch, tmp_path)
        missing = tmp_path / "office" / "missing.xlsx"

        client = _make_client(monkeypatch)
        for method, url, params in (
            ("get", "/excel/formula-result",
             {"file_path": str(missing), "cell_path": "/Sheet1/A1"}),
            ("post", "/excel/recalculate", {"file_path": str(missing)}),
        ):
            resp = getattr(client, method)(url, params=params)
            assert resp.status_code == 400, (
                f"{url} returned {resp.status_code} for a missing in-scope file"
            )
            assert "detail" in resp.text


# ---------------------------------------------------------------------------
# B. CWE-1236 CSV-injection bypass: leading whitespace before a prefix
# ---------------------------------------------------------------------------


class TestCSVInjectionLeadingWhitespace:
    @pytest.mark.parametrize(
        "payload",
        [
            " =1+1",
            "  =HYPERLINK(\"http://evil.example\",\"x\")",
            "\t=1+1",
            "\r@SUM(1,1)",
            "\n+cmd|' /C calc'!A0",
            " \t =2+2",
        ],
    )
    def test_leading_whitespace_formula_sanitized(self, payload):
        from accounting.export_service import _sanitize_csv_cell

        result = _sanitize_csv_cell(payload)

        assert isinstance(result, str) and result.startswith("'"), (
            f"_sanitize_csv_cell did not neutralize leading-whitespace payload: "
            f"{payload!r} -> {result!r} (CWE-1236)"
        )
        assert not result.lstrip().startswith(("=", "+", "-", "@")), (
            f"sanitized cell still starts with a formula prefix after the quote: {result!r}"
        )

    def test_plain_text_untouched(self):
        from accounting.export_service import _sanitize_csv_cell

        assert _sanitize_csv_cell("plain text") == "plain text"
        assert _sanitize_csv_cell(123) == 123
        assert _sanitize_csv_cell(None) is None

    def test_existing_prefix_behavior_unchanged(self):
        from accounting.export_service import _sanitize_csv_cell

        assert _sanitize_csv_cell("=1+1") == "'=1+1"
        assert _sanitize_csv_cell("-2+3") == "'-2+3"


# ---------------------------------------------------------------------------
# C. Formula injection into xlsx: literal '=' text stored as live formula
# ---------------------------------------------------------------------------


class TestXlsxLiteralTextNotFormula:
    def test_equals_text_written_as_string_not_formula(self, monkeypatch, tmp_path):
        monkeypatch.setenv("ATOM_OFFICE_DIR", str(tmp_path / "office"))
        (tmp_path / "office").mkdir()
        f = tmp_path / "office" / "lit.xlsx"

        from core.office_service import ExcelManager

        res = ExcelManager().write_cell(
            str(f), "/Sheet1/A1", "=HYPERLINK(\"http://evil.example\",\"x\")",
            is_formula=False,
        )
        assert res.get("success") is True

        import openpyxl

        wb = openpyxl.load_workbook(f, data_only=False)
        cell = wb["Sheet1"]["A1"]
        assert cell.data_type != "f", (
            "literal '=' text was stored as a live formula in the xlsx "
            "(formula injection)"
        )
        assert cell.data_type == "s"
        assert cell.value == "=HYPERLINK(\"http://evil.example\",\"x\")"

    def test_true_formula_still_written_as_formula(self, monkeypatch, tmp_path):
        monkeypatch.setenv("ATOM_OFFICE_DIR", str(tmp_path / "office"))
        (tmp_path / "office").mkdir()
        f = tmp_path / "office" / "fml.xlsx"

        from core.office_service import ExcelManager

        ExcelManager().write_cell(str(f), "/Sheet1/A1", "=SUM(1,2)", is_formula=True)

        import openpyxl

        wb = openpyxl.load_workbook(f, data_only=False)
        assert wb["Sheet1"]["A1"].data_type == "f"
        assert wb["Sheet1"]["A1"].value == "=SUM(1,2)"


# ---------------------------------------------------------------------------
# D. Dead WebSocket push: wrong broadcast() signature
# ---------------------------------------------------------------------------


class TestOfficeSyncBroadcastChannel:
    @pytest.mark.asyncio
    async def test_broadcast_file_update_pushes_on_canvas_channel(self, monkeypatch, tmp_path):
        """broadcast() signature is (channel, message) — the canvas:update frame
        must be sent on 'canvas:{id}', never swallowed by a TypeError."""
        monkeypatch.setenv("ATOM_OFFICE_DIR", str(tmp_path / "office"))
        f = tmp_path / "office" / "doc.docx"
        f.parent.mkdir()
        f.write_bytes(b"dummy")

        db = MagicMock()
        from core.office_sync_service import OfficeSyncService

        broadcast = AsyncMock()
        with patch("core.office_sync_service.ws_manager.broadcast", new=broadcast), patch(
            "core.office_service.DocumentRenderer.render_to_html",
            return_value={"success": True, "html": "<p>rendered</p>"},
        ), patch.object(OfficeSyncService, "_ingest_document_to_memory_sync"):
            OfficeSyncService(db).broadcast_file_update("canvas-abc", str(f), "u1")

        calls = broadcast.call_args_list
        assert calls, "ws_manager.broadcast was never called"
        args = calls[0].args
        assert len(args) == 2, (
            f"broadcast called with wrong arity {len(args)} — signature is "
            f"(channel, message); the canvas:update frame is silently dropped "
            f"(TypeError) in production"
        )
        assert args[0] == "canvas:canvas-abc", (
            f"canvas update broadcast on wrong channel {args[0]!r}"
        )
        assert args[1]["type"] == "canvas:update"
