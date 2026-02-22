#!/usr/bin/env python3
"""
Test Suite for Excel Export Functionality
Tests Discord and Google Chat analytics Excel export features
"""

import sys
import importlib

# CRITICAL FIX: Ensure numpy/pandas/lancedb/pyarrow/openpyxl are real modules before importing analytics engines.
#
# During test collection, pytest imports test modules in alphabetical order. Some test files mock
# numpy with incomplete implementations (e.g., main_api_app_safe.py creates a mock numpy without
# the .short attribute). When this file is later collected, importing discord_analytics_engine
# (which imports openpyxl) fails because openpyxl tries to access numpy.short and other dtypes.
#
# Root cause: test_cognitive_tier_e2e.py imports main_api_app_safe.py which mocks numpy as:
#   np_mock = mock_package("numpy")  # Creates empty module without .short
#   sys.modules["numpy.core"] = MagicMock()
#   sys.modules["numpy._core"] = MagicMock()
#   sys.modules["numpy.core.multiarray"] = MagicMock()
#   sys.modules["numpy._core.multiarray"] = MagicMock()
#
# When discord_analytics_engine later imports openpyxl, openpyxl/compat/numbers.py tries to access
# numpy.short, which doesn't exist in the mock. Even worse, when we remove the mocked numpy and
# try to import the real one, the mocked submodules cause import errors.
#
# The fix is to detect and remove ALL mocked/incomplete numpy modules (including submodules),
# forcing a clean import of the real numpy module. We also remove any openpyxl modules that may
# have cached references to the corrupted numpy.
#
# This cleanup must happen at module level, before any imports that might trigger the issue.
def _is_mocked_module(module, mod_name):
    """Check if a module is mocked or incomplete."""
    if module is None:
        return True
    # Check for MagicMock (has _spec_class attribute that real modules don't have)
    if hasattr(module, '_spec_class'):
        return True
    # For numpy, check if it has critical attributes
    if mod_name == "numpy" and not hasattr(module, 'short'):
        return True
    return False

_modules_to_clean = ["numpy", "pandas", "lancedb", "pyarrow"]
for mod in _modules_to_clean:
    if mod in sys.modules and _is_mocked_module(sys.modules[mod], mod):
        # Remove the corrupted module and ALL its submodules
        keys_to_remove = [k for k in sys.modules.keys() if k == mod or k.startswith(mod + '.')]
        for key in keys_to_remove:
            sys.modules.pop(key, None)
        # Also remove any openpyxl modules that might have cached references to it
        for openpyxl_mod in list(sys.modules.keys()):
            if openpyxl_mod.startswith('openpyxl'):
                sys.modules.pop(openpyxl_mod, None)

import asyncio
import pytest
import statistics
from datetime import datetime, timedelta, timezone
from typing import List
from unittest.mock import Mock, MagicMock, patch
import io

# Import engines
from integrations.discord_analytics_engine import (
    DiscordAnalyticsEngine,
    DiscordAnalyticsMetric,
    DiscordAnalyticsTimeRange,
    DiscordAnalyticsGranularity,
    DiscordAnalyticsDataPoint,
    OPENPYXL_AVAILABLE
)
from integrations.google_chat_analytics_engine import (
    GoogleChatAnalyticsEngine,
    GoogleChatAnalyticsMetric,
    GoogleChatAnalyticsTimeRange,
    GoogleChatAnalyticsGranularity,
    GoogleChatAnalyticsDataPoint
)


# ============================================================================
# Test Fixtures
# ============================================================================

@pytest.fixture
def discord_engine():
    """Create Discord analytics engine instance"""
    config = {
        'database': Mock(),
        'redis': {'client': None},
        'cache_ttl': 300
    }
    return DiscordAnalyticsEngine(config)


@pytest.fixture
def google_chat_engine():
    """Create Google Chat analytics engine instance"""
    config = {
        'database': Mock(),
        'redis': {'client': None},
        'cache_ttl': 300
    }
    return GoogleChatAnalyticsEngine(config)


@pytest.fixture
def sample_discord_data():
    """Create sample Discord analytics data points"""
    now = datetime.now(timezone.utc)
    return [
        DiscordAnalyticsDataPoint(
            timestamp=now - timedelta(hours=3),
            metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
            value=150,
            dimensions={'channel_id': '123', 'guild_id': '456'},
            metadata={'user_id': 'user1'}
        ),
        DiscordAnalyticsDataPoint(
            timestamp=now - timedelta(hours=2),
            metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
            value=200,
            dimensions={'channel_id': '123', 'guild_id': '456'},
            metadata={'user_id': 'user2'}
        ),
        DiscordAnalyticsDataPoint(
            timestamp=now - timedelta(hours=1),
            metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
            value=175,
            dimensions={'channel_id': '123', 'guild_id': '456'},
            metadata={'user_id': 'user1'}
        ),
    ]


@pytest.fixture
def sample_google_chat_data():
    """Create sample Google Chat analytics data points"""
    now = datetime.now(timezone.utc)
    return [
        GoogleChatAnalyticsDataPoint(
            timestamp=now - timedelta(hours=3),
            metric=GoogleChatAnalyticsMetric.MESSAGE_COUNT,
            value=120,
            dimensions={'space_id': 'space789', 'thread_id': 'thread101'},
            metadata={'user_id': 'user3'}
        ),
        GoogleChatAnalyticsDataPoint(
            timestamp=now - timedelta(hours=2),
            metric=GoogleChatAnalyticsMetric.MESSAGE_COUNT,
            value=180,
            dimensions={'space_id': 'space789', 'thread_id': 'thread102'},
            metadata={'user_id': 'user4'}
        ),
    ]


# ============================================================================
# Discord Analytics Excel Export Tests
# ============================================================================

class TestDiscordExcelExport:
    """Test Discord analytics Excel export functionality"""

    @pytest.mark.asyncio
    async def test_export_to_excel_success(self, discord_engine, sample_discord_data):
        """Test successful Excel export"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        result = await discord_engine.export_analytics_data(
            metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
            time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
            granularity=DiscordAnalyticsGranularity.HOUR,
            format='excel',
            filters=None
        )

        # Mock the analytics retrieval
        with patch.object(discord_engine, 'get_analytics', return_value=sample_discord_data):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

        assert result['ok'] is True
        assert result['format'] == 'excel'
        assert 'data' in result
        assert result['filename'] == 'discord_message_count_last_24_hours.xlsx'
        assert isinstance(result['data'], bytes)
        assert len(result['data']) > 0

    @pytest.mark.asyncio
    async def test_export_to_excel_without_openpyxl(self, discord_engine, sample_discord_data):
        """Test Excel export without openpyxl installed"""
        # First provide sample data, then check openpyxl availability
        with patch.object(discord_engine, 'get_analytics', return_value=sample_discord_data), \
             patch('integrations.discord_analytics_engine.OPENPYXL_AVAILABLE', False):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

        assert result['ok'] is False
        assert 'openpyxl' in result['error'].lower()

    @pytest.mark.asyncio
    async def test_export_to_excel_empty_data(self, discord_engine):
        """Test Excel export with empty data"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        with patch.object(discord_engine, 'get_analytics', return_value=[]):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

        assert result['ok'] is False
        assert 'No data available' in result['error']

    def test_convert_to_excel_structure(self, discord_engine, sample_discord_data):
        """Test Excel file structure and content"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        # Convert to Excel
        excel_data = discord_engine._convert_to_excel(
            sample_discord_data,
            DiscordAnalyticsMetric.MESSAGE_COUNT,
            DiscordAnalyticsTimeRange.LAST_24_HOURS
        )

        assert excel_data is not None
        assert isinstance(excel_data, bytes)

        # Load and verify structure
        wb = load_workbook(io.BytesIO(excel_data))

        # Check sheets exist
        assert 'Summary' in wb.sheetnames
        assert 'Detailed Data' in wb.sheetnames
        assert 'Metadata' in wb.sheetnames

        # Verify Summary sheet has data
        summary_sheet = wb['Summary']
        assert summary_sheet['A1'].value == 'Metric'
        assert summary_sheet['B1'].value == 'Value'

        # Verify Data sheet has headers
        data_sheet = wb['Detailed Data']
        assert data_sheet['A1'].value == 'Timestamp'
        assert data_sheet['B1'].value == 'Metric'
        assert data_sheet['C1'].value == 'Value'
        assert data_sheet['D1'].value == 'Dimensions'
        assert data_sheet['E1'].value == 'Metadata'

        # Verify data rows
        assert data_sheet['A2'].value is not None
        assert data_sheet['B2'].value == 'message_count'
        assert data_sheet['C2'].value == '150'

    def test_convert_to_excel_statistics(self, discord_engine, sample_discord_data):
        """Test Excel export includes correct statistics"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        excel_data = discord_engine._convert_to_excel(
            sample_discord_data,
            DiscordAnalyticsMetric.MESSAGE_COUNT,
            DiscordAnalyticsTimeRange.LAST_24_HOURS
        )

        wb = load_workbook(io.BytesIO(excel_data))
        summary_sheet = wb['Summary']

        # Check statistics are calculated
        values = [150, 200, 175]
        expected_avg = round(statistics.mean(values), 2)
        expected_min = round(min(values), 2)
        expected_max = round(max(values), 2)

        # Find the statistic rows
        stats_found = {}
        for row in range(2, 20):
            metric_name = summary_sheet[f'A{row}'].value
            if metric_name == 'Average Value':
                stats_found['avg'] = summary_sheet[f'B{row}'].value
            elif metric_name == 'Min Value':
                stats_found['min'] = summary_sheet[f'B{row}'].value
            elif metric_name == 'Max Value':
                stats_found['max'] = summary_sheet[f'B{row}'].value

        assert stats_found['avg'] == expected_avg
        assert stats_found['min'] == expected_min
        assert stats_found['max'] == expected_max


# ============================================================================
# Google Chat Analytics Excel Export Tests
# ============================================================================

class TestGoogleChatExcelExport:
    """Test Google Chat analytics Excel export functionality"""

    @pytest.mark.asyncio
    async def test_export_to_excel_success(self, google_chat_engine, sample_google_chat_data):
        """Test successful Excel export"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        with patch.object(google_chat_engine, 'get_analytics', return_value=sample_google_chat_data):
            result = await google_chat_engine.export_analytics_data(
                metric=GoogleChatAnalyticsMetric.MESSAGE_COUNT,
                time_range=GoogleChatAnalyticsTimeRange.LAST_7_DAYS,
                granularity=GoogleChatAnalyticsGranularity.DAY,
                format='excel',
                filters=None
            )

        assert result['ok'] is True
        assert result['format'] == 'excel'
        assert 'data' in result
        assert result['filename'] == 'google_chat_message_count_last_7_days.xlsx'

    def test_convert_to_excel_structure(self, google_chat_engine, sample_google_chat_data):
        """Test Excel file structure and content"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        excel_data = google_chat_engine._convert_to_excel(
            sample_google_chat_data,
            GoogleChatAnalyticsMetric.MESSAGE_COUNT,
            GoogleChatAnalyticsTimeRange.LAST_7_DAYS
        )

        assert excel_data is not None

        wb = load_workbook(io.BytesIO(excel_data))

        # Check sheets exist
        assert 'Summary' in wb.sheetnames
        assert 'Detailed Data' in wb.sheetnames
        assert 'Metadata' in wb.sheetnames

        # Verify Metadata sheet content
        metadata_sheet = wb['Metadata']
        assert metadata_sheet['A1'].value == 'Property'
        assert metadata_sheet['B1'].value == 'Description'

        # Find engine name
        engine_found = False
        for row in range(2, 20):
            if metadata_sheet[f'B{row}'].value == 'Google Chat Analytics Engine':
                engine_found = True
                break

        assert engine_found, "Google Chat Analytics Engine not found in metadata"


# ============================================================================
# Edge Cases and Error Handling Tests
# ============================================================================

class TestExcelExportEdgeCases:
    """Test edge cases and error handling"""

    @pytest.mark.asyncio
    async def test_export_unsupported_format(self, discord_engine, sample_discord_data):
        """Test export with unsupported format"""
        # Need to provide data first, otherwise "No data available" error occurs first
        with patch.object(discord_engine, 'get_analytics', return_value=sample_discord_data):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='pdf',  # Unsupported format
                filters=None
            )

        assert result['ok'] is False
        assert 'Unsupported format' in result['error']

    @pytest.mark.asyncio
    async def test_export_analytics_exception_handling(self, discord_engine):
        """Test exception handling during export"""
        # Mock get_analytics to raise exception
        with patch.object(discord_engine, 'get_analytics', side_effect=Exception('Database error')):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

        assert result['ok'] is False
        assert 'error' in result

    def test_convert_to_excel_none_data(self, discord_engine):
        """Test Excel conversion with None data"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        result = discord_engine._convert_to_excel(
            None,
            DiscordAnalyticsMetric.MESSAGE_COUNT,
            DiscordAnalyticsTimeRange.LAST_24_HOURS
        )

        assert result is None

    def test_convert_to_excel_empty_list(self, discord_engine):
        """Test Excel conversion with empty list"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        result = discord_engine._convert_to_excel(
            [],
            DiscordAnalyticsMetric.MESSAGE_COUNT,
            DiscordAnalyticsTimeRange.LAST_24_HOURS
        )

        assert result is None


# ============================================================================
# Integration Tests
# ============================================================================

class TestExcelExportIntegration:
    """Integration tests for Excel export"""

    @pytest.mark.asyncio
    async def test_full_export_workflow_discord(self, discord_engine, sample_discord_data):
        """Test complete Discord export workflow"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        with patch.object(discord_engine, 'get_analytics', return_value=sample_discord_data):
            # Export to Excel
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

            assert result['ok'] is True

            # Verify we can write to file
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
                f.write(result['data'])
                f.flush()

                # Verify file exists and has content
                import os
                assert os.path.exists(f.name)
                assert os.path.getsize(f.name) > 0

                # Clean up
                os.unlink(f.name)

    @pytest.mark.asyncio
    async def test_export_multiple_metrics(self, discord_engine):
        """Test exporting multiple different metrics"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        metrics = [
            DiscordAnalyticsMetric.MESSAGE_COUNT,
            DiscordAnalyticsMetric.ACTIVE_USERS,
            DiscordAnalyticsMetric.REACTION_COUNT,
        ]

        for metric in metrics:
            with patch.object(discord_engine, 'get_analytics', return_value=[]):
                result = await discord_engine.export_analytics_data(
                    metric=metric,
                    time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                    granularity=DiscordAnalyticsGranularity.HOUR,
                    format='excel',
                    filters=None
                )

            # All should fail due to empty data, but format should be recognized
            assert 'error' in result or result['ok'] is False


# ============================================================================
# Performance Tests
# ============================================================================

class TestExcelExportPerformance:
    """Test Excel export performance"""

    @pytest.mark.asyncio
    async def test_export_large_dataset(self, discord_engine):
        """Test Excel export with large dataset"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        import time
        from datetime import datetime, timezone, timedelta

        # Create large dataset (1000 data points)
        now = datetime.now(timezone.utc)
        large_dataset = [
            DiscordAnalyticsDataPoint(
                timestamp=now - timedelta(hours=i),
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                value=100 + i % 50,
                dimensions={'channel_id': f'channel_{i % 10}', 'guild_id': '456'},
                metadata={'user_id': f'user_{i % 20}'}
            )
            for i in range(1000)
        ]

        start_time = time.time()

        with patch.object(discord_engine, 'get_analytics', return_value=large_dataset):
            result = await discord_engine.export_analytics_data(
                metric=DiscordAnalyticsMetric.MESSAGE_COUNT,
                time_range=DiscordAnalyticsTimeRange.LAST_24_HOURS,
                granularity=DiscordAnalyticsGranularity.HOUR,
                format='excel',
                filters=None
            )

        end_time = time.time()
        export_time = end_time - start_time

        assert result['ok'] is True
        assert export_time < 10.0  # Should complete in under 10 seconds


# ============================================================================
# Run Tests
# ============================================================================

if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
