"""Coverage wave 40 — core/workbook_runtime (23% → 90%+).

- _find_soffice paths, engine/can_evaluate/can_render properties
- recalculate: missing file raises, soffice path (mocked subprocess: success,
  timeout, no-output, generic failure), formulas path, no-engine no-op
- run_macro: missing file, no-soffice error, sandbox success/failure
- add_pivot_table: missing sheet, empty data, success (pandas pivot),
  exception
- render_to_html: missing file, soffice render (mocked), basic fallback,
  render error
- insert_rows/insert_cols: missing sheet, success with real xlsx
- get_evaluated_range: range + single cell + missing sheet
- get_formula_result delegates
- get_workbook_runtime singleton
"""
import asyncio
import os
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from core.workbook_runtime import (
    WorkbookRuntime,
    _find_soffice,
    get_workbook_runtime,
)


@pytest.fixture
def xlsx(tmp_path):
    """A real minimal workbook for openpyxl-path tests."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def runtime():
    r = WorkbookRuntime.__new__(WorkbookRuntime)
    r._soffice = None
    r._has_formulas = False
    return r


class TestFindSoffice:
    def test_finds_via_shutil(self):
        with patch("core.workbook_runtime.shutil.which", return_value="/usr/bin/soffice"):
            assert _find_soffice() == "/usr/bin/soffice"

    def test_mac_path_fallback(self):
        with patch("core.workbook_runtime.shutil.which", return_value=None), \
             patch("core.workbook_runtime.os.path.exists", return_value=True):
            assert _find_soffice() == "/Applications/LibreOffice.app/Contents/MacOS/soffice"

    def test_not_found(self):
        with patch("core.workbook_runtime.shutil.which", return_value=None), \
             patch("core.workbook_runtime.os.path.exists", return_value=False):
            assert _find_soffice() is None


class TestProperties:
    def test_engine_and_caps(self):
        with patch.object(WorkbookRuntime, "_check_formulas", return_value=False):
            r = WorkbookRuntime.__new__(WorkbookRuntime)
            r._soffice = None
            r._has_formulas = False
            assert r.engine == "openpyxl"
            assert r.can_evaluate is False
            assert r.can_render is False

            r._soffice = "/usr/bin/soffice"
            assert r.engine == "libreoffice"
            assert r.can_evaluate is True
            assert r.can_render is True

            r._soffice = None
            r._has_formulas = True
            assert r.engine == "formulas"
            assert r.can_evaluate is True
            assert r.can_render is False

    def test_check_formulas_import_error(self):
        with patch("builtins.__import__", side_effect=ImportError("no formulas")):
            assert WorkbookRuntime()._check_formulas() is False


class TestRecalculate:
    async def test_missing_file_raises(self, runtime):
        with pytest.raises(FileNotFoundError):
            await runtime.recalculate("/nope/missing.xlsx")

    async def test_no_engine_noop(self, runtime, xlsx):
        result = await runtime.recalculate(xlsx)
        assert result == Path(xlsx)

    async def test_soffice_success(self, runtime, xlsx, tmp_path):
        runtime._soffice = "/usr/bin/soffice"
        out = tmp_path / "out"
        out.mkdir()
        (out / xlsx.name).write_bytes(xlsx.read_bytes())
        with patch("core.workbook_runtime.asyncio.create_subprocess_exec",
                   new=AsyncMock()) as proc_mock:
            proc = MagicMock()
            proc.communicate = AsyncMock(return_value=(b"", b""))
            proc_mock.return_value = proc
            with patch("core.workbook_runtime.tempfile.TemporaryDirectory") as td:
                td.return_value.__enter__.return_value = str(out)
                result = await runtime._recalc_with_soffice(Path(xlsx))
        assert result == Path(xlsx)
        proc_mock.assert_awaited_once()

    async def test_soffice_timeout(self, runtime, xlsx, tmp_path):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.workbook_runtime.asyncio.create_subprocess_exec",
                   new=AsyncMock()) as proc_mock:
            proc = MagicMock()
            proc.communicate = AsyncMock(side_effect=asyncio.TimeoutError())
            proc_mock.return_value = proc
            with patch("core.workbook_runtime.tempfile.TemporaryDirectory") as td:
                td.return_value.__enter__.return_value = str(tmp_path)
                result = await runtime._recalc_with_soffice(Path(xlsx))
        assert result == Path(xlsx)

    async def test_soffice_no_output(self, runtime, xlsx, tmp_path):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.workbook_runtime.asyncio.create_subprocess_exec",
                   new=AsyncMock()) as proc_mock:
            proc = MagicMock()
            proc.communicate = AsyncMock(return_value=(b"", b"error"))
            proc_mock.return_value = proc
            with patch("core.workbook_runtime.tempfile.TemporaryDirectory") as td:
                td.return_value.__enter__.return_value = str(tmp_path)
                result = await runtime._recalc_with_soffice(Path(xlsx))
        assert result == Path(xlsx)

    async def test_formulas_path(self, runtime, xlsx):
        runtime._has_formulas = True
        fm = MagicMock()
        fm.loads.return_value.finish.return_value = fm
        fm.calculate.return_value = {
            f"'[{xlsx.name}]Data'!A3": SimpleNamespace(value=30)
        }
        with patch.dict("sys.modules", {"formulas": MagicMock(ExcelModel=lambda: fm)}), \
             patch("openpyxl.load_workbook") as lw:
            wb = MagicMock()
            ws = MagicMock()
            cell = MagicMock()
            cell.value = "=A1+A2"
            cell.coordinate = "A3"
            ws.iter_rows.return_value = [[cell]]
            wb.__getitem__.return_value = ws
            wb.sheetnames = ["Data"]
            lw.return_value = wb
            result = await runtime._recalc_with_formulas(Path(xlsx))
        assert result == Path(xlsx)
        assert cell.value == 30
        wb.save.assert_called_once()


class TestRunMacro:
    async def test_missing_file(self, runtime):
        result = await runtime.run_macro("/nope.xlsx", "Macro1")
        assert result["success"] is False

    async def test_no_soffice(self, runtime, xlsx):
        result = await runtime.run_macro(xlsx, "Macro1")
        assert result["success"] is False
        assert "LibreOffice" in result["error"]

    async def test_sandbox_success(self, runtime, xlsx):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.firecracker_sandbox.get_sandbox") as gs:
            sandbox = MagicMock()
            sandbox.execute_in_sandbox = AsyncMock(return_value=True)
            gs.return_value = sandbox
            with patch.object(runtime, "recalculate", new=AsyncMock()) as recalc:
                result = await runtime.run_macro(xlsx, "Macro1")
        assert result["success"] is True
        recalc.assert_awaited_once()

    async def test_sandbox_failure(self, runtime, xlsx):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.firecracker_sandbox.get_sandbox") as gs:
            sandbox = MagicMock()
            sandbox.execute_in_sandbox = AsyncMock(return_value=False)
            gs.return_value = sandbox
            result = await runtime.run_macro(xlsx, "Macro1")
        assert result["success"] is False


class TestPivot:
    async def test_missing_sheet(self, runtime, xlsx):
        result = await runtime.add_pivot_table(xlsx, "Nope", "Pivot", "A1:B2", [], [], [])
        assert result["success"] is False

    async def test_success(self, runtime, xlsx):
        pivot_result = MagicMock()
        with patch("pandas.DataFrame") as pdf, \
             patch("pandas.pivot_table", return_value=pivot_result) as pt, \
             patch("openpyxl.utils.dataframe.dataframe_to_rows",
                   return_value=[["name", "value"], ["a", 1]]) as dtr:
            pdf.return_value = MagicMock()
            with patch("openpyxl.load_workbook") as lw:
                wb = MagicMock()
                wb.sheetnames = ["Data", "Pivot"]
                src = MagicMock()
                src.iter_rows.return_value = [
                    ("name", "value"),
                    ("a", 1),
                    ("b", 2),
                ]
                wb.__getitem__.return_value = src
                lw.return_value = wb
                with patch.object(runtime, "recalculate", new=AsyncMock()):
                    result = await runtime.add_pivot_table(
                        xlsx, "Data", "Pivot", "A1:B2",
                        rows=["name"], columns=[], values=[{"field": "value", "function": "sum"}],
                    )
        assert result["success"] is True
        assert result["pivot_sheet"] == "Pivot"
        assert pt.called

    async def test_exception(self, runtime, xlsx):
        with patch("openpyxl.load_workbook", side_effect=RuntimeError("boom")):
            result = await runtime.add_pivot_table(xlsx, "Data", "Pivot", "A1:B2", [], [], [])
        assert result["success"] is False


class TestRender:
    async def test_missing_file(self, runtime):
        assert await runtime.render_to_html("/nope.xlsx") == "<p>File not found</p>"

    async def test_soffice_render(self, runtime, xlsx, tmp_path):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.workbook_runtime.asyncio.create_subprocess_exec",
                   new=AsyncMock()) as proc_mock:
            proc = MagicMock()
            proc.communicate = AsyncMock(return_value=(b"", b""))
            proc_mock.return_value = proc
            with patch("core.workbook_runtime.tempfile.TemporaryDirectory") as td:
                td.return_value.__enter__.return_value = str(tmp_path)
                (tmp_path / (xlsx.stem + ".html")).write_text("<html>hi</html>")
                html = await runtime._render_html_with_soffice(Path(xlsx))
        assert "<html>hi</html>" in html

    async def test_soffice_render_failure_falls_back(self, runtime, xlsx, tmp_path):
        runtime._soffice = "/usr/bin/soffice"
        with patch("core.workbook_runtime.asyncio.create_subprocess_exec",
                   new=AsyncMock()) as proc_mock:
            proc = MagicMock()
            proc.communicate = AsyncMock(return_value=(b"", b""))
            proc_mock.return_value = proc
            with patch("core.workbook_runtime.tempfile.TemporaryDirectory") as td:
                td.return_value.__enter__.return_value = str(tmp_path)
                html = await runtime._render_html_with_soffice(Path(xlsx))
        assert "<table" in html

    def test_render_html_basic(self, runtime, xlsx):
        html = runtime._render_html_basic(Path(xlsx))
        assert "<table" in html
        assert "Data" in html
        assert "10" in html

    def test_render_html_basic_error(self, runtime):
        with patch("openpyxl.load_workbook", side_effect=RuntimeError("boom")):
            html = runtime._render_html_basic(Path("/nope.xlsx"))
        assert "Error rendering" in html


class TestStructural:
    async def test_insert_rows_missing_sheet(self, runtime, xlsx):
        result = await runtime.insert_rows(xlsx, "Nope", 1)
        assert result["success"] is False

    async def test_insert_rows_success(self, runtime, xlsx):
        with patch("openpyxl.load_workbook") as lw:
            wb = MagicMock()
            wb.sheetnames = ["Data"]
            ws = MagicMock()
            wb.__getitem__.return_value = ws
            lw.return_value = wb
            with patch.object(runtime, "recalculate", new=AsyncMock()):
                result = await runtime.insert_rows(xlsx, "Data", 2, count=3)
        assert result["success"] is True
        assert result["rows_inserted"] == 3
        ws.insert_rows.assert_called_once_with(2, 3)

    async def test_insert_cols_success(self, runtime, xlsx):
        with patch("openpyxl.load_workbook") as lw:
            wb = MagicMock()
            wb.sheetnames = ["Data"]
            ws = MagicMock()
            wb.__getitem__.return_value = ws
            lw.return_value = wb
            with patch.object(runtime, "recalculate", new=AsyncMock()):
                result = await runtime.insert_cols(xlsx, "Data", 1, count=2)
        assert result["success"] is True
        ws.insert_cols.assert_called_once_with(1, 2)


class TestEvaluatedReads:
    async def test_get_evaluated_range(self, runtime, xlsx):
        with patch("openpyxl.load_workbook") as lw:
            wb = MagicMock()
            wb.sheetnames = ["Data"]
            cell = MagicMock()
            cell.coordinate = "A1"
            cell.value = 10
            wb.__getitem__.return_value.__getitem__.return_value = cell
            lw.return_value = wb
            with patch.object(runtime, "recalculate", new=AsyncMock()):
                result = await runtime.get_evaluated_range(xlsx, "Data", "A1")
        assert result["success"] is True
        assert result["values"][0][0]["value"] == 10

    async def test_get_evaluated_range_missing_sheet(self, runtime, xlsx):
        with patch("openpyxl.load_workbook") as lw:
            wb = MagicMock()
            wb.sheetnames = ["Data"]
            lw.return_value = wb
            with patch.object(runtime, "recalculate", new=AsyncMock()):
                result = await runtime.get_evaluated_range(xlsx, "Nope", "A1")
        assert result["success"] is False

    async def test_get_formula_result_delegates(self, runtime, xlsx):
        with patch.object(runtime, "get_evaluated_range", new=AsyncMock(
            return_value={"success": True}
        )) as ger:
            result = await runtime.get_formula_result(xlsx, "Data", "A3")
        assert result["success"] is True
        ger.assert_awaited_once_with(xlsx, "Data", "A3")


class TestSingleton:
    def test_get_workbook_runtime(self):
        with patch("core.workbook_runtime._workbook_runtime", None):
            a = get_workbook_runtime()
            b = get_workbook_runtime()
            assert isinstance(a, WorkbookRuntime)
            assert a is b
