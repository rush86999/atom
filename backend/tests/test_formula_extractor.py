"""
Test suite for Formula Extractor

Tests cover:
- Formula extraction (Excel, CSV, ODS, XLS)
- Business logic (rule extraction, dependency mapping)
- Calculation execution (evaluation, caching)
- Integration (spreadsheet, database, API)
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

from core.formula_extractor import (
    FormulaExtractor,
    get_formula_extractor,
)


class TestFormulaExtractorInit:
    """Test FormulaExtractor initialization."""

    def test_initialization_with_default_workspace(self):
        """FormulaExtractor initializes with default workspace."""
        # Act
        extractor = FormulaExtractor()

        # Assert
        assert extractor.workspace_id == "default"

    def test_initialization_with_custom_workspace(self):
        """FormulaExtractor initializes with custom workspace."""
        # Act
        extractor = FormulaExtractor(workspace_id="custom-workspace")

        # Assert
        assert extractor.workspace_id == "custom-workspace"

    def test_supported_extensions(self):
        """FormulaExtractor supports expected file extensions."""
        # Assert
        assert ".xlsx" in FormulaExtractor.SUPPORTED_EXTENSIONS
        assert ".xls" in FormulaExtractor.SUPPORTED_EXTENSIONS
        assert ".csv" in FormulaExtractor.SUPPORTED_EXTENSIONS
        assert ".ods" in FormulaExtractor.SUPPORTED_EXTENSIONS
        assert ".gsheet" in FormulaExtractor.SUPPORTED_EXTENSIONS
        assert ".numbers" in FormulaExtractor.SUPPORTED_EXTENSIONS


class TestFormulaDetection:
    """Test formula type detection and parsing."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_detect_sum_formula(self, extractor):
        """Detect SUM formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=SUM(A1:A10)")

        # Assert
        assert formula_type == "SUM"

    def test_detect_average_formula(self, extractor):
        """Detect AVERAGE formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=AVERAGE(B1:B20)")

        # Assert
        assert formula_type == "AVERAGE"

    def test_detect_if_formula(self, extractor):
        """Detect IF formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=IF(A1>0, B1, C1)")

        # Assert
        assert formula_type == "IF"

    def test_detect_vlookup_formula(self, extractor):
        """Detect VLOOKUP formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=VLOOKUP(A1, B1:D10, 3, FALSE)")

        # Assert
        assert formula_type == "VLOOKUP"

    def test_detect_arithmetic_formula(self, extractor):
        """Detect arithmetic formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=A1+B1*C1")

        # Assert
        assert formula_type == "ARITHMETIC"

    def test_detect_custom_formula(self, extractor):
        """Detect custom/unknown formula type."""
        # Act
        formula_type = extractor._detect_formula_type("=CUSTOMFUNC(A1)")

        # Assert
        assert formula_type == "CUSTOM"

    def test_case_insensitive_formula_detection(self, extractor):
        """Formula detection is case-insensitive."""
        # Act
        formula_type = extractor._detect_formula_type("=sum(a1:a10)")

        # Assert
        assert formula_type == "SUM"


class TestCellReferenceExtraction:
    """Test cell reference extraction from formulas."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_extract_single_cell_reference(self, extractor):
        """Extract single cell reference from formula."""
        # Act
        refs = extractor._extract_cell_references("=A1")

        # Assert
        assert len(refs) == 1
        assert refs[0] == ("A", 1)

    def test_extract_multiple_cell_references(self, extractor):
        """Extract multiple cell references from formula."""
        # Act
        refs = extractor._extract_cell_references("=SUM(A1, B2, C3)")

        # Assert
        assert len(refs) == 3
        assert ("A", 1) in refs
        assert ("B", 2) in refs
        assert ("C", 3) in refs

    def test_extract_range_references(self, extractor):
        """Extract range references from formula."""
        # Act
        refs = extractor._extract_cell_references("=SUM(A1:A10)")

        # Assert
        assert len(refs) == 2
        assert ("A", 1) in refs
        assert ("A", 10) in refs

    def test_extract_absolute_references(self, extractor):
        """Extract absolute cell references from formula."""
        # Act
        refs = extractor._extract_cell_references("=$A$1+$B$2")

        # Assert
        assert len(refs) == 2
        assert ("A", 1) in refs
        assert ("B", 2) in refs

    def test_column_letter_to_number(self, extractor):
        """Convert column letter to number."""
        # Act & Assert
        assert extractor._column_letter_to_number("A") == 1
        assert extractor._column_letter_to_number("Z") == 26
        assert extractor._column_letter_to_number("AA") == 27
        assert extractor._column_letter_to_number("AZ") == 52
        assert extractor._column_letter_to_number("ZZ") == 702


class TestSemanticExpression:
    """Test semantic expression creation."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_create_semantic_sum_expression(self, extractor):
        """Create semantic expression for SUM formula."""
        # Arrange
        parameters = [
            {"name": "Revenue", "type": "number"},
            {"name": "Expenses", "type": "number"}
        ]

        # Act
        expression = extractor._create_semantic_expression(
            formula_str="=SUM(A1:A2)",
            formula_type="SUM",
            parameters=parameters
        )

        # Assert
        assert expression == "sum(Revenue, Expenses)"

    def test_create_semantic_average_expression(self, extractor):
        """Create semantic expression for AVERAGE formula."""
        # Arrange
        parameters = [
            {"name": "Price", "type": "number"},
            {"name": "Quantity", "type": "number"}
        ]

        # Act
        expression = extractor._create_semantic_expression(
            formula_str="=AVERAGE(A1:A2)",
            formula_type="AVERAGE",
            parameters=parameters
        )

        # Assert
        assert expression == "average(Price, Quantity)"

    def test_create_semantic_arithmetic_expression(self, extractor):
        """Create semantic expression for arithmetic formula."""
        # Arrange
        parameters = [
            {"name": "Price", "type": "number"},
            {"name": "Quantity", "type": "number"}
        ]

        # Act
        expression = extractor._create_semantic_expression(
            formula_str="=A1*B1",
            formula_type="ARITHMETIC",
            parameters=parameters
        )

        # Assert
        assert expression == "Price * Quantity"


class TestDomainDetection:
    """Test domain detection based on column names."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_detect_finance_domain(self, extractor):
        """Detect finance domain from financial keywords."""
        # Act
        domain = extractor._detect_domain("Total Revenue", ["Sales", "Profit"])

        # Assert
        assert domain == "finance"

    def test_detect_sales_domain(self, extractor):
        """Detect sales domain from sales keywords."""
        # Act
        domain = extractor._detect_domain("Conversion Rate", ["Leads", "Deals"])

        # Assert
        assert domain == "sales"

    def test_detect_hr_domain(self, extractor):
        """Detect HR domain from HR keywords."""
        # Act
        domain = extractor._detect_domain("Employee Count", ["Salary", "Headcount"])

        # Assert
        assert domain == "hr"

    def test_detect_general_domain(self, extractor):
        """Detect general domain when no keywords match."""
        # Act
        domain = extractor._detect_domain("Total", ["Column1", "Column2"])

        # Assert
        assert domain == "general"


class TestUseCaseGeneration:
    """Test use case description generation."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_generate_sum_use_case(self, extractor):
        """Generate use case for SUM formula."""
        # Arrange
        parameters = [
            {"name": "Q1 Revenue", "type": "number"},
            {"name": "Q2 Revenue", "type": "number"}
        ]

        # Act
        use_case = extractor._generate_use_case(
            result_name="Total Revenue",
            formula_type="SUM",
            parameters=parameters
        )

        # Assert
        assert "sum" in use_case.lower()
        assert "Total Revenue" in use_case

    def test_generate_average_use_case(self, extractor):
        """Generate use case for AVERAGE formula."""
        # Arrange
        parameters = [
            {"name": "Score1", "type": "number"},
            {"name": "Score2", "type": "number"}
        ]

        # Act
        use_case = extractor._generate_use_case(
            result_name="Average Score",
            formula_type="AVERAGE",
            parameters=parameters
        )

        # Assert
        assert "average" in use_case.lower()
        assert "Average Score" in use_case


class TestExcelExtraction:
    """Test Excel formula extraction."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_extract_from_excel_file_not_found(self, extractor):
        """Return empty list for non-existent file."""
        # Act
        result = extractor.extract_from_excel("/nonexistent/file.xlsx")

        # Assert
        assert result == []

    @patch('core.formula_extractor.openpyxl')
    def test_extract_from_excel_success(self, mock_openpyxl, extractor):
        """Successfully extract formulas from Excel file."""
        # Arrange
        mock_workbook = Mock()
        mock_sheet = Mock()
        mock_cell = Mock()
        mock_cell.value = "=SUM(A1:A10)"
        mock_cell.column = 1
        mock_cell.coordinate = "C1"

        # Mock sheet iteration
        mock_sheet.iter_rows.return_value = [[mock_cell]]
        mock_sheet[1] = []  # Headers row

        mock_workbook.sheetnames = ["Sheet1"]
        mock_workbook.__getitem__ = Mock(return_value=mock_sheet)
        mock_workbook.__iter__ = Mock(return_value=iter([mock_sheet]))

        mock_openpyxl.load_workbook.return_value = mock_workbook

        # Mock formula storage
        extractor._store_formulas = Mock()

        # Act
        result = extractor.extract_from_excel("/fake/test.xlsx")

        # Assert
        assert isinstance(result, list)


class TestCSVExtraction:
    """Test CSV formula extraction."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_extract_from_csv_with_formulas(self, extractor, tmp_path):
        """Extract formulas from CSV file."""
        # Arrange
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Header1,Header2,Header3\n=SUM(A1:A2),10,20\n")

        # Mock formula storage
        extractor._store_formulas = Mock()

        # Act
        result = extractor.extract_from_csv(str(csv_file))

        # Assert
        assert isinstance(result, list)

    def test_extract_from_csv_no_formulas(self, extractor, tmp_path):
        """Return empty list for CSV without formulas."""
        # Arrange
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Header1,Header2\n10,20\n30,40\n")

        # Act
        result = extractor.extract_from_csv(str(csv_file))

        # Assert
        assert result == []

    def test_detect_implicit_formulas(self, extractor):
        """Detect implicit formulas from column patterns."""
        # Arrange
        rows = [
            ["Price", "Quantity", "Total"],
            ["10", "5", "50"],
            ["20", "3", "60"],
            ["15", "4", "60"]
        ]
        headers = {1: "Price", 2: "Quantity", 3: "Total"}

        # Act
        implicit = extractor._detect_implicit_formulas(rows, headers)

        # Assert
        assert isinstance(implicit, list)


class TestOdsExtraction:
    """Test ODS formula extraction."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_extract_from_ods_without_odfpy(self, extractor):
        """Return empty list when odfpy is not installed."""
        # Arrange
        with patch('core.formula_extractor.odf', side_effect=ImportError):
            # Act
            result = extractor.extract_from_ods("/fake/test.ods")

            # Assert
            assert result == []


class TestFileFormatRouting:
    """Test file format routing for formula extraction."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_extract_from_file_routes_to_xlsx(self, extractor):
        """Route .xlsx files to Excel extractor."""
        # Arrange
        extractor.extract_from_excel = Mock(return_value=[])

        # Act
        extractor.extract_from_file("test.xlsx")

        # Assert
        extractor.extract_from_excel.assert_called_once()

    def test_extract_from_file_routes_to_csv(self, extractor):
        """Route .csv files to CSV extractor."""
        # Arrange
        extractor.extract_from_csv = Mock(return_value=[])

        # Act
        extractor.extract_from_file("test.csv")

        # Assert
        extractor.extract_from_csv.assert_called_once()

    def test_extract_from_file_routes_to_ods(self, extractor):
        """Route .ods files to ODS extractor."""
        # Arrange
        extractor.extract_from_ods = Mock(return_value=[])

        # Act
        extractor.extract_from_file("test.ods")

        # Assert
        extractor.extract_from_ods.assert_called_once()

    def test_extract_from_file_unsupported_format(self, extractor):
        """Return empty list for unsupported file format."""
        # Act
        result = extractor.extract_from_file("test.txt")

        # Assert
        assert result == []


class TestFormulaStorage:
    """Test formula storage in memory."""

    @pytest.fixture
    def extractor(self):
        """FormulaExtractor instance."""
        return FormulaExtractor()

    def test_store_formulas_success(self, extractor):
        """Successfully store formulas in memory."""
        # Arrange
        formulas = [
            {
                "expression": "sum(A, B)",
                "name": "Total",
                "domain": "finance",
                "use_case": "Calculate total",
                "parameters": [{"name": "A", "type": "number"}],
                "source": "excel:test.xlsx"
            }
        ]

        # Mock formula manager
        mock_manager = Mock()
        mock_manager.add_formula.return_value = "formula-001"
        extractor._formula_manager = mock_manager

        # Act
        extractor._store_formulas(formulas, "user-001", "/path/to/test.xlsx")

        # Assert
        mock_manager.add_formula.assert_called()


class TestFactory:
    """Test formula extractor factory."""

    def test_get_formula_extractor_default(self):
        """Get formula extractor with default workspace."""
        # Act
        extractor = get_formula_extractor()

        # Assert
        assert extractor is not None
        assert extractor.workspace_id == "default"

    def test_get_formula_extractor_custom_workspace(self):
        """Get formula extractor with custom workspace."""
        # Act
        extractor = get_formula_extractor("custom-workspace")

        # Assert
        assert extractor is not None
        assert extractor.workspace_id == "custom-workspace"

    def test_get_formula_extractor_returns_new_instance(self):
        """Factory returns new instance each time."""
        # Act
        extractor1 = get_formula_extractor()
        extractor2 = get_formula_extractor()

        # Assert
        assert extractor1 is not extractor2
