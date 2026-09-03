"""
Test suite for auto_document_ingestion.py

Document ingestion service for parsing and processing various file formats.
Target file: backend/core/auto_document_ingestion.py (841 lines)
Target tests: 25-30 tests
Coverage target: 25-30%
"""

import io

import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from datetime import datetime
from typing import Dict, Any, List

# Import target module classes
from core.auto_document_ingestion import (
    FileType,
    IntegrationSource,
    IngestionSettings,
    IngestedDocument,
    DocumentParser,
    AutoDocumentIngestionService,
    get_document_ingestion_service,
    AutoDocumentIngestion,
)


class TestFileTypeEnum:
    """Test FileType enum definition."""

    def test_pdf_file_type(self):
        """FileType.PDF has correct value."""
        assert FileType.PDF == "pdf"

    def test_docx_file_type(self):
        """FileType.DOCX has correct value."""
        assert FileType.DOCX == "docx"

    def test_markdown_file_type(self):
        """FileType.MARKDOWN has correct value."""
        assert FileType.MARKDOWN == "md"

    def test_all_file_types_defined(self):
        """All required file types are defined."""
        required_types = ["pdf", "doc", "docx", "txt", "csv", "xlsx", "xls", "md", "json"]
        defined_types = [ft.value for ft in FileType]
        for rt in required_types:
            assert rt in defined_types


class TestIntegrationSourceEnum:
    """Test IntegrationSource enum definition."""

    def test_google_drive_source(self):
        """IntegrationSource.GOOGLE_DRIVE has correct value."""
        assert IntegrationSource.GOOGLE_DRIVE == "google_drive"

    def test_dropbox_source(self):
        """IntegrationSource.DROPBOX has correct value."""
        assert IntegrationSource.DROPBOX == "dropbox"

    def test_notion_source(self):
        """IntegrationSource.NOTION has correct value."""
        assert IntegrationSource.NOTION == "notion"


class TestIngestionSettings:
    """Test IngestionSettings dataclass."""

    def test_ingestion_settings_creation(self):
        """IngestionSettings can be created with valid parameters."""
        settings = IngestionSettings(
            integration_id="google_drive",
            workspace_id="workspace-001",
            enabled=True,
            auto_sync_new_files=True,
            file_types=["pdf", "docx"],
            max_file_size_mb=50
        )
        assert settings.integration_id == "google_drive"
        assert settings.workspace_id == "workspace-001"
        assert settings.enabled is True
        assert settings.auto_sync_new_files is True
        assert settings.file_types == ["pdf", "docx"]
        assert settings.max_file_size_mb == 50

    def test_ingestion_settings_defaults(self):
        """IngestionSettings uses correct default values."""
        settings = IngestionSettings(
            integration_id="dropbox",
            workspace_id="workspace-002"
        )
        assert settings.enabled is False
        assert settings.auto_sync_new_files is True
        assert settings.file_types == ["pdf", "docx", "txt", "md"]
        assert settings.max_file_size_mb == 50
        assert settings.sync_frequency_minutes == 60
        assert settings.last_sync is None


class TestIngestedDocument:
    """Test IngestedDocument dataclass."""

    def test_ingested_document_creation(self):
        """IngestedDocument can be created with valid parameters."""
        doc = IngestedDocument(
            id="doc-001",
            file_name="test.pdf",
            file_path="/files/test.pdf",
            file_type="pdf",
            integration_id="google_drive",
            workspace_id="workspace-001",
            file_size_bytes=1024000,
            content_preview="This is a preview...",
            ingested_at=datetime.utcnow(),
            external_id="ext-001"
        )
        assert doc.id == "doc-001"
        assert doc.file_name == "test.pdf"
        assert doc.file_type == "pdf"
        assert doc.integration_id == "google_drive"
        assert doc.file_size_bytes == 1024000

    def test_ingested_document_optional_fields(self):
        """IngestedDocument handles optional external_modified_at field."""
        doc = IngestedDocument(
            id="doc-002",
            file_name="test2.pdf",
            file_path="/files/test2.pdf",
            file_type="pdf",
            integration_id="dropbox",
            workspace_id="workspace-001",
            file_size_bytes=2048000,
            content_preview="Another preview...",
            ingested_at=datetime.utcnow(),
            external_id="ext-002",
            external_modified_at=None
        )
        assert doc.external_modified_at is None


@pytest.fixture(autouse=True)
def reset_document_ingestion_singleton():
    """Reset the global document ingestion service singleton before each test."""
    from core import auto_document_ingestion as adi_module
    original_instance = adi_module._doc_ingestion_service
    adi_module._doc_ingestion_service = None
    yield
    adi_module._doc_ingestion_service = original_instance


@pytest.fixture(autouse=True)
def isolated_ingestion_db(monkeypatch):
    """Hermetic DB boundary.

    The service's freshness/mirror helpers persist through
    ``core.database.SessionLocal`` — without this swap those writes (PG
    mirror rows, freshness stamps, deletions) land in the real dev database
    (Aug 2026 journey trace: test rows leaked into data/atom.db).
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    import core.database as core_db
    from core.database import Base

    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    monkeypatch.setattr(core_db, "SessionLocal", sessionmaker(bind=engine))
    yield


class TestDocumentParser:
    """Test DocumentParser class."""

    @pytest.mark.asyncio
    async def test_parse_text_file(self):
        """DocumentParser can parse text files."""
        content = b"This is plain text content"
        text = await DocumentParser.parse_document(content, "txt", "test.txt")
        assert "This is plain text content" in text

    @pytest.mark.asyncio
    async def test_parse_markdown_file(self):
        """DocumentParser can parse markdown files."""
        content = b"# Heading\n\nThis is markdown"
        text = await DocumentParser.parse_document(content, "md", "test.md")
        assert "# Heading" in text

    @pytest.mark.asyncio
    async def test_parse_json_file(self):
        """DocumentParser can parse JSON files."""
        content = b'{"key": "value", "number": 123}'
        text = await DocumentParser.parse_document(content, "json", "test.json")
        assert "key" in text
        assert "value" in text

    @pytest.mark.asyncio
    async def test_parse_csv_file(self):
        """DocumentParser can parse CSV files."""
        content = b"Name,Age\nAlice,30\nBob,25"
        text = await DocumentParser.parse_document(content, "csv", "test.csv")
        assert "Name" in text
        assert "Alice" in text

    @pytest.mark.asyncio
    async def test_parse_unsupported_file_type(self):
        """DocumentParser returns empty string for unsupported types."""
        content = b"Some content"
        text = await DocumentParser.parse_document(content, "unsupported", "test.unsupported")
        assert text == ""

    @pytest.mark.asyncio
    async def test_parse_pdf_with_pypdf2_mock(self):
        """DocumentParser can parse PDF using PyPDF2 (imported locally in _parse_pdf)."""
        content = b"Mock PDF content"
        # Mock PyPDF2 module at the point where it's imported (inside _parse_pdf method)
        # Since it's imported as 'import PyPDF2', we mock the module name directly
        with patch('builtins.__import__', side_effect=lambda name, *args, **kwargs: MagicMock() if name == 'PyPDF2' else __import__(name, *args, **kwargs)):
            text = await DocumentParser.parse_document(content, "pdf", "test.pdf")
            # Should return content (either from docling or fallback)
            assert isinstance(text, str)


class TestAutoDocumentIngestionService:
    """Test AutoDocumentIngestionService class."""

    def test_service_initialization(self):
        """AutoDocumentIngestionService initializes correctly."""
        # Mock get_lancedb_handler where it's imported (inside __init__)
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            assert service.workspace_id == "default"
            assert isinstance(service.settings, dict)
            assert service.parser is not None

    def test_get_settings_creates_new(self):
        """get_settings creates new settings if not exists."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            settings = service.get_settings("new_integration")
            assert settings.integration_id == "new_integration"
            assert settings.workspace_id == "default"

    def test_get_settings_returns_existing(self):
        """get_settings returns existing settings."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            settings1 = service.get_settings("existing")
            settings2 = service.get_settings("existing")
            assert settings1 is settings2

    def test_update_settings_enabled(self):
        """update_settings can enable integration."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            settings = service.update_settings("test_integration", enabled=True)
            assert settings.enabled is True

    def test_update_settings_file_types(self):
        """update_settings can update file types."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            settings = service.update_settings(
                "test_integration",
                file_types=["pdf", "txt", "csv"]
            )
            assert settings.file_types == ["pdf", "txt", "csv"]

    def test_update_settings_max_file_size(self):
        """update_settings can update max file size."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            settings = service.update_settings(
                "test_integration",
                max_file_size_mb=100
            )
            assert settings.max_file_size_mb == 100

    @pytest.mark.asyncio
    async def test_sync_integration_disabled(self):
        """sync_integration skips disabled integrations."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            result = await service.sync_integration("disabled_integration")
            assert result.get("skipped") is True
            assert "not enabled" in result.get("reason", "").lower()

    def test_get_ingested_documents_empty(self):
        """get_ingested_documents returns empty list initially."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            docs = service.get_ingested_documents()
            assert docs == []

    def test_get_ingested_documents_by_integration(self):
        """get_ingested_documents can filter by integration."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            # Add a mock document
            doc = IngestedDocument(
                id="doc-001",
                file_name="test.pdf",
                file_path="/files/test.pdf",
                file_type="pdf",
                integration_id="google_drive",
                workspace_id="default",
                file_size_bytes=1000,
                content_preview="Test",
                ingested_at=datetime.utcnow(),
                external_id="ext-001"
            )
            service.ingested_docs["ext-001"] = doc

            docs = service.get_ingested_documents(integration_id="google_drive")
            assert len(docs) == 1
            assert docs[0].integration_id == "google_drive"

    def test_get_all_settings(self):
        """get_all_settings returns all integration settings."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()
            service.get_settings("integration_1")
            service.get_settings("integration_2")

            all_settings = service.get_all_settings()
            assert len(all_settings) == 2

    @pytest.mark.asyncio
    async def test_remove_integration_documents(self):
        """remove_integration_documents removes all docs for integration."""
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service = AutoDocumentIngestionService()

            # Add mock documents
            doc1 = IngestedDocument(
                id="doc-001",
                file_name="test1.pdf",
                file_path="/files/test1.pdf",
                file_type="pdf",
                integration_id="google_drive",
                workspace_id="default",
                file_size_bytes=1000,
                content_preview="Test1",
                ingested_at=datetime.utcnow(),
                external_id="ext-001"
            )
            doc2 = IngestedDocument(
                id="doc-002",
                file_name="test2.pdf",
                file_path="/files/test2.pdf",
                file_type="pdf",
                integration_id="google_drive",
                workspace_id="default",
                file_size_bytes=2000,
                content_preview="Test2",
                ingested_at=datetime.utcnow(),
                external_id="ext-002"
            )
            service.ingested_docs["ext-001"] = doc1
            service.ingested_docs["ext-002"] = doc2

            result = await service.remove_integration_documents("google_drive")
            assert result["success"] is True
            assert result["documents_removed"] == 2
            assert len(service.ingested_docs) == 0


class TestGlobalServiceInstance:
    """Test global service instance functions."""

    def test_get_document_ingestion_service_singleton(self):
        """get_document_ingestion_service returns singleton instance."""
        # Mock get_lancedb_handler where it's imported (inside __init__)
        with patch('core.lancedb_handler.get_lancedb_handler'):
            service1 = get_document_ingestion_service()
            service2 = get_document_ingestion_service()
            assert service1 is service2

    def test_auto_document_ingestion_alias(self):
        """AutoDocumentIngestion is alias for AutoDocumentIngestionService."""
        assert AutoDocumentIngestion == AutoDocumentIngestionService


class TestIntegration:
    """Integration tests for document ingestion workflow."""

    @pytest.mark.asyncio
    async def test_full_ingestion_workflow_with_mocks(self):
        """Test complete ingestion workflow with mocked dependencies."""
        # Mock dependencies where they're imported (inside __init__)
        with patch('core.lancedb_handler.get_lancedb_handler') as mock_lancedb, \
             patch('core.secrets_redactor.get_secrets_redactor') as mock_redactor:

            # Setup mocks
            mock_memory = AsyncMock()
            mock_memory.add_document.return_value = True
            mock_lancedb.return_value = mock_memory

            mock_redactor_instance = MagicMock()
            mock_redactor_instance.redact.return_value = MagicMock(
                has_secrets=False,
                redacted_text="Safe content"
            )
            mock_redactor.return_value = mock_redactor_instance

            service = AutoDocumentIngestionService()

            # Enable integration
            service.update_settings("test_integration", enabled=True)

            # Mock file listing
            with patch.object(service, '_list_files', return_value=[
                {
                    "id": "file-001",
                    "name": "test.pdf",
                    "path": "/files/test.pdf",
                    "size": 1024,
                    "modified_at": datetime.utcnow()
                }
            ]):
                # Mock file download
                with patch.object(service, '_download_file', return_value=b"PDF content"):
                    result = await service.sync_integration("test_integration", force=True)

                    # Verify sync ran
                    assert "integration_id" in result
                    assert result["integration_id"] == "test_integration"


def _build_minimal_xlsx(
    shared_strings_xml: str | None = None,
    sheet_xml: str | None = None,
) -> bytes:
    """Hand-assemble a small xlsx package (mirrors third-party writers whose
    sharedStrings.xml openpyxl's strict reader rejects, e.g. Zoho Sheet)."""
    import zipfile

    ct = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        + (
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            if shared_strings_xml
            else ""
        )
        + "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    wb = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Price List" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    wb_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        + (
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
            if shared_strings_xml
            else ""
        )
        + "</Relationships>"
    )
    sst = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="4" uniqueCount="4">'
        # Zoho-style defect: phonetic run with non-numeric sb/eb — well-formed
        # XML that openpyxl's Integer descriptor rejects inside read_strings.
        "<si><rPh sb=\"x\" eb=\"y\"><t>ph</t></rPh><t>Consolidated Price List 2019</t></si>"
        "<si><t>HDR-1000</t></si>"
        "<si><t>Hydraulic Press</t></si>"
        "<si><t>1</t></si>"
        "</sst>"
    ) if shared_strings_xml == "defective" else shared_strings_xml
    sheet = sheet_xml or (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
        '<row r="1"><c r="A1" t="s"><v>0</v></c></row>'
        '<row r="2"><c r="A2" t="s"><v>1</v></c><c r="B2" t="s"><v>2</v></c><c r="C2" t="s"><v>3</v></c></row>'
        "</sheetData></worksheet>"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ct)
        z.writestr("_rels/.rels", rels)
        z.writestr("xl/workbook.xml", wb)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        if shared_strings_xml:
            z.writestr("xl/sharedStrings.xml", sst)
        z.writestr("xl/worksheets/sheet1.xml", sheet)
    return buf.getvalue()


class TestExcelMalformedWorkbookFallback:
    """Real-world regression (2026-09-03): 'Consolidated Price List 2019.xlsx'
    from Zoho WorkDrive came back no_text — openpyxl's read_strings raised on
    its sharedStrings.xml, _parse_excel swallowed the error and returned "",
    so the file never became searchable. The raw-XML fallback must extract it."""

    @pytest.mark.asyncio
    async def test_shared_strings_openpyxl_rejects_still_extracts(self):
        """Workbook whose sharedStrings openpyxl chokes on extracts via fallback."""
        content = _build_minimal_xlsx(shared_strings_xml="defective")
        text = await DocumentParser.parse_document(content, "xlsx", "Consolidated Price List 2019.xlsx")
        assert "Consolidated Price List 2019" in text
        assert "HDR-1000" in text
        assert "Hydraulic Press" in text
        # Phonetic hint <rPh> is pronunciation metadata, not content.
        assert "phConsolidated" not in text
        # Real sheet name from workbook.xml, not the part filename.
        assert "=== Sheet: Price List ===" in text
        assert "sheet1.xml" not in text

    @pytest.mark.asyncio
    async def test_truly_truncated_xml_is_clean_skip(self):
        """Genuinely malformed (truncated) XML: no crash, no junk — empty skip."""
        truncated = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<si><t>HDR-1000</t></si>"
        )
        content = _build_minimal_xlsx(shared_strings_xml=truncated)
        text = await DocumentParser.parse_document(content, "xlsx", "broken.xlsx")
        assert text == ""

    @pytest.mark.asyncio
    async def test_inline_strings_without_shared_strings_part(self):
        """Inline-string workbooks (no sharedStrings.xml) extract cell text."""
        sheet = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
            '<row r="1"><c r="A1" t="inlineStr"><is><t>Inline Only</t></is></c><c r="B1"><v>7</v></c></row>'
            "</sheetData></worksheet>"
        )
        content = _build_minimal_xlsx(sheet_xml=sheet)
        text = await DocumentParser.parse_document(content, "xlsx", "inline.xlsx")
        assert "Inline Only" in text
        assert "7" in text

    @pytest.mark.asyncio
    async def test_formula_cells_render_value_and_formula(self):
        """Formula cells must not collapse to their cached value: <v> is the
        last computed number, <f> is the business logic (markup, currency
        conversion). The fallback renders both."""
        sheet = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
            '<row r="1"><c r="A1" t="s"><v>0</v></c></row>'
            '<row r="2"><c r="A2"><v>42</v></c><c r="B2"><v>1.25</v></c>'
            '<c r="C2"><f>A2*B2</f><v>52.5</v></c></row>'
            "</sheetData></worksheet>"
        )
        content = _build_minimal_xlsx(sheet_xml=sheet)
        text = await DocumentParser.parse_document(content, "xlsx", "priced.xlsx")
        assert "52.5 [=A2*B2]" in text

    @pytest.mark.asyncio
    async def test_shared_formula_without_text_keeps_value(self):
        """Dependent cells of a shared formula carry <f t="shared" si="…"/>
        with no body — keep the value, do not invent a formula."""
        sheet = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
            '<row r="1"><c r="A1"><f t="shared" ref="A1:A2" si="0">B1*2</f><v>10</v></c></row>'
            '<row r="2"><c r="A2"><f t="shared" si="0"/><v>20</v></c></row>'
            "</sheetData></worksheet>"
        )
        content = _build_minimal_xlsx(sheet_xml=sheet)
        text = await DocumentParser.parse_document(content, "xlsx", "shared.xlsx")
        assert "20" in text
        assert "[=" not in text.split("\n")[2] or "[=B1*2]" in text  # no phantom formula on the dependent cell

    @pytest.mark.asyncio
    async def test_bytes_ingest_runs_formula_extractor(self):
        """Connector ingests are bytes-only; the Phase-19 formula extractor
        must run for them too (it previously required a disk path, so formula
        memory never saw cloud-drive workbooks)."""
        content = _build_minimal_xlsx(shared_strings_xml="defective")
        with patch("core.formula_extractor.FormulaExtractor.extract_from_excel") as ex:
            ex.return_value = []
            text = await DocumentParser.parse_document(content, "xlsx", "f.xlsx")
        assert ex.called

    @pytest.mark.asyncio
    async def test_normal_workbook_still_uses_pandas_path(self):
        """Regression guard: a well-formed workbook parses unchanged."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Normal File"
        ws["B2"] = 123
        buf = io.BytesIO()
        wb.save(buf)
        text = await DocumentParser.parse_document(buf.getvalue(), "xlsx", "normal.xlsx")
        assert "Normal File" in text

    @pytest.mark.asyncio
    async def test_non_zip_bytes_labeled_xlsx_is_clean_skip(self):
        """OLE2 bytes mislabeled .xlsx: no crash, empty result (no junk text)."""
        content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
        text = await DocumentParser.parse_document(content, "xlsx", "fake.xlsx")
        assert text == ""


class TestSourceUpdateReingest:
    """File UPDATES must be ingested and accounted for (2026-09-03): storage
    drives default to hybrid content mode, and the mode gate silently skipped
    changed files — a stored copy kept reading "fresh" while the source moved
    on. Now: a newer source modified_at on an already-stored file refreshes
    content even under hybrid mode (upsert replaces the old row)."""

    PROBE_ID = "ext_" + __import__("hashlib").sha1(
        b"zoho_workdrive:file-123"
    ).hexdigest()[:24]

    def _svc(self, stored):
        with patch("core.lancedb_handler.get_lancedb_handler"):
            svc = AutoDocumentIngestionService()
        svc.memory_handler.get_document_by_id = MagicMock(return_value=stored)
        return svc

    async def _run(self, svc, source_modified_at):
        from core.vector_upsert import upsert_document as _unused  # noqa: F401

        with patch(
            "core.hybrid_data_ingestion.get_hybrid_ingestion_service"
        ) as gsvc, patch(
            "core.vector_upsert.upsert_document",
            new=AsyncMock(return_value="written"),
        ) as up:
            gsvc.return_value.get_content_mode = MagicMock(return_value="hybrid")
            result = await svc.process_file_bytes(
                content=b"hello v2",
                file_name="quote.txt",
                source="zoho_workdrive",
                user_id="u-1",
                external_id="file-123",
                extra_metadata={"source_modified_at": source_modified_at},
                explicit=False,
            )
            return result, up

    @pytest.mark.asyncio
    async def test_updated_file_refreshes_under_hybrid_mode(self):
        stored = {
            "id": self.PROBE_ID,
            "metadata": {"source_modified_at": "2025-05-01T12:16:00+00:00"},
            "text": "old version",
        }
        svc = self._svc(stored)
        result, up = await self._run(svc, "Sep 2, 2026, 09:00 AM")
        assert result["status"] == "ingested", result
        assert up.call_args.kwargs["extra_columns"]["source_modified_at"].startswith(
            "2026-09-02"
        )

    @pytest.mark.asyncio
    async def test_unchanged_file_still_gated(self):
        stored = {
            "id": self.PROBE_ID,
            "metadata": {"source_modified_at": "2025-05-01T12:16:00+00:00"},
            "text": "old version",
        }
        svc = self._svc(stored)
        result, up = await self._run(svc, "May 1, 2025, 12:16 PM")
        assert result["status"] == "skipped"
        assert result["reason"] == "content_mode_hybrid"
        up.assert_not_called()

    @pytest.mark.asyncio
    async def test_new_file_still_gated_under_hybrid(self):
        svc = self._svc(None)
        result, up = await self._run(svc, "Sep 2, 2026, 09:00 AM")
        assert result["status"] == "skipped"
        assert result["reason"] == "content_mode_hybrid"
        up.assert_not_called()

    @pytest.mark.asyncio
    async def test_unparsable_modified_time_still_gated(self):
        # 'can't tell' must not trigger a refresh
        stored = {
            "id": self.PROBE_ID,
            "metadata": {"source_modified_at": "2025-05-01T12:16:00+00:00"},
            "text": "old version",
        }
        svc = self._svc(stored)
        result, up = await self._run(svc, "recently-ish")
        assert result["status"] == "skipped"
        up.assert_not_called()


class TestParseSourceModified:
    def test_workdrive_listing_format(self):
        from core.auto_document_ingestion import _parse_source_modified

        dt = _parse_source_modified("May 1, 2025, 12:16 PM")
        assert dt is not None and dt.tzinfo is not None
        assert dt.month == 5 and dt.hour == 12

    def test_iso_formats_and_datetime_passthrough(self):
        from datetime import datetime as dtc, timezone as tz
        from core.auto_document_ingestion import _parse_source_modified

        assert _parse_source_modified("2025-05-01T12:16:00+00:00") is not None
        assert _parse_source_modified("2025-05-01") is not None
        now = dtc.now(tz.utc)
        assert _parse_source_modified(now) is now
        assert _parse_source_modified("recently-ish") is None
        assert _parse_source_modified(None) is None
        assert _parse_source_modified("") is None


@pytest.mark.asyncio
async def test_long_file_ingests_as_chunk_family_and_refresh_replaces_it():
    """Integration on a real (isolated) workspace store: a long file lands as
    {doc_id}::c{i} rows; a source-side update refreshes the WHOLE family."""
    import shutil
    from pathlib import Path

    ws = "ws-chunk-ingest-test"
    store = Path(__file__).resolve().parent.parent / "data" / "atom_memory" / ws
    shutil.rmtree(store, ignore_errors=True)
    svc = AutoDocumentIngestionService(workspace_id=ws)
    text_v1 = ("Section 0: " + "pricing detail " * 40) + "".join(
        f"\n\nSection {i}: " + "pricing detail " * 40 for i in range(1, 12)
    )
    common = dict(
        source="zoho_workdrive", user_id="u-chunk",
        workspace_id=ws, external_id="file-chunk-1",
    )
    try:
        r1 = await svc.process_file_bytes(
            content=text_v1.encode(), file_name="big.txt", **common,
            extra_metadata={"source_modified_at": "Sep 1, 2026, 08:00 AM"},
            explicit=True,
        )
        assert r1["status"] == "ingested"
        doc_id = r1["doc_id"]
        family1 = sorted(
            svc.memory_handler.get_document_ids_by_prefix("documents", f"{doc_id}::c")
        )
        assert len(family1) > 3, "long document must chunk"

        # walker-shaped auto ingest with a newer source time → full refresh
        v2 = text_v1.replace("pricing detail", "updated pricing")
        r2 = await svc.process_file_bytes(
            content=v2.encode(), file_name="big.txt", **common,
            extra_metadata={"source_modified_at": "Sep 3, 2026, 10:00 AM"},
            explicit=False,
        )
        assert r2["status"] == "ingested", r2
        family2 = sorted(
            svc.memory_handler.get_document_ids_by_prefix("documents", f"{doc_id}::c")
        )
        assert family2 == family1, "chunk ids stable, content replaced"
        stored = svc.memory_handler.get_document_by_id("documents", f"{doc_id}::c0")
        assert "updated pricing" in stored["text"]
    finally:
        shutil.rmtree(store, ignore_errors=True)
