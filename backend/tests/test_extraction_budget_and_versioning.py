"""Full-text extraction budget + extraction-version hash dedup.

Regression root: Consolidated Price List 2019.xlsx ingested through the old
5-sheet/100-row caps stored ~386 chunks WITHOUT its machine-pricing sheets —
the WG350DSAV row existed nowhere in the corpus, the agent confabulated the
price, and every re-ingest short-circuited "unchanged" because the truncated
extraction hashed identically. These tests pin:
  1. all sheets/rows/pages/paragraphs extract (only the char budget bounds);
  2. the budget truncates with a visible note when set low;
  3. extraction_content_hash changes when EXTRACTION_VERSION changes, so
     hash-dedup can never pin a stale extraction;
  4. upsert_document_chunks replaces a row whose stored hash was produced by
     an older extractor (write, not skip_unchanged).
"""

import io
import os
import uuid
from unittest import mock

import pytest


def _make_workbook(n_sheets: int, n_rows: int, marker_prefix: str = "MARKER") -> bytes:
    """Real multi-sheet xlsx built in-memory (openpyxl), one marker row per
    sheet near the END of that sheet (past the old 100-row cap)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in range(n_sheets):
        ws = wb.create_sheet(title=f"Sheet{s + 1}")
        ws.append(["col_a", "col_b"])
        for r in range(n_rows - 1):
            ws.append([f"filler-{s + 1}-{r}", r])
        ws.append([f"{marker_prefix}{s + 1}END", "value"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_xlsx_all_sheets_and_rows_extract():
    """8 sheets x 150 rows: every sheet's end-marker must survive extraction
    (old code: only sheets 1-5 x first 100 rows). Calls the Excel parser
    directly so the test is independent of whether docling is installed."""
    from core.auto_document_ingestion import DocumentParser

    content = _make_workbook(8, 150)
    text = await DocumentParser._parse_excel(content, "big.xlsx")
    for s in range(1, 9):
        assert f"MARKER{s}END" in text, f"sheet {s} content missing from extraction"


@pytest.mark.asyncio
async def test_xlsx_structured_context_preserved():
    """Spreadsheet structure (sheet -> column -> cell) must survive: a
    workbook-level index, per-sheet column-letter header maps, real row
    numbers, and per-cell formulas. Live 2026-09-03: Consolidated Price List
    2019.xlsx lost all of this — the agent could not cite 'LINMAC R17'."""
    from core.auto_document_ingestion import DocumentParser
    from openpyxl import Workbook
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "LINMAC"
    ws.append(["MODEL", "DESCRIPTION", "LIST PRICE"])  # sheet row 1 = headers
    ws.append(["WG120", "Small bandsaw", 5200])
    ws.append(["WG350DSAV", "Bandsaw 230V/3PH/60HZ Double Miter", 14145])  # row 3
    ws["D1"] = "NET"       # extra column so D can hold a formula cell
    ws["D3"] = "=C3*0.9"   # formula cell (openpyxl stores no cached value)
    buf = _io.BytesIO()
    wb.save(buf)

    text = await DocumentParser._parse_excel(buf.getvalue(), "linmac.xlsx")
    assert "WORKBOOK INDEX:" in text and "LINMAC" in text
    assert "A=MODEL" in text and "C=LIST PRICE" in text  # column letter map
    assert "R3" in text and "WG350DSAV" in text          # real row numbers
    assert "14145" in text

    # raw-XML path (what Zoho exports fall back to): formulas preserved
    raw = DocumentParser._parse_xlsx_raw(buf.getvalue())
    assert "WG350DSAV" in raw and "R3" in raw
    assert "[=" in raw, "raw-XML path must keep the formula text"

    # workbook index carries the header summary for discovery
    assert "MODEL" in text.split("=== Sheet")[0] or "MODEL" in text.split("--- Sheet")[0]


@pytest.mark.asyncio
async def test_xlsx_budget_truncates_with_note(monkeypatch):
    """Above the budget floor, the text is bounded and the truncation note
    renders — later sheets are cut, earlier content survives."""
    from core.auto_document_ingestion import DocumentParser

    monkeypatch.setenv("ATOM_EXTRACTION_MAX_CHARS", "12000")
    content = _make_workbook(10, 250)  # ~6k chars/sheet — sheet 3+ cross the floor
    text = await DocumentParser._parse_excel(content, "bigger.xlsx")
    assert "extraction budget reached" in text
    assert "MARKER1END" in text, "early content must survive the cut"
    assert "MARKER10END" not in text, "budget did not bound the extraction"


@pytest.mark.asyncio
async def test_csv_beyond_1000_rows_extracts():
    """Old cap: 1000 rows. A 1500-row CSV must extract past row 1000."""
    from core.auto_document_ingestion import DocumentParser

    lines = ["name,amount"] + [f"row{i},{i}" for i in range(1, 1501)]
    text = await DocumentParser.parse_document(
        "\n".join(lines).encode(), "csv", "long.csv"
    )
    assert "row1500 | 1500" in text


@pytest.mark.asyncio
async def test_pdf_beyond_50_pages_extracts(monkeypatch):
    """Old cap: 50 pages. A 55-page PDF must extract page 55's body."""
    import sys
    import types
    from core.auto_document_ingestion import DocumentParser

    class FakePage:
        def __init__(self, i):
            self.i = i

        def extract_text(self):
            return f"PAGE{self.i}BODY"

    fake_pypdf = types.ModuleType("pypdf")

    class FakeReader:
        def __init__(self, _buf):
            self.pages = [FakePage(i) for i in range(55)]

    fake_pypdf.PdfReader = FakeReader
    fake_module = monkeypatch.setitem(sys.modules, "pypdf", fake_pypdf)
    text = await DocumentParser._parse_pdf(b"%pdf-fake-bytes")
    assert "PAGE54BODY" in text, "page 55 content missing — page cap still active"


def test_extraction_hash_changes_with_version():
    """Bumping EXTRACTION_VERSION must change the content hash for identical
    text — that is the mechanism that un-pins a stale extraction."""
    from core.doc_freshness_service import (
        EXTRACTION_VERSION,
        extraction_content_hash,
        has_current_extraction_version,
        hash_text,
    )

    text = "WG350DSAV\tBandsaw\t14145.00"
    h = extraction_content_hash(text)
    assert h.startswith(f"ev{EXTRACTION_VERSION}:")
    assert has_current_extraction_version(h)
    assert not has_current_extraction_version("ev1:" + hash_text(text))
    assert not has_current_extraction_version(hash_text(text))  # legacy bare digest
    assert not has_current_extraction_version(None)

    with mock.patch("core.doc_freshness_service.EXTRACTION_VERSION", "X"):
        h2 = extraction_content_hash(text)
    assert h2 != h
    assert h2.startswith("evX:")


@pytest.mark.asyncio
async def test_upsert_replaces_stale_extractor_row():
    """A stored row hashed by an OLD extractor must be REPLACED (write), not
    skipped as unchanged, when the current extractor produces new text."""
    from core.doc_freshness_service import hash_text
    from core.vector_upsert import upsert_document_chunks

    class FakeTable:
        def __init__(self, rows):
            self._rows = rows

    # Minimal fake handler exercising only the probe/delete/add contract used
    # by upsert_document_chunks.
    stored_meta = {"source_content_hash": "ev1:deadbeef"}  # OLD extractor hash
    new_text = "WG350DSAV row with price 14145.00"

    class FakeHandler:
        def __init__(self):
            self.deleted = []
            self.added = []

        def get_document_by_id(self, table_name, doc_id):
            if stored_meta is None:
                return None
            return {"metadata": stored_meta}

        def delete_documents_by_id(self, table_name, doc_id):
            self.deleted.append(doc_id)

        def add_document(self, **kwargs):
            self.added.append(kwargs)
            return True

    handler = FakeHandler()
    status = await upsert_document_chunks(
        handler,
        table_name="documents",
        text=new_text,
        doc_id="ext_test",
        source="zoho_workdrive:big.xlsx",
        metadata={},
        user_id="u1",
    )
    assert status == "written", "stale-extractor hash must not read as unchanged"
    assert handler.deleted == ["ext_test"]
    assert handler.added and handler.added[0]["metadata"]["source_content_hash"].startswith("ev")


@pytest.mark.asyncio
async def test_source_side_update_refreshes_in_hybrid_mode():
    """UPDATE PATH (live 2026-09-04): a source-side edit of an already-stored
    file must re-ingest in hybrid mode. WorkDrive's bulk walker passes the
    modified time under the key 'modified_at'; the funnel only read
    'source_modified_at', so the stored copy never had a baseline time and
    every changed file was silently skipped as content_mode_hybrid."""
    import shutil
    from pathlib import Path

    ws = "ws-update-detect-test"
    store = Path(__file__).resolve().parent.parent / "data" / "atom_memory" / ws
    shutil.rmtree(store, ignore_errors=True)
    from core.auto_document_ingestion import AutoDocumentIngestionService

    svc = AutoDocumentIngestionService(workspace_id=ws)
    run_tag = uuid.uuid4().hex[:8]  # unique content: a cached handler over a
    # re-created workspace must not see the prior run's identical rows
    common = dict(
        source="zoho_workdrive", user_id="u-upd",
        workspace_id=ws, external_id="file-upd-1",
    )
    try:
        r1 = await svc.process_file_bytes(
            content=f"price list v1 {run_tag}: WG350DSAV 14145".encode(), file_name="prices.txt",
            **common,
            extra_metadata={"modified_at": "Sep 1, 2026, 08:00 AM"},  # WorkDrive key
            explicit=True,
        )
        assert r1["status"] == "ingested", r1

        # source-side edit,walker-shaped auto ingest (explicit=False)
        r2 = await svc.process_file_bytes(
            content=f"price list v2 {run_tag}: WG350DSAV 15145".encode(), file_name="prices.txt",
            **common,
            extra_metadata={"modified_at": "Sep 3, 2026, 10:00 AM"},
            explicit=False,
        )
        assert r2["status"] == "ingested", r2  # was: skipped content_mode_hybrid
        stored = (
            svc.memory_handler.get_document_by_id("documents", r1["doc_id"])
            or svc.memory_handler.get_document_by_id(
                "documents", f"{r1['doc_id']}::c0")
        )
        assert stored and "15145" in stored["text"] and run_tag in stored["text"]
        # the update-detection baseline is now stamped for FUTURE probes
        meta = stored.get("metadata") or {}
        assert meta.get("source_modified_at"), "modified time must be stamped"
    finally:
        shutil.rmtree(store, ignore_errors=True)
